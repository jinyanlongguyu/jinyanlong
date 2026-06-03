#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
閲戣壋榫橝I绋庡姟鍔╂墜 - Web 鐣岄潰锛堟敮鎸佺湡瀹?DeepSeek AI锛?
杩愯锛歴treamlit run tax_web_app.py

API Key 閰嶇疆浼樺厛绾э細
  1. 渚ц竟鏍忔墜鍔ㄨ緭鍏ワ紙鏈€楂橈級
  2. .streamlit/secrets.toml 鎴?Streamlit Cloud secrets
  3. .env 鏂囦欢涓殑 DEEPSEEK_API_KEY
  4. 鏈厤缃垯浣跨敤妯℃嫙妯″紡
"""

import sys
import os
import json
import requests
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import streamlit as st
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv

# 鍚姩鏃跺姞杞?.env 鏂囦欢
load_dotenv()

# 瀵煎叆璁＄畻鍙傛暟
from tax_calculator import (
    SOCIAL_INSURANCE_ACTUAL,
    SOCIAL_INSURANCE_COMPANY,
    BASIC_DEDUCTION,
    calc_corporate_income_tax_quarterly,
    format_corporate_tax_report,
    classify_bank_transaction,
    generate_profit_statement,
    validate_quarterly_declaration,
    calc_vat_and_surcharge,
    get_tax_policy_summary,
    calc_disabled_employment_fund,
    calc_stamp_duty,
    validate_salary_data,
)

# ===============================================
#  PDF 鐢熸垚宸ュ叿
# ===============================================

try:
    from fpdf import FPDF
    _HAS_FPDF = True
except ImportError:
    _HAS_FPDF = False


def make_pdf(title: str, body_lines: list, filename: str) -> bytes | None:
    """鐢熸垚 PDF 鏂囦欢瀛楄妭娴侊紝fpdf2 涓嶅彲鐢ㄦ椂杩斿洖 None"""
    if not _HAS_FPDF:
        return None
    try:
        pdf = FPDF()
        pdf.add_page()
        # 浣跨敤鍐呯疆瀛椾綋锛堟棤闇€棰濆瀛椾綋鏂囦欢锛?
        pdf.set_auto_page_break(auto=True, margin=15)
        # 鏍囬
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, title, ln=True, align="C")
        pdf.ln(6)
        # 姝ｆ枃
        pdf.set_font("Helvetica", "", 9)
        for line in body_lines:
            # 澶勭悊涓枃锛氱敤 ASCII 鏇夸唬鏂规
            safe_line = line.encode("ascii", errors="replace").decode("ascii")
            pdf.multi_cell(0, 5, safe_line)
        return bytes(pdf.output())
    except Exception:
        return None


def make_pdf_with_dataframe(title: str, df, summary_lines: list, filename: str) -> bytes | None:
    """鐢熸垚鍚〃鏍肩殑 PDF"""
    if not _HAS_FPDF:
        return None
    try:
        pdf = FPDF()
        pdf.add_page("L")  # 妯悜
        pdf.set_auto_page_break(auto=True, margin=10)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, title, ln=True, align="C")
        pdf.ln(4)

        # 琛ㄥご
        cols = list(df.columns)
        col_w = (pdf.w - 20) / len(cols)
        pdf.set_font("Helvetica", "B", 7)
        for c in cols:
            safe = str(c).encode("ascii", errors="replace").decode("ascii")
            pdf.cell(col_w, 6, safe[:12], border=1, align="C")
        pdf.ln()

        # 鏁版嵁琛?
        pdf.set_font("Helvetica", "", 7)
        for _, row in df.head(30).iterrows():
            for c in cols:
                val = str(row[c]).encode("ascii", errors="replace").decode("ascii")
                pdf.cell(col_w, 5, val[:15], border=1, align="C")
            pdf.ln()

        pdf.ln(6)
        pdf.set_font("Helvetica", "", 9)
        for line in summary_lines:
            safe = line.encode("ascii", errors="replace").decode("ascii")
            pdf.multi_cell(0, 5, safe)

        return bytes(pdf.output())
    except Exception:
        return None

# ===============================================
#  DeepSeek AI 閰嶇疆
# ===============================================

DEEPSEEK_API_URL = "https://api.deepseek.com/chat/completions"
DEEPSEEK_MODEL = "deepseek-chat"

# 浠?.env 璇诲彇榛樿 Key
ENV_API_KEY = os.environ.get("DEEPSEEK_API_KEY", "")


def get_api_key():
    """鑾峰彇褰撳墠鐢熸晥鐨?API Key锛堟墜鍔ㄨ緭鍏?> st.secrets > .env锛?""
    # 1. 渚ц竟鏍忔墜鍔ㄨ緭鍏ワ紙鏈€楂樹紭鍏堢骇锛?
    manual_key = st.session_state.get("deepseek_api_key_manual", "")
    if manual_key:
        return manual_key, "manual"
    # 2. Streamlit Cloud secrets
    try:
        if "DEEPSEEK_API_KEY" in st.secrets:
            return st.secrets["DEEPSEEK_API_KEY"], "secrets"
    except Exception:
        pass
    # 3. .env 鏂囦欢
    if ENV_API_KEY:
        return ENV_API_KEY, "env"
    return "", "none"


def ask_deepseek(prompt: str, system_prompt: str = None) -> str:
    """璋冪敤 DeepSeek API"""
    api_key, _ = get_api_key()
    if not api_key:
        return "[鏈厤缃?API Key锛岃烦杩?AI 鐢熸垚]"

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    messages = []
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt})
    messages.append({"role": "user", "content": prompt})

    payload = {
        "model": DEEPSEEK_MODEL,
        "messages": messages,
        "temperature": 0.3,
        "max_tokens": 1500,
    }

    try:
        resp = requests.post(
            DEEPSEEK_API_URL, headers=headers, json=payload, timeout=30
        )
        resp.raise_for_status()
        return resp.json()["choices"][0]["message"]["content"]
    except Exception as e:
        return f"[AI 璋冪敤澶辫触: {e}]"


# ===============================================
#  鏍稿績璁＄畻鍑芥暟
# ===============================================

def calc_one_employee(
    name, gross_salary, si_base, si_personal,
    special_total, child_edu, infant_care, elderly_care,
    housing_fund_personal=0.0,
) -> dict:
    """璁＄畻鍗曞悕鍛樺伐"""
    taxable_income = (
        gross_salary
        - si_personal
        - housing_fund_personal
        - BASIC_DEDUCTION
        - special_total
    )

    if taxable_income <= 0:
        tax = 0.0
        taxable_income = 0.0
    else:
        if taxable_income <= 3000:
            rate, deduction = 0.03, 0
        elif taxable_income <= 12000:
            rate, deduction = 0.10, 210
        elif taxable_income <= 25000:
            rate, deduction = 0.20, 1410
        elif taxable_income <= 35000:
            rate, deduction = 0.25, 2660
        elif taxable_income <= 55000:
            rate, deduction = 0.30, 4410
        elif taxable_income <= 80000:
            rate, deduction = 0.35, 7160
        else:
            rate, deduction = 0.45, 15160

        tax = taxable_income * rate - deduction

    net_salary = gross_salary - si_personal - housing_fund_personal - round(tax, 2)

    si_company = (
        si_base * SOCIAL_INSURANCE_COMPANY["pension"]
        + si_base * SOCIAL_INSURANCE_COMPANY["medical"]
        + si_base * SOCIAL_INSURANCE_COMPANY["unemployment"]
        + si_base * SOCIAL_INSURANCE_COMPANY["injury"]
    )
    total_cost = gross_salary + si_company

    return {
        "濮撳悕": name,
        "绋庡墠宸ヨ祫": gross_salary,
        "涓汉绀句繚": si_personal,
        "涓撻」闄勫姞鎵ｉ櫎": special_total,
        "瀛愬コ鏁欒偛": child_edu,
        "濠村辜鍎跨収鎶?: infant_care,
        "璧″吇鑰佷汉": elderly_care,
        "搴旂◣鏀跺叆": round(taxable_income, 2),
        "搴旂撼绋庨": round(tax, 2),
        "瀹炲彂宸ヨ祫": round(net_salary, 2),
        "鍏徃绀句繚鎵挎媴": round(si_company, 2),
        "鍏徃鐢ㄤ汉鎬绘垚鏈?: round(total_cost, 2),
    }


# ===============================================
#  AI 鐢虫姤璇存槑鐢熸垚
# ===============================================

def generate_tax_report_ai(results: list) -> str:
    """鐢熸垚涓◣鐢虫姤璇存槑锛堢湡瀹?AI 鎴栨ā鎷燂級"""
    now = datetime.now()
    api_key, _ = get_api_key()
    use_ai = bool(api_key)

    if use_ai:
        rows_text = ""
        for r in results:
            rows_text += (
                f"鍛樺伐 {r['濮撳悕']}锛氱◣鍓嶅伐璧?{r['绋庡墠宸ヨ祫']} 鍏冿紝"
                f"涓汉绀句繚 {r['涓汉绀句繚']} 鍏冿紝"
                f"涓撻」闄勫姞鎵ｉ櫎 {r['涓撻」闄勫姞鎵ｉ櫎']} 鍏?
                f"锛堝瓙濂虫暀鑲?{r['瀛愬コ鏁欒偛']} 鍏冿紝濠村辜鍎跨収鎶?{r['濠村辜鍎跨収鎶?]} 鍏冿紝"
                f"璧″吇鑰佷汉 {r['璧″吇鑰佷汉']} 鍏冿級锛?
                f"搴旂◣鏀跺叆 {r['搴旂◣鏀跺叆']} 鍏冿紝搴旂撼绋庨 {r['搴旂撼绋庨']} 鍏冿紝"
                f"瀹炲彂宸ヨ祫 {r['瀹炲彂宸ヨ祫']} 鍏冦€俓n"
            )

        company_si_total = sum(r["鍏徃绀句繚鎵挎媴"] for r in results)
        total_tax = sum(r["搴旂撼绋庨"] for r in results)
        total_cost = sum(r["鍏徃鐢ㄤ汉鎬绘垚鏈?] for r in results)

        prompt = f"""浣犳槸涓€浣嶄笓涓氱殑绋庡姟椤鹃棶锛岃涓轰互涓嬩紒涓歿now.year}骞磠now.month}鏈堢殑涓◣鍙婄ぞ淇濈敵鎶ユ挵鍐欎竴浠戒笓涓氱殑鐢虫姤璇存槑銆?

## 鍛樺伐鏁版嵁
{rows_text}
## 姹囨€绘暟鎹?
- 鍏徃鎵挎媴绀句繚鎬婚锛歿company_si_total} 鍏?
- 鍏ㄤ綋鍛樺伐搴旂撼绋庨鍚堣锛歿total_tax} 鍏?
- 鍏徃鐢ㄤ汉鎬绘垚鏈細{total_cost} 鍏?

## 瑕佹眰
1. 浠?姝︽眽閲戣壋榫欑鎶€鏈夐檺鍏徃 {now.year}骞磠now.month}鏈?绋庡姟鐢虫姤璇存槑"涓烘爣棰?
2. 鍒嗗洓涓儴鍒嗭細涓€銆佺敵鎶ユ鍐碉紱浜屻€佸憳宸ヤ釜绋庢槑缁嗭紱涓夈€佺ぞ淇濈即绾宠鏄庯紱鍥涖€佺敵鎶ユ敞鎰忎簨椤?
3. 璇皵涓撲笟銆佺畝娲侊紝閫傚悎璐㈠姟鎻愪氦缁欑◣鍔″眬鎴栫暀瀛樺妗?
4. 鎻愰啋鐢ㄦ埛鏍稿涓撻」闄勫姞鎵ｉ櫎淇℃伅鏄惁宸插強鏃舵洿鏂帮紙涓◣APP锛?
5. 璇存槑绀句繚鍩烘暟濡傛湁璋冩暣璇蜂互绀句繚灞€鏍稿畾涓哄噯
6. 鎬诲瓧鏁版帶鍒跺湪 500-800 瀛?
7. 鐢ㄤ腑鏂囪緭鍑猴紝涓嶈杈撳嚭鑻辨枃
"""

        ai_result = ask_deepseek(
            prompt,
            system_prompt="浣犳槸涓€浣嶄笓涓氱殑绋庡姟椤鹃棶锛屾搮闀挎挵鍐欎紒涓氱◣鍔＄敵鎶ヨ鏄庛€?
        )
        if not ai_result.startswith("["):
            return ai_result

    # 妯℃嫙妯″紡
    has_tax = any(r["搴旂撼绋庨"] > 0 for r in results)
    lines = [f"銆恵now.year}骞磠now.month}鏈堜釜绋庣敵鎶ヨ鏄庛€?, ""]
    lines.append(f"鏈湀鍏徃鍏辨湁 {len(results)} 鍚嶅憳宸ラ渶杩涜涓◣鐢虫姤銆?)
    if has_tax:
        total_tax = sum(r["搴旂撼绋庨"] for r in results)
        lines.append(f"鏈湀搴旂撼涓◣鍚堣 {total_tax:.2f} 鍏冿紝璇峰強鏃跺湪鑷劧浜虹數瀛愮◣鍔″眬锛堟墸缂寸锛夊畬鎴愮敵鎶ョ即绋庛€?)
    else:
        lines.append("缁忚绠楋紝鏈湀鎵€鏈夊憳宸ュ簲绋庢敹鍏ュ潎涓?0 鍏冿紝鏃犻渶缂寸撼涓◣銆傝鍦ㄨ嚜鐒朵汉鐢靛瓙绋庡姟灞€杩涜闆剁敵鎶ユ搷浣溿€?)
    lines.append("")
    lines.append("銆愭墸闄ら」璇存槑銆?)
    for r in results:
        lines.append(
            f"  {r['濮撳悕']}锛氱ぞ淇濇墸闄?{r['涓汉绀句繚']} 鍏冿紝"
            f"涓撻」闄勫姞鎵ｉ櫎 {r['涓撻」闄勫姞鎵ｉ櫎']} 鍏?
            f"锛堝瓙濂虫暀鑲瞷r['瀛愬コ鏁欒偛']}+濠村辜鍎縶r['濠村辜鍎跨収鎶?]}+璧″吇鑰佷汉{r['璧″吇鑰佷汉']}锛夈€?
        )
    lines.append("")
    lines.append("銆愭敞鎰忎簨椤广€?)
    lines.append("  1. 璇锋牳瀹炲憳宸ヤ笓椤归檮鍔犳墸闄や俊鎭槸鍚︽渶鏂帮紱")
    lines.append("  2. 绀句繚鍩烘暟濡傛湁璋冩暣锛岃鍙婃椂鏇存柊绯荤粺鍙傛暟锛?)
    lines.append("  3. 闆剁敵鎶ヤ篃闇€鎸夋椂鎻愪氦锛岄伩鍏嶄骇鐢熼€炬湡璁板綍銆?)
    lines.append("")
    lines.append("鈥斺€?鐢?閲戣壋榫橝I绋庡姟鍔╂墜 鑷姩鐢熸垚")
    return "\n".join(lines)


def generate_social_report_ai(results: list) -> str:
    """鐢熸垚绀句繚鐢虫姤璇存槑锛堢湡瀹?AI 鎴栨ā鎷燂級"""
    now = datetime.now()
    api_key, _ = get_api_key()
    use_ai = bool(api_key)

    if use_ai:
        rows_text = ""
        for r in results:
            rows_text += (
                f"  鍛樺伐{r['濮撳悕']}锛氱即璐瑰熀鏁?5000 鍏冿紝"
                f"鍏徃绀句繚鎵挎媴 {r['鍏徃绀句繚鎵挎媴']} 鍏冿紝"
                f"涓汉绀句繚 {r['涓汉绀句繚']} 鍏冦€俓n"
            )
        total_si = sum(r["鍏徃绀句繚鎵挎媴"] for r in results)

        prompt = f"""璇蜂负姝︽眽閲戣壋榫欑鎶€鏈夐檺鍏徃鐢熸垚 {now.year}骞磠now.month}鏈?鐨勭ぞ淇濈敵鎶ユ搷浣滆鏄庛€?

绀句繚鏁版嵁锛?
{rows_text}
姹囨€伙細鍏徃鎵挎媴绀句繚鍚堣 {total_si} 鍏冦€?

瑕佹眰锛?
1. 璇存槑绀句繚缂寸撼鏄庣粏鍜屽叕鍙告壙鎷呴儴鍒?
2. 鎻愪緵鎿嶄綔鎸囧紩锛堢櫥褰曟箹鍖楁斂鍔℃湇鍔＄綉锛岃繘鍏ュ崟浣嶇ぞ淇濈敵鎶ユā鍧楋級
3. 鎻愰啋鐢虫姤鎴鏃堕棿鍜屾敞鎰忎簨椤?
4. 璇皵涓撲笟锛?00-300 瀛楋紝鐢ㄤ腑鏂囪緭鍑?
"""

        ai_result = ask_deepseek(prompt)
        if not ai_result.startswith("["):
            return ai_result

    # 妯℃嫙妯″紡
    total_si = sum(r["鍏徃绀句繚鎵挎媴"] for r in results)
    lines = [
        f"銆恵now.year}骞磠now.month}鏈堢ぞ淇濈敵鎶ヨ鏄庛€?,
        "",
        f"鏈湀闇€涓?{len(results)} 鍚嶅憳宸ョ即绾崇ぞ淇濓紝鍏徃鎵挎媴閮ㄥ垎鍚堣 {total_si:.2f} 鍏冦€?,
        "",
        "銆愮即璐规槑缁嗐€?,
    ]
    for r in results:
        lines.append(
            f"  {r['濮撳悕']}锛氱即璐瑰熀鏁?5000 鍏冿紝"
            f"鍏徃鎵挎媴 {r['鍏徃绀句繚鎵挎媴']} 鍏冿紝"
            f"涓汉鎵挎媴 {r['涓汉绀句繚']} 鍏冦€?
        )
    lines.append("")
    lines.append("銆愭搷浣滄寚寮曘€?)
    lines.append("  1. 鐧诲綍銆屾箹鍖楁斂鍔℃湇鍔＄綉銆嶆垨銆屾姹夌ぞ淇濈敵鎶ョ郴缁熴€嶏紱")
    lines.append("  2. 杩涘叆銆屽崟浣嶇ぞ淇濈敵鎶ャ€嶆ā鍧楋紝鏍稿浜哄憳鍚嶅崟锛?)
    lines.append("  3. 纭缂磋垂鍩烘暟鏃犺鍚庢彁浜ょ敵鎶ワ紱")
    lines.append("  4. 缂磋垂鎴愬姛鍚庣暀瀛樼即璐瑰嚟璇佸鏌ャ€?)
    lines.append("")
    lines.append("銆愭敞鎰忎簨椤广€?)
    lines.append("  绀句繚鐢虫姤鎴鏃堕棿涓烘瘡鏈?25 鏃ワ紝璇锋彁鍓嶅姙鐞嗐€?)
    lines.append("")
    lines.append("鈥斺€?鐢?閲戣壋榫橝I绋庡姟鍔╂墜 鑷姩鐢熸垚")
    return "\n".join(lines)


# ===============================================
#  鐢熸垚涓婁紶妯℃澘锛堝唴瀛樹腑锛?
# ===============================================

def get_template_df():
    """杩斿洖绀鸿寖鐢ㄧ殑涓婁紶妯℃澘 DataFrame"""
    return pd.DataFrame([
        {
            "濮撳悕": "鍛樺伐A",
            "绋庡墠宸ヨ祫": 10522,
            "绀句繚鍩烘暟": 5000,
            "涓汉绀句繚瀹炵即": 522,
            "涓撻」闄勫姞鎵ｉ櫎": 5000,
            "瀛愬コ鏁欒偛": 2000,
            "濠村辜鍎跨収鎶?: 2000,
            "璧″吇鑰佷汉": 1000,
        },
        {
            "濮撳悕": "鍛樺伐B",
            "绋庡墠宸ヨ祫": 8000,
            "绀句繚鍩烘暟": 5000,
            "涓汉绀句繚瀹炵即": 522,
            "涓撻」闄勫姞鎵ｉ櫎": 0,
            "瀛愬コ鏁欒偛": 0,
            "濠村辜鍎跨収鎶?: 0,
            "璧″吇鑰佷汉": 0,
        },
    ])


# ===============================================
#  骞存姤瀵煎叆妯℃澘鐢熸垚
# ===============================================

ANNUAL_TEMPLATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "骞存姤鏁版嵁瀵煎叆妯℃澘.xlsx")


def gen_annual_report_template_bytes() -> bytes:
    """鐢熸垚銆屽勾鎶ユ暟鎹鍏ャ€岴xcel 妯℃澘锛?涓猄heet锛夛紝杩斿洖瀛楄妭娴?""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # 鈹€鈹€ 閫氱敤鏍峰紡 鈹€鈹€
    header_font = Font(name="寰蒋闆呴粦", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name="寰蒋闆呴粦", size=10)
    body_align = Alignment(horizontal="left", vertical="center")
    num_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    section_font = Font(name="寰蒋闆呴粦", bold=True, size=11, color="1E40AF")
    hint_font = Font(name="寰蒋闆呴粦", size=9, color="6B7280")

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def style_body_row(ws, row, cols, is_num=False):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = body_font
            cell.alignment = num_align if (is_num and c >= 2) else body_align
            cell.border = thin_border

    # ============================================================
    #  Sheet 1锛氬勾鎶ユ眹鎬?
    # ============================================================
    ws1 = wb.active
    ws1.title = "骞存姤姹囨€?

    ws1.merge_cells("A1:D1")
    ws1.cell(row=1, column=1, value="姝︽眽閲戣壋榫欑鎶€鏈夐檺鍏徃 路 2025骞村害骞存姤鏁版嵁瀵煎叆妯℃澘").font = Font(name="寰蒋闆呴粦", bold=True, size=14, color="1E3A8A")
    ws1.cell(row=2, column=1, value="濉啓璇存槑锛氫粎闇€鍦ㄣ€屾暟鍊笺€嶅垪濉叆瀹為檯閲戦锛涚伆鑹茶鏃犻渶濉啓锛涘～瀹屽悗鍒囨崲鍒般€屽憳宸ヤ俊鎭€峉heet 濉啓鍛樺伐鏄庣粏").font = hint_font
    ws1.merge_cells("A2:D2")

    headers1 = ["搴忓彿", "椤圭洰鍚嶇О", "鏁板€?, "鍗曚綅 / 澶囨敞"]
    for c, h in enumerate(headers1, 1):
        ws1.cell(row=4, column=c, value=h)
    style_header(ws1, 4, 4)

    items1 = [
        # (搴忓彿, 椤圭洰, 榛樿鍊? 澶囨敞)
        (1, "鍏ㄥ勾钀ヤ笟鏀跺叆锛堝惈绋庯級", 500000.00, "鍏冿紝鍙栬嚜鍒╂鼎琛ㄧ1琛?),
        (2, "鍏ㄥ勾钀ヤ笟鎴愭湰", 200000.00, "鍏冿紝鍙栬嚜鍒╂鼎琛ㄧ2琛?),
        (3, "鍏ㄥ勾鍒╂鼎鎬婚", 50000.00, "鍏冿紝鍙栬嚜鍒╂鼎琛ㄧ12琛?),
        (4, "澧炲€肩◣璁＄◣鏀跺叆锛堜笉鍚◣锛?, 495049.50, "鍏冿紝灏忚妯＄撼绋庝汉濉惈绋幟?.01"),
        (5, "骞冲潎浠庝笟浜烘暟", 2, "浜猴紝鍏ㄥ勾鍚勫搴﹀钩鍧?),
        (6, "骞冲潎璧勪骇鎬婚", 50.00, "涓囧厓"),
        (7, "鈥斺€?浠ヤ笅涓哄叾浠栫◣绉嶅弬鏁?鈥斺€?, "", "鈥斺€?),
        (8, "娉ㄥ唽璧勬湰瀹炵即棰濓紙绱锛?, 300000.00, "鍏冿紝濉疄闄呭凡鍒颁綅閲戦锛堥潪璁ょ即棰濓級锛屼緥锛氭敞鍐岃祫鏈?00涓囧凡鍒颁綅30涓囧垯濉?00000"),
        (9, "鏈勾澧炶祫棰?, 0.00, "鍏冿紝鏈勾鏂板瀹炴敹璧勬湰锛屾棤澧炶祫濉?"),
        (10, "鍏ㄥ勾宸ヨ祫鎬婚", 111132.00, "鍏冿紝鍏ㄩ儴鍛樺伐鐨勭◣鍓嶅伐璧勫勾鍚堣"),
        (11, "鍏ㄥ勾绀句繚鍏徃鎵挎媴閮ㄥ垎", 34560.00, "鍏冿紝绾︿负宸ヨ祫鐨?5.6%"),
        (12, "涓婂勾骞冲潎鍦ㄨ亴鑱屽伐浜烘暟", 2, "浜猴紝鐢ㄤ簬娈嬩繚閲戞祴绠?),
        (13, "涓婂勾瀹夋帓娈嬬柧浜哄氨涓氫汉鏁?, 0, "浜猴紝闇€娈嬭仈瀹℃牳纭"),
        (14, "涓婂勾鑱屽伐骞村钩鍧囧伐璧?, 60000.00, "鍏?骞?),
        (15, "褰撳湴绀句細骞冲潎宸ヨ祫锛堝厓/骞达級", 90000.00, "鍏?骞达紝姝︽眽2024骞寸害8~9涓?),
        (16, "鈥斺€?瀛ｅ害鍒嗘憡鏂瑰紡锛堜簩閫変竴锛夆€斺€?, "", "鈥斺€?),
        (17, "瀛ｅ害鍒嗘憡鏂瑰紡", "骞冲潎", "濉€屽钩鍧囥€嶆寜4瀛ｅ害鍧囧垎锛涘～銆屾槑缁嗐€嶈鍦⊿heet3濉啓鍚勫搴︽暟鎹?),
    ]
    for i, (seq, name, val, note) in enumerate(items1):
        r = 5 + i
        ws1.cell(row=r, column=1, value=seq)
        ws1.cell(row=r, column=2, value=name)
        ws1.cell(row=r, column=3, value=val if val != "" else "")
        ws1.cell(row=r, column=4, value=note)
        # 鍒嗛殧琛岀伆鑹?
        if name.startswith("鈥斺€?):
            for c in range(1, 5):
                ws1.cell(row=r, column=c).font = Font(name="寰蒋闆呴粦", size=9, color="9CA3AF", italic=True)
        else:
            style_body_row(ws1, r, 4, is_num=(isinstance(val, (int, float)) and val != ""))

    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 42

    # ============================================================
    #  Sheet 2锛氬憳宸ヤ俊鎭?
    # ============================================================
    ws2 = wb.create_sheet("鍛樺伐淇℃伅")
    ws2.merge_cells("A1:H1")
    ws2.cell(row=1, column=1, value="鍛樺伐宸ヨ祫涓庝笓椤归檮鍔犳墸闄ゆ槑缁?).font = Font(name="寰蒋闆呴粦", bold=True, size=13, color="1E3A8A")
    ws2.cell(row=2, column=1, value="濉啓璇存槑锛氭瘡琛屼竴鍚嶅憳宸ワ紝閲戦鍧囦负鏈堝潎鏁帮紱鍛樺伐浜烘暟涓?Sheet1銆屽钩鍧囦粠涓氫汉鏁般€嶄竴鑷?).font = hint_font
    ws2.merge_cells("A2:H2")

    headers2 = ["濮撳悕", "绋庡墠鏈堝伐璧?, "绀句繚缂磋垂鍩烘暟", "涓汉绀句繚鏈堝疄缂?, "涓撻」闄勫姞鎵ｉ櫎鍚堣",
                 "瀛愬コ鏁欒偛", "濠村辜鍎跨収鎶?, "璧″吇鑰佷汉"]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=4, column=c, value=h)
    style_header(ws2, 4, 8)

    sample_emps = [
        ["寮犱笁", 10522, 5000, 522, 5000, 2000, 2000, 1000],
        ["鏉庡洓", 8000, 5000, 522, 0, 0, 0, 0],
    ]
    for i, emp in enumerate(sample_emps):
        r = 5 + i
        for c, val in enumerate(emp, 1):
            ws2.cell(row=r, column=c, value=val)
        style_body_row(ws2, r, 8, is_num=False)
        # 鏁板瓧鍒楀彸瀵归綈
        for c in range(2, 9):
            ws2.cell(row=r, column=c).alignment = num_align

    for c, w in enumerate([10, 14, 14, 16, 18, 12, 12, 12], 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    # ============================================================
    #  Sheet 3锛氬搴﹀垎鎽婃槑缁嗭紙鍙€夛級
    # ============================================================
    ws3 = wb.create_sheet("瀛ｅ害鍒嗘憡鏄庣粏")
    ws3.merge_cells("A1:E1")
    ws3.cell(row=1, column=1, value="鍚勫搴﹁惀涓氭敹鍏?鎴愭湰/鍒╂鼎鏄庣粏锛堝彲閫夛級").font = Font(name="寰蒋闆呴粦", bold=True, size=13, color="1E3A8A")
    ws3.cell(row=2, column=1, value="濉啓璇存槑锛氫粎褰?Sheet1 瀛ｅ害鍒嗘憡鏂瑰紡閫夈€屾槑缁嗐€嶆椂闇€濉啓锛涢€夈€屽钩鍧囥€嶅垯蹇界暐鏈〃").font = hint_font
    ws3.merge_cells("A2:E2")

    headers3 = ["瀛ｅ害", "钀ヤ笟鏀跺叆锛堝厓锛?, "钀ヤ笟鎴愭湰锛堝厓锛?, "鍒╂鼎鎬婚锛堝厓锛?, "澧炲€肩◣璁＄◣鏀跺叆锛堝厓锛?]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=4, column=c, value=h)
    style_header(ws3, 4, 5)

    for i, q in enumerate(["Q1", "Q2", "Q3", "Q4"]):
        r = 5 + i
        ws3.cell(row=r, column=1, value=q)
        for c in range(2, 6):
            ws3.cell(row=r, column=c, value=0)
        style_body_row(ws3, r, 5, is_num=True)

    ws3.column_dimensions["A"].width = 8
    for c in range(2, 6):
        ws3.column_dimensions[get_column_letter(c)].width = 18

    # 淇濆瓨鍒板瓧鑺傛祦
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def gen_annual_report_template_pdf_bytes() -> bytes | None:
    """鐢熸垚銆屽勾鎶ユ暟鎹鍏ャ€峆DF 妯℃澘锛堟棤闇€濉啓锛岀敤浜庢墦鍗?瀛樻。锛夛紝杩斿洖瀛楄妭娴?""
    if not _HAS_FPDF:
        return None
    try:
        from fpdf import FPDF

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=12)

        # 娉ㄥ唽涓枃瀛椾綋
        font_path = r"C:\Windows\Fonts\msyh.ttc"
        if not os.path.exists(font_path):
            font_path = r"C:\Windows\Fonts\simsunb.ttf"
        if os.path.exists(font_path):
            pdf.add_font("CJK", "", font_path)
            pdf.add_font("CJK", "B", font_path)
            use_cjk = True
        else:
            use_cjk = False

        def text_cjk(w, h, txt, **kw):
            if use_cjk:
                pdf.set_font("CJK", kw.pop("style", "").replace("B", "B") or "", kw.pop("size", 10))
                pdf.cell(w, h, txt, **kw)
            else:
                safe = txt.encode("ascii", errors="replace").decode("ascii")
                pdf.set_font("Helvetica", kw.pop("style", "").replace("B", "B") or "", kw.pop("size", 10))
                pdf.cell(w, h, safe, **kw)

        def multi_cjk(w, h, txt, **kw):
            if use_cjk:
                pdf.set_font("CJK", kw.pop("style", "").replace("B", "B") or "", kw.pop("size", 10))
                pdf.multi_cell(w, h, txt, **kw)
            else:
                safe = txt.encode("ascii", errors="replace").decode("ascii")
                pdf.set_font("Helvetica", kw.pop("style", "").replace("B", "B") or "", kw.pop("size", 10))
                pdf.multi_cell(w, h, safe, **kw)

        # 鈹€鈹€ 灏侀潰 鈹€鈹€
        pdf.add_page()
        pdf.ln(30)
        if use_cjk:
            pdf.set_font("CJK", "B", 20)
        else:
            pdf.set_font("Helvetica", "B", 20)
        pdf.cell(0, 12, "姝︽眽閲戣壋榫欑鎶€鏈夐檺鍏徃", ln=True, align="C")
        pdf.ln(4)
        if use_cjk:
            pdf.set_font("CJK", "B", 16)
        else:
            pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, "2025 骞村害骞存姤鏁版嵁瀵煎叆妯℃澘", ln=True, align="C")
        pdf.ln(10)
        if use_cjk:
            pdf.set_font("CJK", "", 10)
        else:
            pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 8, "鐢熸垚鏃ユ湡锛? + datetime.now().strftime("%Y-%m-%d"), ln=True, align="C")
        pdf.cell(0, 8, "璇存槑锛氭湰 PDF 涓烘ā鏉垮弬鑰冩枃浠讹紝瀹為檯鏁版嵁璇烽€氳繃 Excel 鐗堝鍏?, ln=True, align="C")
        pdf.cell(0, 8, "Excel 瀵煎叆璺緞锛氱郴缁熶晶杈规爮 鈫?骞存姤瀵煎叆 鈫?涓婁紶 .xlsx 鏂囦欢", ln=True, align="C")

        # 鈹€鈹€ 绗?閮ㄥ垎锛氬勾鎶ユ眹鎬?鈹€鈹€
        pdf.add_page()
        if use_cjk:
            pdf.set_font("CJK", "B", 14)
        else:
            pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "涓€銆佸勾鎶ユ眹鎬绘暟鎹?, ln=True)
        pdf.ln(4)

        col_w = [12, 86, 28, 62]
        headers = ["搴忓彿", "椤圭洰鍚嶇О", "鏁板€?, "鍗曚綅/澶囨敞"]
        if use_cjk:
            pdf.set_font("CJK", "B", 8)
        else:
            pdf.set_font("Helvetica", "B", 8)
        for i, (h, w) in enumerate(zip(headers, col_w)):
            pdf.cell(w, 7, h, border=1, align="C")
        pdf.ln()

        items = [
            ("1", "鍏ㄥ勾钀ヤ笟鏀跺叆锛堝惈绋庯級", "500,000.00", "鍏?),
            ("2", "鍏ㄥ勾钀ヤ笟鎴愭湰", "200,000.00", "鍏?),
            ("3", "鍏ㄥ勾鍒╂鼎鎬婚", "50,000.00", "鍏?),
            ("4", "澧炲€肩◣璁＄◣鏀跺叆锛堜笉鍚◣锛?, "495,049.50", "鍏?),
            ("5", "骞冲潎浠庝笟浜烘暟", "2", "浜?),
            ("6", "骞冲潎璧勪骇鎬婚", "50.00", "涓囧厓"),
            ("鈥?, "鈥斺€?浠ヤ笅涓哄叾浠栫◣绉嶅弬鏁?鈥斺€?, "", ""),
            ("7", "娉ㄥ唽璧勬湰瀹炵即棰濓紙绱锛?, "300,000.00", "鍏冿紙宸插埌浣嶉儴鍒嗭紝闈炶缂村叏棰濓級"),
            ("8", "鏈勾澧炶祫棰?, "0.00", "鍏?),
            ("9", "鍏ㄥ勾宸ヨ祫鎬婚", "111,132.00", "鍏?),
            ("10", "鍏ㄥ勾绀句繚鍏徃鎵挎媴閮ㄥ垎", "34,560.00", "鍏?),
            ("11", "涓婂勾骞冲潎鍦ㄨ亴鑱屽伐浜烘暟", "2", "浜?),
            ("12", "涓婂勾瀹夋帓娈嬬柧浜哄氨涓氫汉鏁?, "0", "浜?),
            ("13", "涓婂勾鑱屽伐骞村钩鍧囧伐璧?, "60,000.00", "鍏?骞?),
            ("14", "褰撳湴绀句細骞冲潎宸ヨ祫", "90,000.00", "鍏?骞?),
            ("鈥?, "鈥斺€?瀛ｅ害鍒嗘憡鏂瑰紡 鈥斺€?, "", ""),
            ("15", "瀛ｅ害鍒嗘憡鏂瑰紡", "骞冲潎", "濉€屽钩鍧囥€嶆垨銆屾槑缁嗐€?),
        ]
        if use_cjk:
            pdf.set_font("CJK", "", 8)
        else:
            pdf.set_font("Helvetica", "", 8)
        for seq, name, val, note in items:
            is_section = seq == "鈥?
            fs = "CJK" if use_cjk else "Helvetica"
            if is_section:
                pdf.set_font(fs, "", 8)
                pdf.set_text_color(150, 150, 150)
            else:
                pdf.set_font(fs, "", 8)
                pdf.set_text_color(0, 0, 0)
            pdf.cell(col_w[0], 6, seq, border=1, align="C")
            pdf.cell(col_w[1], 6, name, border=1, align="L")
            pdf.cell(col_w[2], 6, val, border=1, align="R")
            pdf.cell(col_w[3], 6, note, border=1, align="L")
            pdf.ln()

        # 鈹€鈹€ 绗?閮ㄥ垎锛氬憳宸ヤ俊鎭?鈹€鈹€
        pdf.add_page()
        if use_cjk:
            pdf.set_font("CJK", "B", 14)
        else:
            pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "浜屻€佸憳宸ュ伐璧勪笌涓撻」闄勫姞鎵ｉ櫎鏄庣粏", ln=True)
        pdf.ln(4)

        emp_col_w = [18, 23, 23, 25, 28, 20, 22, 22]
        emp_headers = ["濮撳悕", "绋庡墠鏈堝伐璧?, "绀句繚缂磋垂鍩烘暟", "涓汉绀句繚鏈堝疄缂?,
                        "涓撻」闄勫姞鎵ｉ櫎鍚堣", "瀛愬コ鏁欒偛", "濠村辜鍎跨収鎶?, "璧″吇鑰佷汉"]
        if use_cjk:
            pdf.set_font("CJK", "B", 7)
        else:
            pdf.set_font("Helvetica", "B", 7)
        for h, w in zip(emp_headers, emp_col_w):
            pdf.cell(w, 7, h, border=1, align="C")
        pdf.ln()

        sample_emps = [
            ["寮犱笁", "10,522", "5,000", "522", "5,000", "2,000", "2,000", "1,000"],
            ["鏉庡洓", "8,000", "5,000", "522", "0", "0", "0", "0"],
        ]
        if use_cjk:
            pdf.set_font("CJK", "", 7)
        else:
            pdf.set_font("Helvetica", "", 7)
        for emp in sample_emps:
            for val, w in zip(emp, emp_col_w):
                pdf.cell(w, 6, val, border=1, align="R" if val.replace(",", "").replace(".", "").isdigit() else "L")
            pdf.ln()

        # 鈹€鈹€ 绗?閮ㄥ垎锛氬搴﹀垎鎽婃槑缁?鈹€鈹€
        pdf.ln(8)
        if use_cjk:
            pdf.set_font("CJK", "B", 14)
        else:
            pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "涓夈€佸搴﹀垎鎽婃槑缁嗭紙鍙€夛級", ln=True)
        pdf.ln(4)

        q_col_w = [20, 40, 40, 40, 40]
        q_headers = ["瀛ｅ害", "钀ヤ笟鏀跺叆锛堝厓锛?, "钀ヤ笟鎴愭湰锛堝厓锛?, "鍒╂鼎鎬婚锛堝厓锛?, "澧炲€肩◣璁＄◣鏀跺叆锛堝厓锛?]
        if use_cjk:
            pdf.set_font("CJK", "B", 7)
        else:
            pdf.set_font("Helvetica", "B", 7)
        for h, w in zip(q_headers, q_col_w):
            pdf.cell(w, 7, h, border=1, align="C")
        pdf.ln()

        if use_cjk:
            pdf.set_font("CJK", "", 7)
        else:
            pdf.set_font("Helvetica", "", 7)
        for q in ["Q1", "Q2", "Q3", "Q4"]:
            pdf.cell(q_col_w[0], 6, q, border=1, align="C")
            for w in q_col_w[1:]:
                pdf.cell(w, 6, "0", border=1, align="R")
            pdf.ln()

        pdf.ln(6)
        if use_cjk:
            pdf.set_font("CJK", "", 8)
        else:
            pdf.set_font("Helvetica", "", 8)
        pdf.multi_cell(0, 5,
            "娉ㄦ剰锛氭湰 PDF 涓烘ā鏉垮弬鑰冩枃浠讹紝涓嶅彲鐩存帴瀵煎叆銆傝涓嬭浇 Excel 鐗堟湰濉啓鏁版嵁鍚庝笂浼犲鍏ャ€俓n"
            "姝︽眽閲戣壋榫欑鎶€鏈夐檺鍏徃 路 灏忚妯＄撼绋庝汉 路 灏忓瀷寰埄浼佷笟\n"
            f"妯℃澘鐢熸垚鏃ユ湡锛歿datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )

        return bytes(pdf.output())
    except Exception as e:
        return None


def parse_annual_report_excel(file_bytes: bytes) -> dict:
    """
    瑙ｆ瀽骞存姤瀵煎叆 Excel锛岃繑鍥炵粨鏋勫寲鏁版嵁銆?
    杩斿洖鏍煎紡锛?
    {
      "summary": { 骞存姤姹囨€诲瓧娈?},
      "employees": [ 鍛樺伐鍒楄〃 ],
      "quarterly": { "Q1": {...}, "Q2": {...}, "Q3": {...}, "Q4": {...} } or None,
      "warnings": [ 鏍￠獙璀﹀憡 ],
    }
    """
    import openpyxl
    import io

    result = {
        "summary": {},
        "employees": [],
        "quarterly": None,
        "warnings": [],
    }

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # 鈹€鈹€ 瑙ｆ瀽 Sheet1锛氬勾鎶ユ眹鎬?鈹€鈹€
    if "骞存姤姹囨€? in wb.sheetnames:
        ws1 = wb["骞存姤姹囨€?]
        key_map = {
            "鍏ㄥ勾钀ヤ笟鏀跺叆锛堝惈绋庯級": "annual_revenue",
            "鍏ㄥ勾钀ヤ笟鎴愭湰": "annual_cost",
            "鍏ㄥ勾鍒╂鼎鎬婚": "annual_profit",
            "澧炲€肩◣璁＄◣鏀跺叆锛堜笉鍚◣锛?: "annual_vat_revenue",
            "骞冲潎浠庝笟浜烘暟": "avg_employees",
            "骞冲潎璧勪骇鎬婚": "avg_assets",
            "娉ㄥ唽璧勬湰瀹炵即棰濓紙绱锛?: "reg_capital",
            "鏈勾澧炶祫棰?: "capital_increase",
            "鍏ㄥ勾宸ヨ祫鎬婚": "total_salary",
            "鍏ㄥ勾绀句繚鍏徃鎵挎媴閮ㄥ垎": "total_si_company",
            "涓婂勾骞冲潎鍦ㄨ亴鑱屽伐浜烘暟": "prev_employees",
            "涓婂勾瀹夋帓娈嬬柧浜哄氨涓氫汉鏁?: "prev_disabled",
            "涓婂勾鑱屽伐骞村钩鍧囧伐璧?: "prev_avg_salary",
            "褰撳湴绀句細骞冲潎宸ヨ祫锛堝厓/骞达級": "local_avg_salary",
            "瀛ｅ害鍒嗘憡鏂瑰紡": "split_method",
        }
        for row in ws1.iter_rows(min_row=5, values_only=True):
            if row[1] is None:
                continue
            name = str(row[1]).strip()
            if name in key_map:
                val = row[2]
                if val is None:
                    val = 0
                if isinstance(val, str):
                    val = val.strip()
                    if val == "骞冲潎":
                        result["summary"][key_map[name]] = "骞冲潎"
                    elif val == "鏄庣粏":
                        result["summary"][key_map[name]] = "鏄庣粏"
                    else:
                        try:
                            val = float(val.replace(",", "").replace("锛?, ""))
                        except ValueError:
                            result["warnings"].append(f"銆寋name}銆嶇殑鍊笺€寋val}銆嶆棤娉曡瘑鍒紝宸茶烦杩?)
                            continue
                elif isinstance(val, (int, float)):
                    pass  # 淇濇寔鍘熷€?
                else:
                    result["warnings"].append(f"銆寋name}銆嶇殑鏍煎紡涓嶆敮鎸侊紝宸茶烦杩?)
                    continue
                result["summary"][key_map[name]] = val

    # 鈹€鈹€ 瑙ｆ瀽 Sheet2锛氬憳宸ヤ俊鎭?鈹€鈹€
    if "鍛樺伐淇℃伅" in wb.sheetnames:
        ws2 = wb["鍛樺伐淇℃伅"]
        for row in ws2.iter_rows(min_row=5, values_only=True):
            if row[0] is None or str(row[0]).strip() == "":
                continue
            name = str(row[0]).strip()
            if name.startswith("锛堢ず渚嬶級") or name == "濮撳悕":
                continue
            emp = {
                "name": name,
                "gross_salary": float(row[1] or 0),
                "si_base": float(row[2] or 0),
                "si_personal_actual": float(row[3] or 0),
                "special_deductions": float(row[4] or 0),
                "child_education": float(row[5] or 0),
                "infant_care": float(row[6] or 0),
                "elderly_care": float(row[7] or 0),
            }
            result["employees"].append(emp)

    # 鈹€鈹€ 瑙ｆ瀽 Sheet3锛氬搴﹀垎鎽婃槑缁嗭紙鍙€夛級鈹€鈹€
    if "瀛ｅ害鍒嗘憡鏄庣粏" in wb.sheetnames:
        ws3 = wb["瀛ｅ害鍒嗘憡鏄庣粏"]
        quarterly = {}
        for row in ws3.iter_rows(min_row=5, values_only=True):
            if row[0] is None:
                continue
            q_name = str(row[0]).strip().upper()
            if q_name in ("Q1", "Q2", "Q3", "Q4"):
                quarterly[q_name] = {
                    "revenue": float(row[1] or 0),
                    "cost": float(row[2] or 0),
                    "period_profit": float(row[3] or 0),
                    "vat_revenue": float(row[4] or 0),
                }
        if quarterly:
            result["quarterly"] = quarterly

    wb.close()

    # 鈹€鈹€ 鏍￠獙 鈹€鈹€
    s = result["summary"]
    if s.get("annual_revenue", 0) <= 0 and s.get("annual_cost", 0) <= 0:
        result["warnings"].append("钀ヤ笟鏀跺叆鍜岃惀涓氭垚鏈潎涓?0锛岃纭鏄惁宸插～鍏ュ勾鎶ユ暟鎹?)
    if s.get("avg_employees", 0) <= 0:
        result["warnings"].append("骞冲潎浠庝笟浜烘暟涓?0")
    if len(result["employees"]) == 0:
        result["warnings"].append("鏈В鏋愬埌鍛樺伐淇℃伅锛岃妫€鏌ャ€屽憳宸ヤ俊鎭€峉heet 鏄惁宸插～鍐?)
    if len(result["employees"]) > 0 and s.get("avg_employees", 0) > 0:
        if len(result["employees"]) != int(s["avg_employees"]):
            result["warnings"].append(
                f"鍛樺伐浜烘暟锛坽len(result['employees'])}浜猴級涓庛€屽钩鍧囦粠涓氫汉鏁般€嶏紙{int(s['avg_employees'])}浜猴級涓嶄竴鑷?
            )

    return result


def parse_annual_report_pdf(file_bytes: bytes) -> dict:
    """
    鐢?pdfplumber 鎻愬彇 PDF 鏂囨湰锛屽啀鐢?DeepSeek AI 瑙ｆ瀽涓虹粨鏋勫寲骞存姤鏁版嵁銆?
    杩斿洖鏍煎紡涓?parse_annual_report_excel() 涓€鑷淬€?
    """
    import pdfplumber
    import io
    import re

    result = {
        "summary": {},
        "employees": [],
        "quarterly": None,
        "warnings": [],
    }

    # 鈹€鈹€ Step 1锛氭彁鍙?PDF 鍏ㄦ枃 鈹€鈹€
    full_text_parts = []
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text_parts.append(text)
        full_text = "\n".join(full_text_parts)
    except Exception as e:
        result["warnings"].append(f"PDF 鏂囨湰鎻愬彇澶辫触锛歿e}")
        return result

    if not full_text.strip():
        result["warnings"].append("PDF 涓湭鎻愬彇鍒版枃瀛楀唴瀹癸紝鍙兘涓烘壂鎻忎欢锛堝浘鐗囧瀷 PDF锛夛紝璇蜂娇鐢?Excel 瀵煎叆")
        return result

    # 鈹€鈹€ Step 2锛欴eepSeek AI 瑙ｆ瀽 鈹€鈹€
    api_key, _ = get_api_key()
    if not api_key:
        result["warnings"].append("鏈厤缃?DeepSeek API Key锛屾棤娉?AI 瑙ｆ瀽 PDF銆傝鍏堥厤缃?API Key 鎴栨敼鐢?Excel 瀵煎叆")
        return result

    system_prompt = """浣犳槸涓€涓◣鍔℃暟鎹彁鍙栧姪鎵嬨€備粠骞存姤 PDF 鏂囨湰涓彁鍙栫粨鏋勫寲鏁版嵁锛岃繑鍥炵函 JSON锛堜笉瑕?Markdown 浠ｇ爜鍧楋級銆?

JSON 鏍煎紡锛?
{
  "summary": {
    "annual_revenue": 鏁板瓧锛堝叏骞磋惀涓氭敹鍏ュ惈绋庯紝鍏冿級,
    "annual_cost": 鏁板瓧锛堝叏骞磋惀涓氭垚鏈紝鍏冿級,
    "annual_profit": 鏁板瓧锛堝叏骞村埄娑︽€婚锛屽厓锛?
    "annual_vat_revenue": 鏁板瓧锛堝鍊肩◣璁＄◣鏀跺叆涓嶅惈绋庯紝鍏冿級,
    "avg_employees": 鏁板瓧锛堝钩鍧囦粠涓氫汉鏁帮級,
    "avg_assets": 鏁板瓧锛堝钩鍧囪祫浜ф€婚锛屼竾鍏冿級,
    "reg_capital": 鏁板瓧锛堟敞鍐岃祫鏈疄缂撮绱锛屽厓锛?
    "capital_increase": 鏁板瓧锛堟湰骞村璧勯锛屽厓锛?
    "total_salary": 鏁板瓧锛堝叏骞村伐璧勬€婚锛屽厓锛?
    "total_si_company": 鏁板瓧锛堝叏骞寸ぞ淇濆叕鍙告壙鎷呴儴鍒嗭紝鍏冿級,
    "prev_employees": 鏁板瓧锛堜笂骞村钩鍧囧湪鑱岃亴宸ヤ汉鏁帮級,
    "prev_disabled": 鏁板瓧锛堜笂骞村畨鎺掓畫鐤句汉灏变笟浜烘暟锛?
    "prev_avg_salary": 鏁板瓧锛堜笂骞磋亴宸ュ勾骞冲潎宸ヨ祫锛屽厓/骞达級,
    "local_avg_salary": 鏁板瓧锛堝綋鍦扮ぞ浼氬钩鍧囧伐璧勶紝鍏?骞达級
  },
  "employees": [{"name": "濮撳悕", "gross_salary": 鏈堝伐璧? "si_base": 绀句繚鍩烘暟, "si_personal_actual": 涓汉绀句繚鏈堝疄缂? "special_deductions": 涓撻」鎵ｉ櫎鍚堣, "child_education": 瀛愬コ鏁欒偛, "infant_care": 濠村辜鍎跨収鎶? "elderly_care": 璧″吇鑰佷汉}],
  "notes": "琛ュ厖璇存槑"
}

瑙勫垯锛?
- 鏈壘鍒扮殑瀛楁濉?0
- 鏁板瓧涓嶈甯﹂€楀彿鎴栧崟浣?
- 娉ㄦ剰鍖哄垎銆屽叏骞淬€嶅拰銆屾湀鍧囥€嶆暟鎹?
- 娉ㄦ剰鍖哄垎銆屽惈绋庛€嶅拰銆屼笉鍚◣銆?
- 娉ㄥ唽璧勬湰鍙彁鍙栧疄闄呭埌浣嶏紙瀹炵即锛夐噾棰濓紝涓嶆槸璁ょ即閲戦
- 濡傛灉 PDF 鍙寘鍚勾鎶ユ眹鎬绘暟鎹紙鏃犲憳宸ユ槑缁嗭級锛宔mployees 杩斿洖绌烘暟缁?
"""

    prompt = f"""璇蜂粠浠ヤ笅骞存姤 PDF 鏂囨湰涓彁鍙栫粨鏋勫寲鏁版嵁锛?

===== PDF 鏂囨湰寮€濮?=====
{full_text[:8000]}
===== PDF 鏂囨湰缁撴潫 =====

璇疯繑鍥?JSON锛?""

    try:
        ai_response = ask_deepseek(prompt, system_prompt=system_prompt)
    except Exception as e:
        result["warnings"].append(f"AI 瑙ｆ瀽璋冪敤澶辫触锛歿e}")
        return result

    # 鈹€鈹€ Step 3锛氳В鏋?AI 杩斿洖鐨?JSON 鈹€鈹€
    # 灏濊瘯鎻愬彇 JSON锛堝鐞嗗彲鑳界殑 Markdown 浠ｇ爜鍧楋級
    json_match = re.search(r'\{[\s\S]*\}', ai_response)
    if not json_match:
        result["warnings"].append(f"AI 鏈繑鍥炴湁鏁?JSON銆傚師濮嬪搷搴旓細\n{ai_response[:500]}")
        return result

    try:
        parsed = json.loads(json_match.group())
    except json.JSONDecodeError:
        result["warnings"].append(f"AI 杩斿洖鐨?JSON 瑙ｆ瀽澶辫触銆傚師濮嬪搷搴旓細\n{ai_response[:500]}")
        return result

    # 鈹€鈹€ Step 4锛氭爣鍑嗗寲瀛楁 鈹€鈹€
    s = parsed.get("summary", {})
    key_map = {
        "annual_revenue": "annual_revenue",
        "annual_cost": "annual_cost",
        "annual_profit": "annual_profit",
        "annual_vat_revenue": "annual_vat_revenue",
        "avg_employees": "avg_employees",
        "avg_assets": "avg_assets",
        "reg_capital": "reg_capital",
        "capital_increase": "capital_increase",
        "total_salary": "total_salary",
        "total_si_company": "total_si_company",
        "prev_employees": "prev_employees",
        "prev_disabled": "prev_disabled",
        "prev_avg_salary": "prev_avg_salary",
        "local_avg_salary": "local_avg_salary",
    }
    for key, mapped in key_map.items():
        val = s.get(key, 0)
        try:
            result["summary"][mapped] = float(val) if val else 0.0
        except (ValueError, TypeError):
            result["summary"][mapped] = 0.0

    # 鍛樺伐鏁版嵁
    raw_emps = parsed.get("employees", [])
    for emp in raw_emps:
        if isinstance(emp, dict) and emp.get("name"):
            result["employees"].append({
                "name": str(emp.get("name", "")),
                "gross_salary": float(emp.get("gross_salary", 0) or 0),
                "si_base": float(emp.get("si_base", 0) or 0),
                "si_personal_actual": float(emp.get("si_personal_actual", 0) or 0),
                "special_deductions": float(emp.get("special_deductions", 0) or 0),
                "child_education": float(emp.get("child_education", 0) or 0),
                "infant_care": float(emp.get("infant_care", 0) or 0),
                "elderly_care": float(emp.get("elderly_care", 0) or 0),
            })

    # 闄勫姞璇存槑
    notes = parsed.get("notes", "")
    if notes:
        result["warnings"].append(f"AI 瑙ｆ瀽澶囨敞锛歿notes}")

    # 鈹€鈹€ 鏍￠獙 鈹€鈹€
    s2 = result["summary"]
    if s2.get("annual_revenue", 0) <= 0 and s2.get("annual_cost", 0) <= 0:
        result["warnings"].append("AI 鏈兘浠?PDF 涓彁鍙栧埌鏈夋晥鐨勮惀鏀?鎴愭湰鏁版嵁锛岃妫€鏌?PDF 鏄惁涓烘爣鍑嗗勾鎶ユ牸寮?)
    if len(result["employees"]) == 0:
        result["warnings"].append("PDF 涓湭鎻愬彇鍒板憳宸ユ槑缁嗭紙杩欐槸姝ｅ父鐨勶紝璇锋墜鍔ㄨˉ鍏呭憳宸ユ暟鎹級")

    # 瀛ｅ害鍒嗘憡榛樿骞冲潎
    result["summary"]["split_method"] = "骞冲潎"

    return result


# ===============================================
#  椤甸潰閰嶇疆
# ===============================================

st.set_page_config(
    page_title="閲戣壋榫橝I绋庡姟鍔╂墜",
    page_icon="馃Ь",
    layout="wide",
)

# ===============================================
#  鍏ㄥ眬鑷畾涔夋牱寮?
# ===============================================
st.markdown("""
<style>
/* 鈹€鈹€ 鏁翠綋椋庢牸 鈹€鈹€ */
.main .block-container {
    padding-top: 1.5rem;
}

/* 鈹€鈹€ 鎸囨爣鍗＄編鍖?鈹€鈹€ */
div[data-testid="stMetric"] {
    background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 100%);
    border: 1px solid #e2e8f0;
    border-radius: 10px;
    padding: 12px 16px;
    transition: all 0.2s ease;
}
div[data-testid="stMetric"]:hover {
    border-color: #94a3b8;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
}
div[data-testid="stMetric"] label {
    font-size: 0.8rem;
    color: #64748b;
}
div[data-testid="stMetric"] div[data-testid="stMetricValue"] {
    font-size: 1.4rem;
    font-weight: 700;
    color: #0f172a;
}

/* 鈹€鈹€ 琛ㄦ牸缇庡寲 鈹€鈹€ */
div[data-testid="stDataFrame"] {
    border: 1px solid #e2e8f0;
    border-radius: 8px;
    overflow: hidden;
}
div[data-testid="stDataFrame"] th {
    background: #f1f5f9 !important;
    color: #334155 !important;
    font-weight: 600 !important;
    font-size: 0.82rem !important;
}
div[data-testid="stDataFrame"] td {
    font-size: 0.85rem !important;
}

/* 鈹€鈹€ 涓绘寜閽?鈹€鈹€ */
button[kind="primary"] {
    background: linear-gradient(135deg, #2563eb 0%, #1d4ed8 100%) !important;
    border: none !important;
    font-weight: 600 !important;
    border-radius: 8px !important;
    letter-spacing: 0.02em;
}
button[kind="primary"]:hover {
    box-shadow: 0 4px 12px rgba(37,99,235,0.35) !important;
    transform: translateY(-1px);
}

/* 鈹€鈹€ 灞曞紑闈㈡澘 鈹€鈹€ */
div[data-testid="stExpander"] {
    border: 1px solid #e2e8f0;
    border-radius: 8px;
    margin-bottom: 0.5rem;
}

/* 鈹€鈹€ Radio鎸夐挳缁?鈹€鈹€ */
div[data-testid="stRadio"] label {
    font-weight: 500;
}

/* 鈹€鈹€ 鍒嗛殧绾?鈹€鈹€ */
hr {
    margin: 1.2rem 0;
    border-color: #e2e8f0;
}

/* 鈹€鈹€ Caption鏂囧瓧 鈹€鈹€ */
.stCaption {
    color: #64748b;
    font-size: 0.85rem;
}

/* 鈹€鈹€ Toast娑堟伅 鈹€鈹€ */
div[data-testid="stToast"] {
    border-radius: 10px !important;
}

/* 鈹€鈹€ 绉诲姩绔€傞厤 鈹€鈹€ */
@media (max-width: 768px) {
    div[data-testid="stMetric"] div[data-testid="stMetricValue"] {
        font-size: 1.1rem !important;
    }
    .stTabs button {
        font-size: 0.75rem !important;
        padding: 0.4rem 0.6rem !important;
    }
}
</style>
""", unsafe_allow_html=True)

# ===============================================
#  渚ц竟鏍?
# ===============================================

with st.sidebar:
    st.title("鈿欙笍 閰嶇疆")

    # API Key 鐘舵€佹樉绀?
    st.subheader("DeepSeek AI")
    current_key, key_source = get_api_key()

    if key_source in ("secrets", "env"):
        st.success("鉁?宸查€氳繃閰嶇疆鏂囦欢閰嶇疆 API Key")
        st.caption("濡傞渶涓存椂瑕嗙洊锛屽彲鍦ㄤ笅鏂硅緭鍏?)
    elif key_source == "manual":
        st.success("鉁?宸叉墜鍔ㄩ厤缃?API Key")
    else:
        st.warning("鈿狅笍 鏈厤缃?API Key")
        st.caption("璇峰湪 .streamlit/secrets.toml 涓坊鍔?DEEPSEEK_API_KEY=sk-xxx锛屾垨鍦ㄤ笅鏂硅緭鍏?)

    # 鎵嬪姩杈撳叆锛堝彲瑕嗙洊 .env锛?
    api_key_manual = st.text_input(
        "鎵嬪姩杈撳叆 API Key锛堝彲閫夛紝瑕嗙洊閰嶇疆鏂囦欢锛?,
        value=st.session_state.get("deepseek_api_key_manual", ""),
        type="password",
        help="鍦?platform.deepseek.com 鑾峰彇銆傜暀绌哄垯浣跨敤閰嶇疆鏂囦欢涓殑閰嶇疆銆?,
        key="deepseek_api_key_manual_input",
    )
    # 鍚屾鍒?session_state
    if api_key_manual != st.session_state.get("deepseek_api_key_manual", ""):
        st.session_state["deepseek_api_key_manual"] = api_key_manual
        st.rerun()

    # 娴嬭瘯杩炴帴鎸夐挳
    if st.button("馃攳 娴嬭瘯 AI 杩炴帴", use_container_width=True):
        test_result = ask_deepseek("璇风敤涓€鍙ヨ瘽浠嬬粛浣犺嚜宸?)
        if test_result.startswith("["):
            st.error(f"杩炴帴澶辫触锛歿test_result}")
        else:
            st.success("鉁?AI 杩炴帴鎴愬姛锛?)
            st.caption(test_result[:100] + "...")

    st.divider()

    # 绀句繚鍙傛暟璇存槑锛堝彲鎶樺彔锛?
    with st.expander("馃搵 绀句繚鍙傛暟璇存槑锛堟姹?2026锛?, expanded=False):
        st.caption("渚濇嵁锛氶剛浜虹ぞ鍙戙€?023銆曞強姝︽眽鍖讳繚灞€鏈€鏂版爣鍑?)
        st.markdown("**涓汉缂寸撼**")
        st.markdown("- 鍏昏€?8%\n- 鍖荤枟 2%\n- 澶变笟 **0.3%**\n- 澶х梾鍖讳繚 **7 鍏?鏈?*锛堝畾棰濓級")
        st.markdown(f"- 鍚堣 鈮?**522 鍏?鏈?*锛堝熀鏁?000脳10.3%+7锛?)
        st.markdown("**鍏徃缂寸撼**")
        st.markdown("- 鍏昏€?16%\n- 鍖荤枟 8.7%锛堝惈鐢熻偛+澶х梾锛塡n- 澶变笟 **0.7%**\n- 宸ヤ激 **0.2%**锛堜竴绫婚闄╄涓氾級")
        st.markdown(f"- 鍚堣 鈮?**1,280 鍏?鏈?*锛堝熀鏁?000脳25.6%锛?)
        st.markdown(f"**涓◣璧峰緛鐐?*锛歿BASIC_DEDUCTION} 鍏?鏈?)

    st.divider()
    st.subheader("馃摐 鍗拌姳绋庯紙璧勯噾璐︾翱锛?)
    st.caption("鏁版嵁鏉ヨ嚜銆岎煑傦笍 骞存姤瀵煎叆銆嶏紝姝ゅ涓哄彧璇诲睍绀?)

    stamp_reg = st.session_state.get("stamp_reg_capital", 0.0)
    stamp_inc = st.session_state.get("stamp_capital_increase", 0.0)

    rc1, rc2 = st.columns(2)
    with rc1:
        st.metric("娉ㄥ唽璧勬湰瀹炵即锛堢疮璁★級", f"{stamp_reg:,.0f} 鍏?)
    with rc2:
        st.metric("鏈湡澧炶祫棰?, f"{stamp_inc:,.0f} 鍏?)

    if stamp_reg > 0 or stamp_inc > 0:
        from tax_calculator import calc_stamp_duty as _csd
        _preview = _csd(stamp_reg, stamp_inc, 0, 0, 0, 0, 0, True)
        st.caption(f"馃挕 棰勪及璧勯噾璐︾翱鍗拌姳绋庯細**{_preview['鍗拌姳绋庡悎璁★紙搴旂即锛?]:,.2f} 鍏?*锛堝噺鍗婂悗锛?)
    else:
        st.caption("馃挕 鏁版嵁涓?0锛岃鍏堝埌銆岎煑傦笍 骞存姤瀵煎叆銆嶅鍏ュ勾鎶ユ暟鎹?)

    st.divider()
    st.subheader("鈾?娈嬩繚閲戝弬鏁?)
    def_prev_employees_key = "def_prev_employees"
    if def_prev_employees_key not in st.session_state:
        st.session_state[def_prev_employees_key] = 2
    def_prev_disabled_key = "def_prev_disabled"
    if def_prev_disabled_key not in st.session_state:
        st.session_state[def_prev_disabled_key] = 0
    def_prev_avg_salary_key = "def_prev_avg_salary"
    if def_prev_avg_salary_key not in st.session_state:
        st.session_state[def_prev_avg_salary_key] = 60000.0
    def_local_avg_salary_key = "def_local_avg_salary"
    if def_local_avg_salary_key not in st.session_state:
        st.session_state[def_local_avg_salary_key] = 90000.0
    def_year_key = "def_year"
    if def_year_key not in st.session_state:
        st.session_state[def_year_key] = 2026

    st.session_state[def_prev_employees_key] = st.number_input(
        "涓婂勾鍦ㄨ亴鑱屽伐浜烘暟", min_value=0, value=st.session_state[def_prev_employees_key], step=1,
        help="涓婂勾鐢ㄤ汉鍗曚綅骞村钩鍧囧湪鑱岃亴宸ヤ汉鏁?, key="sb_prev_employees")
    st.session_state[def_prev_disabled_key] = st.number_input(
        "涓婂勾瀹夋帓娈嬬柧浜哄氨涓氫汉鏁?, min_value=0, value=st.session_state[def_prev_disabled_key], step=1,
        help="涓婂勾瀹為檯瀹夋帓鐨勬畫鐤句汉灏变笟浜烘暟锛堥渶娈嬭仈瀹℃牳纭锛?, key="sb_prev_disabled")
    st.session_state[def_prev_avg_salary_key] = st.number_input(
        "涓婂勾鑱屽伐骞村钩鍧囧伐璧勶紙鍏冿級", min_value=0.0, value=st.session_state[def_prev_avg_salary_key], step=1000.0,
        help="涓婂勾鐢ㄤ汉鍗曚綅鍦ㄨ亴鑱屽伐骞村钩鍧囧伐璧?, key="sb_prev_avg_salary")
    st.session_state[def_local_avg_salary_key] = st.number_input(
        "褰撳湴绀句細骞冲潎宸ヨ祫锛堝厓/骞达級", min_value=0.0, value=st.session_state[def_local_avg_salary_key], step=1000.0,
        help="姝︽眽2024骞寸ぞ骞冲伐璧勭害8~9涓?骞达紝鐢ㄤ簬2鍊嶅皝椤?, key="sb_local_avg_salary")
    st.session_state[def_year_key] = st.number_input(
        "鐢虫姤骞村害", min_value=2024, max_value=2030, value=st.session_state[def_year_key], step=1, key="sb_year")

    st.divider()
    st.caption("閲戣壋榫橝I绋庡姟鍔╂墜 v1.6")
    st.caption("浠呬緵鍙傝€冿紝鐢虫姤鍓嶈鏍稿疄")

# ===============================================
#  涓荤晫闈?
# ===============================================

st.title("馃挵 閲戣壋榫橝I绋庡姟鍔╂墜")
st.caption("姝︽眽閲戣壋榫欑鎶€鏈夐檺鍏徃 路 涓€绔欏紡浼佷笟绋庡姟璁＄畻涓嶢I鐢虫姤鎸囧紩")

tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs(["馃梻锔?骞存姤瀵煎叆", "馃搳 瀛ｅ害鐢虫姤", "馃挵 宸ヨ祫璁＄畻", "馃搵 鎵归噺瀵煎叆", "馃彟 绋庢缂寸撼娓呭崟", "鈾?娈嬩繚閲?, "馃搫 鐢虫姤璇存槑", "馃摉 鐢虫姤鎸囧崡"])

# ---- Tab1锛氭墜鍔ㄥ綍鍏?----
with tab3:
    st.header("鍛樺伐宸ヨ祫褰曞叆")

    # 鍒濆鍖栨寔涔呭寲瀛樺偍
    EMP_DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "鍛樺伐鏁版嵁_鑽夌.json")
    if "employees_saved" not in st.session_state:
        # 灏濊瘯浠庢枃浠舵仮澶?
        if os.path.exists(EMP_DATA_FILE):
            try:
                with open(EMP_DATA_FILE, "r", encoding="utf-8") as f:
                    st.session_state["employees_saved"] = json.load(f)
            except Exception:
                st.session_state["employees_saved"] = []
        else:
            st.session_state["employees_saved"] = []

    saved_emps = st.session_state["employees_saved"]
    saved_count = len(saved_emps)

    # 鈹€鈹€ 椤堕儴鎿嶄綔鏍?鈹€鈹€
    top_col1, top_col2, top_col3 = st.columns([2, 1, 1])
    with top_col1:
        default_n = max(saved_count, 1)
        num_emp = st.number_input("鍛樺伐浜烘暟", min_value=1, max_value=20, value=default_n, step=1, key="num_emp_tab1")
    with top_col2:
        if st.button("馃捑 淇濆瓨鑽夌", use_container_width=True):
            # 浠庡綋鍓?widget 鍊兼敹闆嗘暟鎹?
            draft = []
            for i in range(num_emp):
                draft.append({
                    "name": st.session_state.get(f"name_{i}", f"鍛樺伐{i+1}"),
                    "gross_salary": st.session_state.get(f"salary_{i}", 10522.0 if i == 0 else 8000.0),
                    "si_base": st.session_state.get(f"si_base_{i}", 5000.0),
                    "si_personal_actual": st.session_state.get(f"si_personal_{i}", float(SOCIAL_INSURANCE_ACTUAL)),
                    "special_deductions": st.session_state.get(f"special_{i}", 5000.0 if i == 0 else 0.0),
                    "child_education": st.session_state.get(f"child_{i}", 2000.0 if i == 0 else 0.0),
                    "infant_care": st.session_state.get(f"infant_{i}", 2000.0 if i == 0 else 0.0),
                    "elderly_care": st.session_state.get(f"elderly_{i}", 1000.0 if i == 0 else 0.0),
                })
            with open(EMP_DATA_FILE, "w", encoding="utf-8") as f:
                json.dump(draft, f, ensure_ascii=False, indent=2)
            st.session_state["employees_saved"] = draft
            st.toast("鉁?鑽夌宸蹭繚瀛?, icon="馃捑")
    with top_col3:
        if saved_count > 0:
            if st.button("馃搨 鍔犺浇鑽夌", use_container_width=True):
                st.toast(f"鉁?宸插姞杞?{saved_count} 鍚嶅憳宸?, icon="馃搨")
                st.rerun()
        else:
            st.button("馃搨 鏃犺崏绋?, disabled=True, use_container_width=True)

    # 鏄剧ず涓婃淇濆瓨鏃堕棿
    if saved_count > 0:
        try:
            mtime = datetime.fromtimestamp(os.path.getmtime(EMP_DATA_FILE))
            st.caption(f"馃挕 涓婃淇濆瓨锛歿mtime.strftime('%m-%d %H:%M')} 路 {saved_count} 鍚嶅憳宸?路 鍒囨崲銆屽憳宸ヤ汉鏁般€嶅悗鐐广€屽姞杞借崏绋裤€嶆仮澶?)
        except Exception:
            pass

    employees_data = []

    for i in range(num_emp):
        st.divider()
        st.subheader(f"鍛樺伐 {i+1}")

        # 榛樿鍊间紭鍏堜粠鑽夌鍙?
        default_name = saved_emps[i]["name"] if i < saved_count else f"鍛樺伐{i+1}"
        default_salary = saved_emps[i]["gross_salary"] if i < saved_count else (10522.0 if i == 0 else 8000.0)
        default_si_base = saved_emps[i]["si_base"] if i < saved_count else 5000.0
        default_si_personal = saved_emps[i]["si_personal_actual"] if i < saved_count else float(SOCIAL_INSURANCE_ACTUAL)
        default_special = saved_emps[i]["special_deductions"] if i < saved_count else (5000.0 if i == 0 else 0.0)
        default_child = saved_emps[i]["child_education"] if i < saved_count else (2000.0 if i == 0 else 0.0)
        default_infant = saved_emps[i]["infant_care"] if i < saved_count else (2000.0 if i == 0 else 0.0)
        default_elderly = saved_emps[i]["elderly_care"] if i < saved_count else (1000.0 if i == 0 else 0.0)

        c1, c2 = st.columns(2)
        with c1:
            name = st.text_input(
                "濮撳悕", value=default_name, key=f"name_{i}"
            )
            salary = st.number_input(
                "绋庡墠宸ヨ祫锛堝厓锛?, min_value=0.0,
                value=default_salary,
                step=100.0, key=f"salary_{i}"
            )
            si_base = st.number_input(
                "绀句繚缂磋垂鍩烘暟锛堝厓锛?, min_value=0.0,
                value=default_si_base, step=100.0, key=f"si_base_{i}"
            )
        with c2:
            si_personal = st.number_input(
                "涓汉绀句繚瀹炵即锛堝厓锛?, min_value=0.0,
                value=default_si_personal,
                step=10.0, key=f"si_personal_{i}"
            )
            special = st.number_input(
                "涓撻」闄勫姞鎵ｉ櫎鍚堣锛堝厓锛?, min_value=0.0,
                value=default_special,
                step=500.0, key=f"special_{i}"
            )
            st.markdown("**涓撻」闄勫姞鎵ｉ櫎鏄庣粏**")
            cc1, cc2, cc3 = st.columns(3)
            with cc1:
                child = st.number_input(
                    "瀛愬コ鏁欒偛", min_value=0.0,
                    value=default_child,
                    step=500.0, key=f"child_{i}"
                )
            with cc2:
                infant = st.number_input(
                    "濠村辜鍎跨収鎶?, min_value=0.0,
                    value=default_infant,
                    step=500.0, key=f"infant_{i}"
                )
            with cc3:
                elderly = st.number_input(
                    "璧″吇鑰佷汉", min_value=0.0,
                    value=default_elderly,
                    step=500.0, key=f"elderly_{i}"
                )

        employees_data.append({
            "name": name,
            "gross_salary": salary,
            "si_base": si_base,
            "si_personal_actual": si_personal,
            "special_deductions": special,
            "child_education": child,
            "infant_care": infant,
            "elderly_care": elderly,
        })

    if st.button("馃殌 寮€濮嬭绠?, use_container_width=True, type="primary"):
        # 鈹€鈹€ 杈撳叆鏍￠獙 鈹€鈹€
        validation_warnings = []
        for emp in employees_data:
            nm = emp["name"]
            gs = emp["gross_salary"]
            sb = emp["si_base"]
            sp = emp["si_personal_actual"]
            sd = emp["special_deductions"]
            ce = emp["child_education"]
            ic = emp["infant_care"]
            ec = emp["elderly_care"]
            if gs <= 0:
                validation_warnings.append(f"鈿狅笍 {nm}锛氱◣鍓嶅伐璧勪负 0锛岃纭鏄惁閬楁紡")
            if gs > 0 and sb <= 0:
                validation_warnings.append(f"鈿狅笍 {nm}锛氱ぞ淇濈即璐瑰熀鏁颁负 0锛岃纭")
            if gs > 0 and sb > 0 and sb < gs * 0.4:
                validation_warnings.append(f"馃挕 {nm}锛氱ぞ淇濆熀鏁帮紙{sb:.0f}锛夊亸浣庯紝閫氬父涓哄伐璧勭殑 60%~300%")
            if sd > 0 and ce + ic + ec != sd:
                if abs(ce + ic + ec - sd) > 1:
                    validation_warnings.append(f"馃挕 {nm}锛氫笓椤归檮鍔犳墸闄ゅ悎璁★紙{sd:.0f}锛変笌鏄庣粏涔嬪拰锛坽ce+ic+ec:.0f}锛変笉涓€鑷?)
            if ce > 2000:
                validation_warnings.append(f"馃挕 {nm}锛氬瓙濂虫暀鑲叉墸闄?{ce:.0f} 鍏冭秴鍑烘爣鍑嗭紙2000鍏?浜猴級锛岃鏍稿疄")
            if ic > 2000:
                validation_warnings.append(f"馃挕 {nm}锛氬┐骞煎効鐓ф姢鎵ｉ櫎 {ic:.0f} 鍏冭秴鍑烘爣鍑嗭紙2000鍏?浜猴級锛岃鏍稿疄")
            if ec > 3000:
                validation_warnings.append(f"馃挕 {nm}锛氳怠鍏昏€佷汉鎵ｉ櫎 {ec:.0f} 鍏冭秴鍑烘爣鍑嗭紙鏈€楂?000鍏冿級锛岃鏍稿疄")

        if validation_warnings:
            with st.expander(f"馃攳 鏁版嵁鏍￠獙鎻愮ず锛坽len(validation_warnings)} 鏉★級", expanded=True):
                for w in validation_warnings:
                    if w.startswith("鈿狅笍"):
                        st.warning(w)
                    else:
                        st.info(w)

        results = []
        for emp in employees_data:
            r = calc_one_employee(
                emp["name"],
                emp["gross_salary"],
                emp["si_base"],
                emp["si_personal_actual"],
                emp["special_deductions"],
                emp["child_education"],
                emp["infant_care"],
                emp["elderly_care"],
            )
            results.append(r)

        st.session_state["results"] = results
        st.success("鉁?璁＄畻瀹屾垚锛?)

        df = pd.DataFrame(results)
        st.subheader("馃搳 璁＄畻缁撴灉")
        # 鍙鏁板瓧鍒楁牸寮忓寲
        numeric_cols = df.select_dtypes(include=["float64", "int64"]).columns
        st.dataframe(
            df.style.format("{:.2f}", subset=numeric_cols),
            use_container_width=True,
        )

        # 姹囨€?
        st.subheader("馃搱 姹囨€?)
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("宸ヨ祫鎬婚", f"{df['绋庡墠宸ヨ祫'].sum():.2f} 鍏?)
        m2.metric("涓◣鎬婚", f"{df['搴旂撼绋庨'].sum():.2f} 鍏?)
        m3.metric("瀹炲彂鎬婚", f"{df['瀹炲彂宸ヨ祫'].sum():.2f} 鍏?)
        m4.metric("鍏徃鎬绘垚鏈?, f"{df['鍏徃鐢ㄤ汉鎬绘垚鏈?].sum():.2f} 鍏?)

        # 涓嬭浇 CSV + PDF
        csv_data = df.to_csv(index=False, encoding="utf-8-sig")
        dl_col1, dl_col2 = st.columns(2)
        with dl_col1:
            st.download_button(
                label="馃摜 涓嬭浇鐢虫姤搴曠锛圕SV锛?,
                data=csv_data,
                file_name=f"鐢虫姤搴曠_{datetime.now().strftime('%Y%m')}.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with dl_col2:
            summary_lines = [
                f"宸ヨ祫鎬婚: {df['绋庡墠宸ヨ祫'].sum():.2f} 鍏?,
                f"涓◣鎬婚: {df['搴旂撼绋庨'].sum():.2f} 鍏?,
                f"瀹炲彂鎬婚: {df['瀹炲彂宸ヨ祫'].sum():.2f} 鍏?,
                f"鍏徃鎬绘垚鏈? {df['鍏徃鐢ㄤ汉鎬绘垚鏈?].sum():.2f} 鍏?,
            ]
            pdf_bytes = make_pdf_with_dataframe("涓◣鐢虫姤搴曠", df, summary_lines, "")
            if pdf_bytes:
                st.download_button(
                    label="馃摜 涓嬭浇鐢虫姤搴曠锛圥DF锛?,
                    data=pdf_bytes,
                    file_name=f"鐢虫姤搴曠_{datetime.now().strftime('%Y%m')}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )

        # 鈹€鈹€ 宸ヨ祫鏁版嵁鏍￠獙锛堜笁閲嶏級 鈹€鈹€
        with st.expander("馃攷 宸ヨ祫鏁版嵁鏍￠獙锛堥摱琛屾祦姘?vs 涓◣鐢虫姤 vs 骞存姤锛?, expanded=False):
            st.caption("涓婁紶閾惰娴佹按鍜?鎴栦釜绋庣敵鎶ヨ褰曪紝涓庣郴缁熷綍鍏ュ伐璧勪氦鍙夋瘮瀵?)

            val_col1, val_col2 = st.columns(2)
            with val_col1:
                bank_file_val = st.file_uploader(
                    "涓婁紶閾惰娴佹按锛堢敤浜庢牎楠屽伐璧勬敮鍑猴級",
                    type=["csv", "xlsx", "xls"],
                    key="salary_val_bank",
                )
            with val_col2:
                tax_file_val = st.file_uploader(
                    "涓婁紶涓◣鐢虫姤璁板綍锛堢敤浜庢牎楠岀疮璁℃敹鍏ワ級",
                    type=["csv", "xlsx", "xls"],
                    key="salary_val_tax",
                )

            annual_salary_input = st.number_input(
                "骞存姤涓殑銆屽叏骞村伐璧勬€婚銆嶏紙鍏冿紝閫夊～锛?,
                min_value=0.0,
                value=0.0,
                step=1000.0,
                key="salary_val_annual",
                help="濉?鍒欒烦杩囧勾鎶ユ牎楠?,
            )

            if st.button("馃攳 寮€濮嬫牎楠?, key="run_salary_val", use_container_width=True):
                # 鍑嗗鍛樺伐鏁版嵁
                emp_list = []
                for i in range(emp_count):
                    name = st.session_state.get(f"name_{i}", "")
                    if not name:
                        continue
                    emp_list.append({
                        "name": name,
                        "gross_salary": st.session_state.get(f"salary_{i}", 0.0),
                        "si_base": st.session_state.get(f"si_base_{i}", 5000.0),
                        "si_personal_actual": st.session_state.get(f"si_personal_{i}", 522.0),
                        "special_deductions": (
                            st.session_state.get(f"child_{i}", 0.0)
                            + st.session_state.get(f"infant_{i}", 0.0)
                            + st.session_state.get(f"elderly_{i}", 0.0)
                        ),
                        "child_education": st.session_state.get(f"child_{i}", 0.0),
                        "infant_care": st.session_state.get(f"infant_{i}", 0.0),
                        "elderly_care": st.session_state.get(f"elderly_{i}", 0.0),
                    })

                bank_df = None
                tax_df = None

                if bank_file_val:
                    try:
                        if bank_file_val.name.endswith(".csv"):
                            try:
                                bank_df = pd.read_csv(bank_file_val, encoding="utf-8-sig")
                            except Exception:
                                bank_file_val.seek(0)
                                bank_df = pd.read_csv(bank_file_val, encoding="gbk")
                        else:
                            bank_df = pd.read_excel(bank_file_val)
                        st.success(f"鉁?閾惰娴佹按璇诲彇 {len(bank_df)} 鏉?)
                    except Exception as e:
                        st.error(f"閾惰娴佹按璇诲彇澶辫触锛歿e}")

                if tax_file_val:
                    try:
                        if tax_file_val.name.endswith(".csv"):
                            try:
                                tax_df = pd.read_csv(tax_file_val, encoding="utf-8-sig")
                            except Exception:
                                tax_file_val.seek(0)
                                tax_df = pd.read_csv(tax_file_val, encoding="gbk")
                        else:
                            tax_df = pd.read_excel(tax_file_val)
                        st.success(f"鉁?涓◣鐢虫姤璁板綍璇诲彇 {len(tax_df)} 鏉?)
                    except Exception as e:
                        st.error(f"涓◣鐢虫姤璁板綍璇诲彇澶辫触锛歿e}")

                with st.spinner("姝ｅ湪鏍￠獙..."):
                    val_result = validate_salary_data(
                        employees=emp_list,
                        bank_df=bank_df,
                        tax_filing_df=tax_df,
                        annual_total_salary=annual_salary_input if annual_salary_input > 0 else 0.0,
                    )

                # 灞曠ず缁撴灉
                st.divider()

                # 鏍￠獙1锛氶摱琛屾祦姘?
                if bank_df is not None:
                    st.markdown("**馃彟 鏍￠獙涓€锛氶摱琛屾祦姘?vs 绯荤粺宸ヨ祫**")
                    bm = val_result.get("bank_match")
                    if bm:
                        c1, c2, c3 = st.columns(3)
                        c1.metric("閾惰娴佹按宸ヨ祫鏀嚭", f"{bm['bank_salary_total']:,.0f} 鍏?)
                        c2.metric("绯荤粺骞村伐璧勫悎璁?, f"{bm['sys_annual_total']:,.0f} 鍏?)
                        c3.metric("宸紓", f"{bm['diff']:+,.0f} 鍏?, delta=f"{bm['diff_pct']:+.1f}%")
                        if bm["match"]:
                            st.success("鉁?閾惰娴佹按涓庣郴缁熷伐璧勪竴鑷?)
                        else:
                            st.error(f"鈿狅笍 宸紓杈冨ぇ锛佸缓璁鏌ラ摱琛屾憳瑕佸叧閿瘝鎴栫郴缁熷伐璧勫綍鍏?)
                    else:
                        st.warning('鏈湪閾惰娴佹按涓瘑鍒埌宸ヨ祫绫绘敮鍑猴紙鎽樿闇€鍚?宸ヨ祫""濂栭噾""缁╂晥"绛夊叧閿瘝锛?)
                    st.divider()

                # 鏍￠獙2锛氫釜绋庣敵鎶?
                if tax_df is not None:
                    st.markdown("**馃搵 鏍￠獙浜岋細涓◣鐢虫姤璁板綍 vs 绯荤粺宸ヨ祫**")
                    tm = val_result.get("tax_match", [])
                    if tm:
                        tm_rows = []
                        for r in tm:
                            tm_rows.append({
                                "濮撳悕": r["name"],
                                "涓◣鐢虫姤绱鏀跺叆": f"{r['tax_filing_income']:,.0f}",
                                "绯荤粺骞村伐璧?: f"{r['sys_annual']:,.0f}",
                                "宸紓": f"{r['diff']:+,.0f}",
                                "鐘舵€?: "鉁? if r["match"] else "鈿狅笍",
                            })
                        st.dataframe(pd.DataFrame(tm_rows), use_container_width=True, hide_index=True)
                    else:
                        st.warning("涓◣鐢虫姤璁板綍涓湭鎵惧埌涓庣郴缁熷憳宸ュ尮閰嶇殑濮撳悕")
                    st.divider()

                # 鏍￠獙3锛氬勾鎶ュ伐璧勬€婚
                if annual_salary_input > 0:
                    st.markdown("**馃搳 鏍￠獙涓夛細骞存姤宸ヨ祫鎬婚 vs 绯荤粺骞村伐璧勫悎璁?*")
                    am = val_result.get("annual_match", {})
                    if am:
                        c1, c2, c3 = st.columns(3)
                        c1.metric("骞存姤宸ヨ祫鎬婚", f"{am['annual_total_salary']:,.0f} 鍏?)
                        c2.metric("绯荤粺骞村伐璧勫悎璁?, f"{am['sys_annual_total']:,.0f} 鍏?)
                        c3.metric("宸紓", f"{am['diff']:+,.0f} 鍏?, delta=f"{am['diff_pct']:+.1f}%")
                        if am["match"]:
                            st.success("鉁?骞存姤宸ヨ祫鎬婚涓庣郴缁熷伐璧勪竴鑷?)
                        else:
                            st.error("鈿狅笍 骞存姤宸ヨ祫鎬婚涓庣郴缁熷勾宸ヨ祫鍚堣宸紓杈冨ぇ锛?)
                    st.divider()

                # 姹囨€昏鍛?
                warnings = val_result.get("warnings", [])
                if warnings:
                    with st.expander(f"馃搵 鏍￠獙璇存槑锛坽len(warnings)} 鏉★級", expanded=True):
                        for w in warnings:
                            if "鉁? in w:
                                st.success(w)
                            elif "鈿狅笍" in w or "宸? in w:
                                st.warning(w)
                            else:
                                st.info(w)

# ---- Tab2锛氭壒閲忓鍏?----
with tab4:
    st.header("鎵归噺瀵煎叆鍛樺伐宸ヨ祫琛?)

    # 涓婁紶璇存槑 + 涓嬭浇妯℃澘 骞舵帓
    col_info, col_template = st.columns([2, 1])
    with col_info:
        st.info(
            "璇蜂笂浼?CSV 鎴?Excel 鏂囦欢锛岄渶鍖呭惈浠ヤ笅鍒楋細\n"
            "濮撳悕, 绋庡墠宸ヨ祫, 绀句繚鍩烘暟, 涓汉绀句繚瀹炵即, 涓撻」闄勫姞鎵ｉ櫎, "
            "瀛愬コ鏁欒偛, 濠村辜鍎跨収鎶? 璧″吇鑰佷汉"
        )
    with col_template:
        # 璇诲彇鏈湴妯℃澘鏂囦欢鎻愪緵涓嬭浇
        template_csv_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "鐢虫姤搴曠妯℃澘.csv")
        if os.path.exists(template_csv_path):
            with open(template_csv_path, "rb") as f:
                csv_bytes = f.read()
            st.download_button(
                label="馃摜 涓嬭浇涓婁紶妯℃澘锛圕SV锛?,
                data=csv_bytes,
                file_name="鐢虫姤搴曠妯℃澘.csv",
                mime="text/csv",
                use_container_width=True,
            )
        # Excel 妯℃澘
        template_xlsx_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "鐢虫姤搴曠妯℃澘.xlsx")
        if os.path.exists(template_xlsx_path):
            with open(template_xlsx_path, "rb") as f:
                xlsx_bytes = f.read()
            st.download_button(
                label="馃摜 涓嬭浇涓婁紶妯℃澘锛圗xcel锛?,
                data=xlsx_bytes,
                file_name="鐢虫姤搴曠妯℃澘.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    st.divider()

    uploaded = st.file_uploader(
        "閫夋嫨鏂囦欢涓婁紶", type=["csv", "xlsx", "xls"], key="uploader"
    )

    if uploaded is not None:
        try:
            # 姣忔涓婁紶鏂版枃浠舵椂锛岄噸鏂拌鍙栧苟缂撳瓨鍒?session_state
            file_id = getattr(uploaded, "file_id", id(uploaded))
            if st.session_state.get("_uploaded_file_id") != file_id:
                if uploaded.name.endswith(".csv"):
                    try:
                        df_up = pd.read_csv(uploaded, encoding="utf-8-sig")
                    except Exception:
                        uploaded.seek(0)
                        df_up = pd.read_csv(uploaded, encoding="gbk")
                else:
                    df_up = pd.read_excel(uploaded)
                st.session_state["_uploaded_df"] = df_up
                st.session_state["_uploaded_file_id"] = file_id
            else:
                df_up = st.session_state["_uploaded_df"]

            st.write("鏂囦欢棰勮锛?)
            st.dataframe(df_up.head(), use_container_width=True)

            if st.button("馃殌 瀵煎叆骞惰绠?, key="btn_upload"):
                results = []
                for _, row in df_up.iterrows():
                    r = calc_one_employee(
                        str(row.get("濮撳悕", "鍛樺伐")),
                        float(row.get("绋庡墠宸ヨ祫") or 0),
                        float(row.get("绀句繚鍩烘暟") or 5000),
                        float(row.get("涓汉绀句繚瀹炵即") or SOCIAL_INSURANCE_ACTUAL),
                        float(row.get("涓撻」闄勫姞鎵ｉ櫎") or 0),
                        float(row.get("瀛愬コ鏁欒偛") or 0),
                        float(row.get("濠村辜鍎跨収鎶?) or 0),
                        float(row.get("璧″吇鑰佷汉") or 0),
                    )
                    results.append(r)

                st.session_state["results"] = results

                df_result = pd.DataFrame(results)
                st.success(f"鉁?璁＄畻瀹屾垚锛佸叡 {len(results)} 鍚嶅憳宸?)
                numeric_cols = df_result.select_dtypes(include=["float64", "int64"]).columns
                st.dataframe(
                    df_result.style.format("{:.2f}", subset=numeric_cols),
                    use_container_width=True,
                )

                csv_data = df_result.to_csv(index=False, encoding="utf-8-sig")
                st.download_button(
                    label="馃摜 涓嬭浇鐢虫姤搴曠锛圕SV锛?,
                    data=csv_data,
                    file_name=f"鐢虫姤搴曠_{datetime.now().strftime('%Y%m')}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
        except Exception as e:
            st.error(f"鏂囦欢璇诲彇澶辫触锛歿e}")

# ---- Tab3锛氱敵鎶ヨ鏄?----
with tab7:
    st.header("AI 鐢虫姤璇存槑")

    if "results" not in st.session_state:
        st.info("馃挕 璇峰厛鍦ㄣ€岎煉?宸ヨ祫璁＄畻銆嶉〉闈㈠綍鍏ュ憳宸ヤ俊鎭苟鐐瑰嚮銆屽紑濮嬭绠椼€嶏紝鍐嶈繑鍥炴湰椤垫煡鐪?AI 鐢熸垚鐨勭敵鎶ヨ鏄庛€?)
        st.caption("AI 灏嗘牴鎹偍鐨勫伐璧勬暟鎹嚜鍔ㄧ敓鎴愪釜绋庡拰绀句繚鐨勭敵鎶ユ搷浣滆鏄庛€?)
    else:
        results = st.session_state["results"]
        now_str = datetime.now().strftime("%Y骞?m鏈?)

        # 涓◣璇存槑
        st.subheader("馃搫 涓◣鐢虫姤璇存槑")
        with st.spinner("AI 姝ｅ湪鐢熸垚涓◣鐢虫姤璇存槑..."):
            tax_text = generate_tax_report_ai(results)
        st.text_area("涓◣鐢虫姤璇存槑", tax_text, height=400, key="tax_area")

        # 绀句繚璇存槑
        st.subheader("馃搫 绀句繚鐢虫姤璇存槑")
        with st.spinner("AI 姝ｅ湪鐢熸垚绀句繚鐢虫姤璇存槑..."):
            social_text = generate_social_report_ai(results)
        st.text_area("绀句繚鐢虫姤璇存槑", social_text, height=400, key="social_area")

        # 涓嬭浇 TXT + PDF
        full_text = tax_text + "\n\n" + "=" * 50 + "\n\n" + social_text
        dl_a1, dl_a2 = st.columns(2)
        with dl_a1:
            st.download_button(
                label="馃摜 涓嬭浇鐢虫姤璇存槑锛圱XT锛?,
                data=full_text,
                file_name=f"鐢虫姤璇存槑_{datetime.now().strftime('%Y%m')}.txt",
                mime="text/plain",
                use_container_width=True,
            )
        with dl_a2:
            pdf_bytes = make_pdf(
                f"鐢虫姤璇存槑 - {datetime.now().strftime('%Y骞?m鏈?)}",
                full_text.split("\n"),
                ""
            )
            if pdf_bytes:
                st.download_button(
                    label="馃摜 涓嬭浇鐢虫姤璇存槑锛圥DF锛?,
                    data=pdf_bytes,
                    file_name=f"鐢虫姤璇存槑_{datetime.now().strftime('%Y%m')}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )

# ===============================================
#  瀛ｅ害鐢虫姤鏁版嵁鎸佷箙鍖?
# ===============================================

QUARTER_DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "瀛ｅ害鐢虫姤鏁版嵁.json")


def load_quarter_data(year: int) -> dict:
    """鍔犺浇鏌愬勾搴︾殑瀛ｅ害鐢虫姤鏁版嵁"""
    if not os.path.exists(QUARTER_DATA_FILE):
        return {}
    try:
        with open(QUARTER_DATA_FILE, "r", encoding="utf-8") as f:
            all_data = json.load(f)
            return all_data.get(str(year), {})
    except Exception:
        return {}


def save_quarter_data(year: int, quarter: int, data: dict):
    """淇濆瓨瀛ｅ害鐢虫姤鏁版嵁"""
    if os.path.exists(QUARTER_DATA_FILE):
        with open(QUARTER_DATA_FILE, "r", encoding="utf-8") as f:
            all_data = json.load(f)
    else:
        all_data = {}

    if str(year) not in all_data:
        all_data[str(year)] = {}

    all_data[str(year)][str(quarter)] = data
    all_data[str(year)]["_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with open(QUARTER_DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)


def get_ytd_values(year: int, quarter: int) -> dict:
    """鑾峰彇鏈勾绱鍊硷紙鍩轰簬宸蹭繚瀛樼殑涓婂搴︽暟鎹級"""
    data = load_quarter_data(year)
    ytd_revenue = 0.0
    ytd_cost = 0.0
    ytd_profit = 0.0

    for q in range(1, quarter):
        if str(q) in data:
            q_data = data[str(q)]
            ytd_revenue += q_data.get("revenue", 0)
            ytd_cost += q_data.get("cost", 0)
            ytd_profit += q_data.get("period_profit", 0)

    return {
        "ytd_revenue": ytd_revenue,
        "ytd_cost": ytd_cost,
        "ytd_profit": ytd_profit,
    }


# ===============================================
#  Tab4锛氬搴︿紒涓氭墍寰楃◣鐢虫姤
# ===============================================

# ---- Tab4锛氬搴︿紒涓氭墍寰楃◣鐢虫姤 ----
with tab2:
    st.header("馃搳 浼佷笟鎵€寰楃◣瀛ｅ害棰勭即鐢虫姤")

    st.info(
        "灏忓瀷寰埄浼佷笟浼樻儬绋庣巼 5%锛?024-2027骞存斂绛栵級銆俓n"
        "绯荤粺浼氳嚜鍔ㄥ姞杞戒笂瀛ｅ害鏁版嵁锛岃绠楁湰骞寸疮璁″€笺€?
    )

    # ========== 瀛ｅ害閫夋嫨 ==========
    col_q, col_y = st.columns([1, 1])
    with col_q:
        quarter = st.selectbox("鐢虫姤瀛ｅ害", [1, 2, 3, 4], index=min(datetime.now().month // 3, 3))
    with col_y:
        year = st.number_input("骞村害", min_value=2024, max_value=2030, value=datetime.now().year)

    # ========== 鍔犺浇涓婂搴︽暟鎹?==========
    ytd = get_ytd_values(year, quarter)
    prev_saved = load_quarter_data(year)

    if quarter > 1 and str(quarter - 1) in prev_saved:
        st.success(
            f"鉁?宸插姞杞?Q{quarter-1} 鏁版嵁锛?
            f"绱鏀跺叆 {ytd['ytd_revenue']:.2f} 鍏冿紝"
            f"绱鍒╂鼎 {ytd['ytd_profit']:.2f} 鍏?
        )

    st.divider()

    # ========== 閾惰娴佹按瀵煎叆鍖哄煙 ==========
    with st.expander("馃摜 瀵煎叆閾惰娴佹按锛堣嚜鍔ㄥ～琛級", expanded=False):
        st.caption("鏀寔姘戠敓閾惰銆佸缓璁鹃摱琛岀瓑 CSV/Excel 娴佹按鏂囦欢锛岃嚜鍔ㄥ垎绫诲苟濉叆涓嬫柟琛ㄥ崟")

        bank_file = st.file_uploader(
            "涓婁紶閾惰娴佹按鏂囦欢锛堝彲澶氭涓婁紶涓嶅悓閾惰锛?,
            type=["csv", "xlsx", "xls"],
            key="bank_uploader",
            accept_multiple_files=True,
        )

        if bank_file:
            try:
                all_txns = []
                for bf in bank_file:
                    # 璇诲彇鏂囦欢
                    if bf.name.endswith(".csv"):
                        try:
                            df_bank = pd.read_csv(bf, encoding="utf-8-sig")
                        except Exception:
                            bf.seek(0)
                            df_bank = pd.read_csv(bf, encoding="gbk")
                    else:
                        df_bank = pd.read_excel(bf)

                    # 鈹€鈹€ 姘戠敓閾惰銆屾椿鏈熻处鎴锋槑缁嗐€嶆牸寮忔娴?鈹€鈹€
                    # 鍓?7琛屼负鍏冩暟鎹紙璐︽埛淇℃伅/鏌ヨ鍙傛暟锛夛紝绗?8琛屼负鐪熷疄琛ㄥご
                    if not bf.name.endswith(".csv"):
                        first_col = str(df_bank.columns[0]) if len(df_bank.columns) > 0 else ""
                        if "璐︽埛鍚嶇О" in first_col:
                            bf.seek(0)
                            df_bank = pd.read_excel(bf, skiprows=17, header=0)
                            # 璺宠繃姹囨€昏锛堝€熸柟绱/璐锋柟绱锛夊拰绌鸿
                            df_bank = df_bank[df_bank.iloc[:, 0].notna()].copy()
                            df_bank = df_bank[~df_bank.iloc[:, 0].astype(str).str.contains("绱|绗旀暟", na=False)].copy()

                    # 缁熶竴鍒楀悕锛堝父瑙侀摱琛屾牸寮忓吋瀹癸級
                    col_map = {}
                    for col in df_bank.columns:
                        col_lower = str(col).strip().lower()
                        if any(k in col_lower for k in ["鏃ユ湡", "date", "浜ゆ槗鏃ユ湡", "璁拌处鏃ユ湡", "浜ゆ槗鏃堕棿", "鏃堕棿"]):
                            col_map[col] = "浜ゆ槗鏃ユ湡"
                        elif any(k in col_lower for k in ["鎽樿", "澶囨敞", "鐢ㄩ€?, "description", "鎽樿璇存槑", "闄勮█"]):
                            col_map[col] = "鎽樿"
                        elif any(k in col_lower for k in ["鏀跺叆", "璐锋柟", "瀛樻", "credit", "瀛樺叆"]):
                            col_map[col] = "鏀跺叆閲戦"
                        elif any(k in col_lower for k in ["鏀嚭", "鍊熸柟", "鍙栨", "debit", "杞嚭"]):
                            col_map[col] = "鏀嚭閲戦"
                        elif any(k in col_lower for k in ["閲戦", "鍙戠敓棰?, "transaction"]):
                            col_map[col] = "閲戦"
                        elif any(k in col_lower for k in ["浣欓", "balance"]):
                            col_map[col] = "浣欓"
                        elif any(k in col_lower for k in ["鍊熻捶", "鏀舵敮鏂瑰悜", "绫诲瀷"]):
                            col_map[col] = "鍊熻捶鏍囪瘑"

                    df_bank = df_bank.rename(columns=col_map)

                    # 濡傛灉娌℃湁鏄庣‘鐨勬敹鍏?鏀嚭鍒楋紝灏濊瘯浠?閲戦"+"鍊熻捶鏍囪瘑"鎺ㄦ柇
                    if "閲戦" in df_bank.columns and "鍊熻捶鏍囪瘑" in df_bank.columns:
                        for _, row in df_bank.iterrows():
                            amount = abs(float(row.get("閲戦") or 0))
                            flag = str(row.get("鍊熻捶鏍囪瘑", "")).strip()
                            txn = {
                                "閾惰": bf.name,
                                "鏃ユ湡": row.get("浜ゆ槗鏃ユ湡", ""),
                                "鎽樿": row.get("鎽樿", ""),
                                "鏀跺叆閲戦": amount if flag in ["璐?, "鏀跺叆", "瀛樺叆", "CREDIT"] else 0,
                                "鏀嚭閲戦": amount if flag in ["鍊?, "鏀嚭", "杞嚭", "DEBIT"] else 0,
                            }
                            all_txns.append(txn)
                    else:
                        # 鐩存帴鍙栨敹鍏?鏀嚭鍒?
                        for _, row in df_bank.iterrows():
                            all_txns.append({
                                "閾惰": bf.name,
                                "鏃ユ湡": row.get("浜ゆ槗鏃ユ湡", row.get("鏃ユ湡", "")),
                                "鎽樿": row.get("鎽樿", ""),
                                "鏀跺叆閲戦": float(row.get("鏀跺叆閲戦", 0) or 0),
                                "鏀嚭閲戦": float(row.get("鏀嚭閲戦", 0) or 0),
                            })

                df_txns = pd.DataFrame(all_txns)
                st.success(f"鉁?鎴愬姛璇诲彇 {len(df_txns)} 鏉′氦鏄撹褰?)

                # 浣跨敤浼樺寲鍚庣殑鍒嗙被鍑芥暟锛堥伒寰皬浼佷笟浼氳鍑嗗垯锛?
                df_txns["鑷姩鍒嗙被"] = df_txns["鎽樿"].apply(
                    lambda x: classify_bank_transaction(x)["category"]
                )
                df_txns["浼氳绉戠洰"] = df_txns["鎽樿"].apply(
                    lambda x: classify_bank_transaction(x)["account"]
                )
                df_txns["鍒╂鼎琛ㄩ」鐩?] = df_txns["鎽樿"].apply(
                    lambda x: classify_bank_transaction(x)["pl_item"]
                )
                
                st.dataframe(
                    df_txns[["鏃ユ湡", "鎽樿", "鏀跺叆閲戦", "鏀嚭閲戦", "鑷姩鍒嗙被", "鍒╂鼎琛ㄩ」鐩?]].head(10), 
                    use_container_width=True
                )

                st.subheader("璇风‘璁や氦鏄撳垎绫伙紙鍙墜鍔ㄤ慨鏀癸級")
                edited_df = st.data_editor(
                    df_txns[["鏃ユ湡", "鎽樿", "鏀跺叆閲戦", "鏀嚭閲戦", "鑷姩鍒嗙被", "鍒╂鼎琛ㄩ」鐩?]],
                    use_container_width=True,
                    num_rows="dynamic",
                    key="txn_editor",
                )

                # 鐢熸垚鍒╂鼎琛ㄩ瑙?
                st.subheader("馃搳 鑷姩鐢熸垚鍒╂鼎琛紙灏忎紒涓氫細璁″噯鍒欙級")
                profit_data = generate_profit_statement(edited_df)
                
                profit_df = pd.DataFrame({
                    "鍒╂鼎琛ㄩ」鐩?: [
                        "涓€銆佽惀涓氭敹鍏ワ紙绗?琛岋級",
                        "鍑忥細钀ヤ笟鎴愭湰锛堢2琛岋級",
                        "    绋庨噾鍙婇檮鍔狅紙绗?琛岋級",
                        "    绠＄悊璐圭敤锛堢5琛岋級",
                        "    璐㈠姟璐圭敤锛堢6琛岋級",
                        "    璧勪骇鍑忓€兼崯澶憋紙绗?琛岋級",
                        "鍔狅細鎶曡祫鏀剁泭锛堢8琛岋級",
                        "浜屻€佽惀涓氬埄娑︼紙绗?琛岋級",
                        "鍔狅細钀ヤ笟澶栨敹鍏ワ紙绗?0琛岋級",
                        "鍑忥細钀ヤ笟澶栨敮鍑猴紙绗?1琛岋級",
                        "涓夈€佸埄娑︽€婚锛堢12琛岋級",
                        "鍑忥細鎵€寰楃◣璐圭敤锛堢13琛岋級",
                        "鍥涖€佸噣鍒╂鼎锛堢14琛岋級",
                    ],
                    "鏈湡閲戦": [
                        f"{profit_data['钀ヤ笟鏀跺叆']:,.2f}",
                        f"{profit_data['钀ヤ笟鎴愭湰']:,.2f}",
                        f"{profit_data['绋庨噾鍙婇檮鍔?]:,.2f}",
                        f"{profit_data['绠＄悊璐圭敤']:,.2f}",
                        f"{profit_data['璐㈠姟璐圭敤']:,.2f}",
                        "0.00",
                        f"{profit_data['鎶曡祫鏀剁泭']:,.2f}",
                        f"{profit_data['钀ヤ笟鍒╂鼎']:,.2f}",
                        f"{profit_data['钀ヤ笟澶栨敹鍏?]:,.2f}",
                        f"{profit_data['钀ヤ笟澶栨敮鍑?]:,.2f}",
                        f"{profit_data['鍒╂鼎鎬婚']:,.2f}",
                        f"{profit_data['鎵€寰楃◣璐圭敤']:,.2f}",
                        f"{profit_data['鍑€鍒╂鼎']:,.2f}",
                    ]
                })
                st.dataframe(profit_df, use_container_width=True, hide_index=True)
                st.caption("馃挕 鍒╂鼎琛ㄦ牴鎹€婂皬浼佷笟浼氳鍑嗗垯銆嬬敓鎴愶紝鍙笌鐢虫姤琛ㄧ1~3琛屼氦鍙夋牎楠?)

                # 璁＄畻姹囨€伙紙鐢ㄤ簬濉叆鐢虫姤琛級
                revenue_total = profit_data["钀ヤ笟鏀跺叆"]
                cost_total = profit_data["钀ヤ笟鎴愭湰"]
                profit = profit_data["鍒╂鼎鎬婚"]

                expense_total = profit_data["绠＄悊璐圭敤"]
                st.subheader("馃搱 鑷姩姹囨€荤粨鏋滐紙鏈湡锛?)
                col_a, col_b, col_c, col_d = st.columns(4)
                col_a.metric("钀ヤ笟鏀跺叆", f"{revenue_total:.2f}")
                col_b.metric("钀ヤ笟鎴愭湰", f"{cost_total:.2f}")
                col_c.metric("绠＄悊璐圭敤", f"{expense_total:.2f}")
                col_d.metric("鍒╂鼎鎬婚", f"{profit:.2f}")

                if st.button("鉁?纭骞跺～鍏ョ敵鎶ヨ〃", use_container_width=True, type="primary", key="btn_fill_quarter"):
                    st.session_state["auto_revenue"] = revenue_total
                    st.session_state["auto_cost"] = cost_total
                    st.session_state["auto_profit"] = profit
                    st.session_state["profit_data"] = profit_data  # 淇濆瓨鍒╂鼎琛ㄦ暟鎹?
                    st.success("鉁?宸茶嚜鍔ㄥ～鍏ョ敵鎶ヨ〃锛岃鍚戜笅婊氬姩纭鏁版嵁锛?)
                    st.rerun()

            except Exception as e:
                st.error(f"閾惰娴佹按瑙ｆ瀽澶辫触锛歿e}")
                st.caption("璇风‘淇濇枃浠跺寘鍚細鏃ユ湡銆佹憳瑕併€佹敹鍏ラ噾棰濄€佹敮鍑洪噾棰?绛夊垪")

    st.divider()

    # ========== 鎵嬪姩杈撳叆鍖哄煙 ==========
    c1, c2 = st.columns(2)
    with c1:
        st.subheader("鏈湡鏁帮紙Q" + str(quarter) + "锛?)

        # 鑷姩濉叆锛堝鏋滈摱琛屾祦姘村鍏ヤ簡锛?
        rev_val = st.session_state.get("auto_revenue", 0.0)
        revenue = st.number_input("瀛ｅ害钀ヤ笟鏀跺叆锛堝厓锛?, min_value=0.0, value=rev_val, step=1000.0, key="q_revenue")

        cost_val = st.session_state.get("auto_cost", 0.0)
        cost = st.number_input("瀛ｅ害钀ヤ笟鎴愭湰锛堝厓锛?, min_value=0.0, value=cost_val, step=1000.0, key="q_cost")

    with c2:
        st.subheader("鏈湡鍒╂鼎鍙婁紒涓氫俊鎭?)

        profit_val = st.session_state.get("auto_profit", 0.0)
        period_profit = st.number_input("瀛ｅ害鍒╂鼎鎬婚锛堝厓锛?, value=profit_val, step=1000.0, key="q_profit")

        num_employees = st.number_input("瀛ｅ害骞冲潎浠庝笟浜烘暟", min_value=1, value=1, step=1)
        total_assets = st.number_input("瀛ｅ害骞冲潎璧勪骇鎬婚锛堜竾鍏冿級", min_value=0.0, value=0.0, step=10.0)

    # ========== 绱鏁帮紙鑷姩璁＄畻锛?=========
    st.divider()
    st.subheader("馃搱 绱鏁帮紙鑷姩璁＄畻锛?)

    ytd_revenue = ytd["ytd_revenue"] + revenue
    ytd_cost = ytd["ytd_cost"] + cost
    ytd_profit = ytd["ytd_profit"] + period_profit

    col_y1, col_y2, col_y3 = st.columns(3)
    col_y1.metric("鏈勾绱钀ヤ笟鏀跺叆", f"{ytd_revenue:.2f} 鍏?)
    col_y2.metric("鏈勾绱钀ヤ笟鎴愭湰", f"{ytd_cost:.2f} 鍏?)
    col_y3.metric("鏈勾绱鍒╂鼎鎬婚", f"{ytd_profit:.2f} 鍏?)

    # ========== 澧炲€肩◣鍙婇檮鍔犵◣娴嬬畻 ==========
    st.divider()
    st.subheader("馃Ь 澧炲€肩◣鍙婇檮鍔犵◣娴嬬畻")
    st.caption("渚濇嵁2026骞存箹鍖楃渷浼樻儬鏀跨瓥锛氬皬瑙勬ā绾崇◣浜哄噺鎸?% + 鍏◣涓よ垂鍑忓崐")

    vat_col1, vat_col2 = st.columns([2, 1])
    with vat_col1:
        is_small_scale = st.radio(
            "绾崇◣浜虹被鍨?,
            ["灏忚妯＄撼绋庝汉锛?%锛?, "涓€鑸撼绋庝汉锛堟寜瀹為檯绋庣巼锛?],
            index=0,
            horizontal=True,
        ) == "灏忚妯＄撼绋庝汉锛?%锛?
        vat_revenue_input = st.number_input(
            "瀛ｅ害鍚◣钀ヤ笟鏀跺叆锛堝厓锛岀敤浜庤绠楀鍊肩◣锛?,
            min_value=0.0,
            value=float(st.session_state.get("auto_revenue", revenue)),
            step=1000.0,
            key="vat_revenue",
            help="灏忚妯＄撼绋庝汉锛氬搴︿笉鍚◣鏀跺叆 鈮?30涓囧厓鍙厤寰佸鍊肩◣",
        )
    with vat_col2:
        st.markdown("**姝︽眽闄勫姞绋庣巼锛堝叚绋庝袱璐瑰噺鍗婂悗锛?*")
        st.markdown("- 鍩庡缓绋庯細**3.5%**锛堝師7%脳50%锛?)
        st.markdown("- 鏁欒偛璐归檮鍔狅細**1.5%**锛堝師3%脳50%锛?)
        st.markdown("- 鍦版柟鏁欒偛闄勫姞锛?*1%**锛堝師2%脳50%锛?)
        st.markdown("- 鍚堣锛?*6%** 脳 澧炲€肩◣")

    # 瀹炴椂棰勭畻澧炲€肩◣
    vat_preview = calc_vat_and_surcharge(
        revenue=vat_revenue_input,
        vat_rate=0.03,
        is_small_scale=is_small_scale,
        is_small_low_profit=True,
    )

    vc1, vc2, vc3, vc4 = st.columns(4)
    vc1.metric(
        "澧炲€肩◣",
        f"{vat_preview['澧炲€肩◣搴旂即']:,.2f} 鍏?,
        delta="鍏嶇◣" if vat_preview['澧炲€肩◣搴旂即'] == 0 else None,
        delta_color="normal",
    )
    vc2.metric("鍩庡缓绋?, f"{vat_preview['鍩庡缓绋?7%)']:,.2f} 鍏?)
    vc3.metric("鏁欒偛璐归檮鍔?, f"{vat_preview['鏁欒偛璐归檮鍔?3%)'] + vat_preview['鍦版柟鏁欒偛闄勫姞(2%)']:,.2f} 鍏?)
    vc4.metric("闄勫姞绋庡悎璁?, f"{vat_preview['闄勫姞绋庡悎璁?]:,.2f} 鍏?)

    if vat_preview["澧炲€肩◣搴旂即"] == 0:
        st.success(f"鉁?{vat_preview['澧炲€肩◣鍏嶇◣璇存槑']}锛岄檮鍔犵◣鍚屾涓洪浂銆?)
    else:
        st.info(f"鈩癸笍 {vat_preview['澧炲€肩◣鍏嶇◣璇存槑']}")

    # ========== 璁＄畻鎸夐挳 ==========
    st.divider()

    if st.button("馃殌 璁＄畻瀛ｅ害棰勭即绋庨锛堝惈澧炲€肩◣鍙婇檮鍔狅級", use_container_width=True, type="primary"):
        # 鈹€鈹€ 杈撳叆鏍￠獙 鈹€鈹€
        q_warnings = []
        if revenue <= 0 and cost <= 0:
            q_warnings.append("鈿狅笍 钀ヤ笟鏀跺叆鍜岃惀涓氭垚鏈潎涓?0锛岃纭鏈搴︽槸鍚︽湁缁忚惀娲诲姩")
        if revenue > 0 and cost <= 0:
            q_warnings.append("馃挕 鏈夎惀涓氭敹鍏ヤ絾鏃犺惀涓氭垚鏈紝鏈嶅姟绫讳紒涓氬彲鑳芥甯革紝璇锋牳瀹?)
        if vat_revenue_input <= 0:
            q_warnings.append("馃挕 澧炲€肩◣璁＄畻鏀跺叆涓?0锛屽鏈鏃犳敹鍏ワ紝澧炲€肩◣鍙浂鐢虫姤")
        if num_employees <= 0:
            q_warnings.append("鈿狅笍 浠庝笟浜烘暟涓?0锛岃濉啓瀹為檯浜烘暟")
        if period_profit != 0 and abs(period_profit - (revenue - cost)) > revenue * 0.5:
            q_warnings.append("馃挕 鍒╂鼎鎬婚涓庛€屾敹鍏?鎴愭湰銆嶅樊寮傝緝澶э紝璇风‘璁ゆ槸鍚﹀凡璁″叆绠＄悊璐圭敤绛夋湡闂磋垂鐢?)
        if q_warnings:
            with st.expander(f"馃攳 鏁版嵁鏍￠獙鎻愮ず锛坽len(q_warnings)} 鏉★級", expanded=True):
                for w in q_warnings:
                    if w.startswith("鈿狅笍"):
                        st.warning(w)
                    else:
                        st.info(w)

        # 璁＄畻鏈勾绱宸查缂寸◣棰濓紙绗?2琛岋級
        tax_paid_ytd = 0.0
        for q in range(1, quarter):
            if str(q) in prev_saved:
                tax_paid_ytd += prev_saved[str(q)].get("tax_payable", 0)

        # 澧炲€肩◣鍙婇檮鍔犵◣璁＄畻
        vat_data = calc_vat_and_surcharge(
            revenue=vat_revenue_input,
            vat_rate=0.03,
            is_small_scale=is_small_scale,
            is_small_low_profit=True,
        )
        st.session_state["vat_data"] = vat_data

        result = calc_corporate_income_tax_quarterly(
            revenue, cost, period_profit, ytd_profit,
            int(num_employees), total_assets,
            tax_paid_ytd=tax_paid_ytd,
            vat_data=vat_data,
        )
        st.session_state["corp_tax_result"] = result

        # 淇濆瓨鏈湡鏁版嵁
        save_quarter_data(year, quarter, {
            "revenue": revenue,
            "cost": cost,
            "period_profit": period_profit,
            "ytd_revenue": ytd_revenue,
            "ytd_cost": ytd_cost,
            "ytd_profit": ytd_profit,
            "num_employees": int(num_employees),
            "total_assets": total_assets,
            "tax_payable": result["鏈湡搴旂撼绋庨"],
            "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })

        # 淇濆瓨骞翠唤/瀛ｅ害渚?Tab5 浣跨敤
        st.session_state["_tax_year"] = year
        st.session_state["_tax_quarter"] = quarter

        st.success(f"鉁?璁＄畻瀹屾垚锛丵{quarter} 鏁版嵁宸蹭繚瀛橈紝涓嬫鐢虫姤 Q{quarter+1} 鏃朵細鑷姩鍔犺浇銆傝鍒囨崲鍒般€岎煆?绋庢缂寸撼娓呭崟銆嶆煡鐪嬫眹鎬汇€?)

    # ========== 璁＄畻缁撴灉灞曠ず锛堝尮閰?A200000 鐢虫姤琛ㄦ牸寮忥級==========
    if "corp_tax_result" in st.session_state:
        r = st.session_state["corp_tax_result"]

        st.subheader("馃搵 鐢虫姤琛ㄩ瑙堬紙A200000鏍煎紡锛?)

        # 绗竴鍖哄煙锛氭敹鍏ユ垚鏈埄娑︼紙绗?~3琛岋級
        st.markdown("**绗竴閮ㄥ垎锛氭敹鍏ャ€佹垚鏈€佸埄娑?*")
        form_df_1 = pd.DataFrame({
            "琛屾": ["绗?琛?, "绗?琛?, "绗?琛?],
            "椤圭洰": ["钀ヤ笟鏀跺叆", "钀ヤ笟鎴愭湰", "鍒╂鼎鎬婚"],
            "鏈湡閲戦": [f"{r['钀ヤ笟鏀跺叆']:,.2f}", f"{r['钀ヤ笟鎴愭湰']:,.2f}", f"{r['鍒╂鼎鎬婚']:,.2f}"],
            "绱閲戦": [f"{ytd_revenue:,.2f}", f"{ytd_cost:,.2f}", f"{ytd_profit:,.2f}"],
        })
        st.dataframe(form_df_1, use_container_width=True, hide_index=True)

        # 绗簩鍖哄煙锛氬簲绾崇◣鎵€寰楅璁＄畻锛堢4~8琛岋級
        st.markdown("**绗簩閮ㄥ垎锛氬簲绾崇◣鎵€寰楅璁＄畻**")
        form_df_2 = pd.DataFrame({
            "琛屾": ["绗?琛?, "绗?琛?, "绗?琛?, "绗?琛?, "绗?琛?],
            "椤圭洰": ["鐗瑰畾涓氬姟璋冩暣", "涓嶅緛绋庢敹鍏?, "鍥哄畾璧勪骇鎶樻棫璋冩暣", "寮ヨˉ浠ュ墠骞村害浜忔崯", "瀹為檯鍒╂鼎棰?],
            "鏈湡閲戦": ["0.00", "0.00", "0.00", "0.00", f"{r['瀹為檯鍒╂鼎棰?]:,.2f}"],
            "绱閲戦": ["0.00", "0.00", "0.00", "0.00", f"{ytd_profit:,.2f}"],
        })
        st.dataframe(form_df_2, use_container_width=True, hide_index=True)

        # 绗笁鍖哄煙锛氱◣娆捐绠楋紙绗?~13琛岋級
        st.markdown("**绗笁閮ㄥ垎锛氱◣娆捐绠?*")
        form_df_3 = pd.DataFrame({
            "琛屾": ["绗?琛?, "绗?0琛?, "绗?1琛?, "绗?2琛?, "绗?3琛?],
            "椤圭洰": ["绋庣巼(25%)", "搴旂撼鎵€寰楃◣棰?, "鍑忓厤鎵€寰楃◣棰?, "鏈勾绱宸查缂?, "鏈湡搴旇ˉ(閫€)绋庨"],
            "鏈湡閲戦": [
                "25%",
                f"{r['搴旂撼绋庨_鏍囧噯']:,.2f}",
                f"{r['鍑忓厤鎵€寰楃◣棰?]:,.2f}",
                f"{r['鏈勾绱宸查缂?]:,.2f}",
                f"{r['鏈湡搴旇ˉ(閫€)绋庨']:,.2f}",
            ],
        })
        st.dataframe(form_df_3, use_container_width=True, hide_index=True)

        # ===== 鏁版嵁鏍￠獙锛堝埄娑﹁〃 vs 鐢虫姤琛級=====
        if "profit_data" in st.session_state:
            profit_data = st.session_state["profit_data"]
            validation = validate_quarterly_declaration(
                profit_data, 
                r["钀ヤ笟鏀跺叆"], 
                r["钀ヤ笟鎴愭湰"], 
                r["鍒╂鼎鎬婚"]
            )
            
            st.subheader("馃搵 鏁版嵁鏍￠獙缁撴灉")
            all_pass = True
            for passed, msg in validation:
                if passed:
                    st.success(f"鉁?{msg}")
                else:
                    st.error(f"鈿狅笍 {msg}")
                    all_pass = False
            
            if all_pass:
                st.success("馃帀 鎵€鏈夋牎楠岄€氳繃锛佸埄娑﹁〃涓庣敵鎶ヨ〃鏁版嵁涓€鑷淬€?)
            else:
                st.warning("鈿狅笍 璇锋鏌ラ摱琛屾祦姘村垎绫绘槸鍚︽纭紝鎴栨墜鍔ㄨ皟鏁寸敵鎶ヨ〃鏁版嵁銆?)

        # 鍏抽敭鎸囨爣鍗＄墖
        st.subheader("馃搳 鍏抽敭鎸囨爣")
        k1, k2, k3 = st.columns(3)
        k1.metric("瀹為檯鍒╂鼎棰?, f"{r['瀹為檯鍒╂鼎棰?]:,.2f} 鍏?)
        k2.metric("浼佷笟鎵€寰楃◣锛堟湰鏈熷簲琛ョ即锛?, f"{r['鏈湡搴旇ˉ(閫€)绋庨']:,.2f} 鍏?)
        k3.metric("鏈湡绋庤垂鍚堣", f"{r.get('鏈湡绋庤垂鍚堣', r['鏈湡搴旇ˉ(閫€)绋庨']):,.2f} 鍏?)

        # ===== 绋庤垂姹囨€昏〃 =====
        st.subheader("馃挵 鏈湡绋庤垂姹囨€绘祴绠?)
        vat_d = st.session_state.get("vat_data", {})
        tax_summary = pd.DataFrame({
            "绋庣": [
                "鈶?浼佷笟鎵€寰楃◣锛圓200000鐢虫姤锛?,
                "鈶?澧炲€肩◣锛堝惈绋庢敹鍏ョ敵鎶ワ級",
                "鈶?鍩庡缓绋庯紙澧炲€肩◣脳7%锛?,
                "鈶?鏁欒偛璐归檮鍔狅紙澧炲€肩◣脳3%锛?,
                "鈶?鍦版柟鏁欒偛闄勫姞锛堝鍊肩◣脳2%锛?,
                "鍚堣搴旂即绋庤垂",
            ],
            "鏈湡搴旂即锛堝厓锛?: [
                f"{r['鏈湡搴旇ˉ(閫€)绋庨']:,.2f}",
                f"{vat_d.get('澧炲€肩◣搴旂即', 0.0):,.2f}",
                f"{vat_d.get('鍩庡缓绋?7%)', 0.0):,.2f}",
                f"{vat_d.get('鏁欒偛璐归檮鍔?3%)', 0.0):,.2f}",
                f"{vat_d.get('鍦版柟鏁欒偛闄勫姞(2%)', 0.0):,.2f}",
                f"{r.get('鏈湡绋庤垂鍚堣', r['鏈湡搴旇ˉ(閫€)绋庨']):,.2f}",
            ],
            "璁＄◣渚濇嵁": [
                f"鍒╂鼎鎬婚 {r['鍒╂鼎鎬婚']:,.2f} 脳 {'5%锛堝皬寰紭鎯狅級' if r['鏄惁灏忓瀷寰埄浼佷笟']=='鏄? else '25%'}",
                vat_d.get("澧炲€肩◣鍏嶇◣璇存槑", "-"),
                f"澧炲€肩◣ {vat_d.get('澧炲€肩◣搴旂即', 0.0):,.2f} 脳 7%",
                f"澧炲€肩◣ {vat_d.get('澧炲€肩◣搴旂即', 0.0):,.2f} 脳 3%",
                f"澧炲€肩◣ {vat_d.get('澧炲€肩◣搴旂即', 0.0):,.2f} 脳 2%",
                "鈶?+ 鈶?+ 鈶?+ 鈶?+ 鈶?,
            ],
        })
        st.dataframe(tax_summary, use_container_width=True, hide_index=True)

        # 鍒ゆ柇鐘舵€?
        if r['鍒╂鼎鎬婚'] <= 0:
            st.info("馃搶 鏈湡浜忔崯锛屾棤闇€缂寸撼浼佷笟鎵€寰楃◣銆傚噺鍏嶆墍寰楃◣棰濆拰搴旇ˉ閫€绋庨鍧囦负0銆?)
        else:
            if r['鏄惁灏忓瀷寰埄浼佷笟'] == '鏄?:
                st.success(f"馃搶 灏忓瀷寰埄浼佷笟浼樻儬宸茬敓鏁堬細瀹為檯绋庤礋浠?%锛堟爣鍑?5%锛屽噺鍏峽r['鍑忓厤鎵€寰楃◣棰?]:,.2f}鍏冿級")

        # AI 鐢虫姤璇存槑
        st.subheader("馃搫 鐢虫姤璇存槑")
        vat_d_report = st.session_state.get("vat_data", None)
        report_text = format_corporate_tax_report(r, quarter, year, vat_data=vat_d_report)
        st.text_area("鐢虫姤璇存槑", report_text, height=400, key="corp_tax_area")

        # 涓嬭浇
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.download_button(
                label="馃摜 涓嬭浇鐢虫姤璇存槑锛圱XT锛?,
                data=report_text,
                file_name=f"浼佷笟鎵€寰楃◣棰勭即鐢虫姤_{year}Q{quarter}.txt",
                mime="text/plain",
                use_container_width=True,
            )
        with col_dl2:
            csv_corp = pd.DataFrame([r]).to_csv(index=False, encoding="utf-8-sig")
            st.download_button(
                label="馃摜 涓嬭浇鐢虫姤搴曠锛圕SV锛?,
                data=csv_corp,
                file_name=f"浼佷笟鎵€寰楃◣棰勭即鐢虫姤_{year}Q{quarter}.csv",
                mime="text/csv",
                use_container_width=True,
            )

# ===============================================
#  Tab5锛氱◣娆剧即绾虫竻鍗?
# ===============================================

with tab5:
    st.header("馃彟 绋庢缂寸撼娓呭崟 & 浼樻儬鏀跨瓥閫傜敤璇存槑")

    # 璇诲彇璁＄畻缁撴灉
    r5 = st.session_state.get("corp_tax_result", None)
    vat5 = st.session_state.get("vat_data", None)

    if r5 is None:
        st.info("馃挕 璇峰厛鍦ㄣ€岎煋?瀛ｅ害鐢虫姤銆嶉〉闈㈠～鍐欐暟鎹苟鐐瑰嚮銆岃绠楀搴﹂缂寸◣棰濄€嶏紝瀹屾垚鍚庤繑鍥炴湰椤垫煡鐪嬬◣娆剧即绾虫竻鍗曘€?, icon="鈩癸笍")
    if r5 is not None:
        # 骞翠唤/瀛ｅ害
        all_qdata = {}
        if os.path.exists(QUARTER_DATA_FILE):
            with open(QUARTER_DATA_FILE, "r", encoding="utf-8") as _f:
                all_qdata = json.load(_f)

        latest_year = max(all_qdata.keys(), default=str(datetime.now().year))
        latest_qdata_key = str(st.session_state.get("_tax_year", int(latest_year)))
        disp_year = st.session_state.get("_tax_year", int(latest_year))
        disp_quarter = st.session_state.get("_tax_quarter", datetime.now().month // 3 or 1)

        # ===== 宸ヨ祫涓◣鏁版嵁 =====
        payroll_results = st.session_state.get("results", [])
        total_personal_tax = sum(float(e.get("搴旂撼绋庨", 0)) for e in payroll_results)
        total_company_social = sum(float(e.get("鍏徃绀句繚", 0)) for e in payroll_results)
        total_personal_social = sum(float(e.get("涓汉绀句繚", 0)) for e in payroll_results)
        employee_count = len(payroll_results)

        # ===== 澧炲€肩◣鍙婇檮鍔?=====
        vat_amount = vat5.get("澧炲€肩◣搴旂即", 0.0) if vat5 else 0.0
        urban_tax = vat5.get("鍩庡缓绋?7%)", 0.0) if vat5 else 0.0
        edu_surcharge = vat5.get("鏁欒偛璐归檮鍔?3%)", 0.0) if vat5 else 0.0
        local_edu = vat5.get("鍦版柟鏁欒偛闄勫姞(2%)", 0.0) if vat5 else 0.0
        surcharge_total = vat5.get("闄勫姞绋庡悎璁?, 0.0) if vat5 else 0.0
        vat_note = vat5.get("澧炲€肩◣鍏嶇◣璇存槑", "-") if vat5 else "鏈绠?
        six_two_relief = vat5.get("鍏◣涓よ垂鍑忓厤閲戦", 0.0) if vat5 else 0.0
        vat_policy = vat5.get("澧炲€肩◣浼樻儬渚濇嵁", "-") if vat5 else "-"
        surcharge_policy = vat5.get("闄勫姞绋庝紭鎯犺鏄?, "") if vat5 else ""

        # ===== 浼佷笟鎵€寰楃◣ =====
        corp_tax = r5.get("鏈湡搴旇ˉ(閫€)绋庨", 0.0) if r5 else 0.0
        corp_relief = r5.get("鍑忓厤鎵€寰楃◣棰?, 0.0) if r5 else 0.0
        is_small = (r5.get("鏄惁灏忓瀷寰埄浼佷笟", "鍚?) == "鏄?) if r5 else True

        # ===== 鍗拌姳绋?=====
        # 璧勯噾璐︾翱锛氭敞鍐岃祫鏈埌浣?/ 澧炶祫
        stamp_reg_capital = st.session_state.get("stamp_reg_capital", 0.0)
        stamp_capital_increase = st.session_state.get("stamp_capital_increase", 0.0)
        # 璐攢鍚堝悓锛氭寜褰撳鏀跺叆浼扮畻锛堝彲瑙嗘敹鍏ヤ负鍚◣璐攢棰濓級
        stamp_purchase = vat5.get("瀛ｅ害鍚◣鏀跺叆", 0.0) if vat5 else 0.0
        stamp_data = calc_stamp_duty(
            registered_capital=stamp_reg_capital,
            capital_increase=stamp_capital_increase,
            purchase_amount=stamp_purchase,
            is_small_low_profit=is_small,
        )
        stamp_total = stamp_data["鍗拌姳绋庡悎璁★紙搴旂即锛?]
        stamp_nominal = stamp_data["鍗拌姳绋庡悎璁★紙鍚嶄箟锛?]
        stamp_relief = stamp_data["鍏◣涓よ垂鍑忓厤"]

        # ===== 鍚堣锛堝惈鍗拌姳绋庯級=====
        total_tax = round(total_personal_tax + vat_amount + surcharge_total + corp_tax + stamp_total, 2)

        # ===== 1. 浼樻儬鏀跨瓥閫傜敤娓呭崟锛堝厛灞曠ず锛?=====
        st.subheader(f"馃搵 {disp_year}骞寸{disp_quarter}瀛ｅ害 鈥?浼樻儬绋庣巼閫傜敤璇存槑")
        st.caption(f"鏀跨瓥渚濇嵁锛氭箹鍖楃渷銆婂叧浜庡姞鍔涘姪浼佽В闅炬帹鍔ㄤ腑灏忎紒涓氱ǔ鍋ュ彂灞曠殑鑻ュ共鎺柦銆嬶紙鏈夋晥鏈熻嚦2027.12.31锛?)

        # 鑾峰彇鏀跨瓥姹囨€?
        revenue_used = st.session_state.get("vat_revenue", r5.get("钀ヤ笟鏀跺叆", 0.0) if r5 else 0.0)
        emp_used = r5.get("浠庝笟浜烘暟", 1) if r5 else 1
        asset_used = r5.get("璧勪骇鎬婚_涓囧厓", 0.0) if r5 else 0.0

        policy_summary = get_tax_policy_summary(
            is_small_scale=(vat5.get("鏄惁灏忚妯＄撼绋庝汉", "鏄?) == "鏄?) if vat5 else True,
            is_small_low_profit=is_small,
            quarter_revenue=revenue_used,
            num_employees=emp_used,
            total_assets=asset_used,
            quarter=disp_quarter,
        )

        for i, p in enumerate(policy_summary["policies"], 1):
            with st.expander(f"馃幆 {p['绋庣']} 鈥?{p['浼樻儬鍚嶇О']}", expanded=(i == 1)):
                cols = st.columns([2, 1])
                with cols[0]:
                    st.markdown(f"""
    - **浼樻儬鍐呭**锛歿p['浼樻儬鍐呭']}
    - **鏀跨瓥渚濇嵁**锛歿p['鏀跨瓥渚濇嵁']}
    - **閫傜敤鏉′欢**锛歿p['閫傜敤鏉′欢']}
                    """)
                with cols[1]:
                    st.metric("浼樻儬鍔涘害", p['浼樻儬鍔涘害'])

        st.success(f"鉁?{policy_summary['tip']}")
        st.info(f"馃搮 {policy_summary['valid_until']}")

        st.divider()

        # ===== 2. 绋庢缂寸撼娓呭崟锛堝惈浼樻儬鏍囨敞锛?=====
        st.subheader("馃搵 绋庢鏄庣粏娓呭崟锛堝惈浼樻儬绋庣巼鏍囨敞锛?)

        # 纭畾澧炲€肩◣鐨勬爣绉板€煎拰浼樻儬璇存槑
        vat_display_rate = "0%锛堝厤绋庯級" if vat_amount == 0 else "1%锛堜紭鎯犲悗锛屽師3%锛?
        if not (vat5 and vat5.get("鏄惁灏忚妯＄撼绋庝汉") == "鏄?):
            vat_display_rate = f"{vat5.get('澧炲€肩◣鍚嶄箟绋庣巼', 0.03)*100:.0f}%锛堜竴鑸撼绋庝汉锛? if vat5 else "-"

        rows = [
            {
                "搴忓彿": "1",
                "绋庢绫诲瀷": "馃彚 澧炲€肩◣",
                "鏍囩О绋庣巼": "3%锛堝皬瑙勬ā锛?,
                "浼樻儬鍚庣◣鐜?: vat_display_rate,
                "閫傜敤浼樻儬": "灏忚妯″噺鎸?% + 瀛ｂ墹30涓囧厤绋? if vat_amount == 0 else "灏忚妯″噺鎸?%寰佹敹",
                "鏀跨瓥渚濇嵁": "璐㈢◣銆?023銆?9鍙?,
                "鏈湡搴旂即锛堝厓锛?: f"{vat_amount:,.2f}",
                "鐘舵€?: "鍏嶇◣ 鉁? if vat_amount == 0 else "寰呯即 鈴?,
            },
            {
                "搴忓彿": "2",
                "绋庢绫诲瀷": "馃彊锔?鍩庡缓绋?,
                "鏍囩О绋庣巼": "7%锛堝競鍖猴級",
                "浼樻儬鍚庣◣鐜?: "3.5%锛堝噺鍗婏級" if (vat5 and vat5.get("鏄惁浜彈鍏◣涓よ垂鍑忓崐") == "鏄?) else "7%",
                "閫傜敤浼樻儬": "銆屽叚绋庝袱璐广€嶅噺鍗婂緛鏀? if (vat5 and vat5.get("鏄惁浜彈鍏◣涓よ垂鍑忓崐") == "鏄?) else "鏃?,
                "鏀跨瓥渚濇嵁": "璐㈢◣銆?022銆?0鍙?,
                "鏈湡搴旂即锛堝厓锛?: f"{urban_tax:,.2f}",
                "鐘舵€?: "鍏嶇◣ 鉁? if urban_tax == 0 else "寰呯即 鈴?,
            },
            {
                "搴忓彿": "3",
                "绋庢绫诲瀷": "馃摎 鏁欒偛璐归檮鍔?,
                "鏍囩О绋庣巼": "3%",
                "浼樻儬鍚庣◣鐜?: "1.5%锛堝噺鍗婏級" if (vat5 and vat5.get("鏄惁浜彈鍏◣涓よ垂鍑忓崐") == "鏄?) else "3%",
                "閫傜敤浼樻儬": "銆屽叚绋庝袱璐广€嶅噺鍗婂緛鏀?,
                "鏀跨瓥渚濇嵁": "璐㈢◣銆?022銆?0鍙?,
                "鏈湡搴旂即锛堝厓锛?: f"{edu_surcharge:,.2f}",
                "鐘舵€?: "鍏嶇◣ 鉁? if edu_surcharge == 0 else "寰呯即 鈴?,
            },
            {
                "搴忓彿": "4",
                "绋庢绫诲瀷": "馃帗 鍦版柟鏁欒偛闄勫姞",
                "鏍囩О绋庣巼": "2%",
                "浼樻儬鍚庣◣鐜?: "1%锛堝噺鍗婏級" if (vat5 and vat5.get("鏄惁浜彈鍏◣涓よ垂鍑忓崐") == "鏄?) else "2%",
                "閫傜敤浼樻儬": "銆屽叚绋庝袱璐广€嶅噺鍗婂緛鏀?,
                "鏀跨瓥渚濇嵁": "璐㈢◣銆?022銆?0鍙?,
                "鏈湡搴旂即锛堝厓锛?: f"{local_edu:,.2f}",
                "鐘舵€?: "鍏嶇◣ 鉁? if local_edu == 0 else "寰呯即 鈴?,
            },
            {
                "搴忓彿": "5",
                "绋庢绫诲瀷": "馃捈 浼佷笟鎵€寰楃◣锛堝棰勭即锛?,
                "鏍囩О绋庣巼": "25%",
                "浼樻儬鍚庣◣鐜?: "5%锛堝皬寰紭鎯狅級" if is_small else "25%",
                "閫傜敤浼樻儬": "灏忓瀷寰埄浼佷笟锛氬噺鎸?5%璁＄◣脳20%绋庣巼=5%" if is_small else "涓嶉€傜敤灏忓井浼樻儬",
                "鏀跨瓥渚濇嵁": "璐㈢◣銆?023銆?2鍙?,
                "鏈湡搴旂即锛堝厓锛?: f"{corp_tax:,.2f}",
                "鐘舵€?: "鏃犻渶缂寸撼 鉁? if corp_tax == 0 else "寰呯即 鈴?,
            },
            {
                "搴忓彿": "6",
                "绋庢绫诲瀷": "馃懁 涓汉鎵€寰楃◣锛堜唬鎵ｄ唬缂达級",
                "鏍囩О绋庣巼": "3%-45%绱繘",
                "浼樻儬鍚庣◣鐜?: "3%-45%锛堣捣寰佺偣5000鍏?鏈堬級",
                "閫傜敤浼樻儬": f"鍩烘湰鍑忛櫎5000鍏?涓撻」闄勫姞鎵ｉ櫎锛坽employee_count}鍚嶅憳宸ワ級",
                "鏀跨瓥渚濇嵁": "涓汉鎵€寰楃◣娉?,
                "鏈湡搴旂即锛堝厓锛?: f"{total_personal_tax:,.2f}",
                "鐘舵€?: "鏃犱釜绋?鉁? if total_personal_tax == 0 else "寰呯即 鈴?,
            },
        ]

        # 鍗拌姳绋庢槑缁嗚
        for si in stamp_data.get("鏄庣粏", []):
            rows.append({
                "搴忓彿": f"7-{stamp_data['鏄庣粏'].index(si)+1}" if len(stamp_data.get("鏄庣粏", [])) > 1 else "7",
                "绋庢绫诲瀷": f"馃摐 鍗拌姳绋?{si['绋庣洰']}",
                "鏍囩О绋庣巼": si["鍚嶄箟绋庣巼"],
                "浼樻儬鍚庣◣鐜?: si["浼樻儬鍚庣◣鐜?],
                "閫傜敤浼樻儬": "銆屽叚绋庝袱璐广€嶅噺鍗婂緛鏀? if stamp_data["鏄惁鍏◣涓よ垂鍑忓崐"] == "鏄? else "鏍囧噯绋庣巼",
                "鏀跨瓥渚濇嵁": "鍗拌姳绋庢硶 + 璐㈢◣銆?022銆?0鍙?,
                "鏈湡搴旂即锛堝厓锛?: f"{si['搴旂撼绋庨锛堝厓锛?]:,.2f}",
                "鐘舵€?: "鍏嶇◣ 鉁? if si['搴旂撼绋庨锛堝厓锛?] == 0 else "寰呯即 鈴?,
            })

        if stamp_data["绋庣洰鏁伴噺"] == 0:
            rows.append({
                "搴忓彿": "7",
                "绋庢绫诲瀷": "馃摐 鍗拌姳绋?,
                "鏍囩О绋庣巼": "瑙佸悇绋庣洰",
                "浼樻儬鍚庣◣鐜?: "鍏◣涓よ垂鍑忓崐",
                "閫傜敤浼樻儬": "灏忓瀷寰埄 鈫?鍚勭◣鐩噺鍗?,
                "鏀跨瓥渚濇嵁": "鍗拌姳绋庢硶锛?022.7.1锛?,
                "鏈湡搴旂即锛堝厓锛?: "0.00",
                "鐘舵€?: "鏈鏃犻渶缂寸撼 鉁?,
            })

        df_tax_list = pd.DataFrame(rows)
        st.dataframe(
            df_tax_list.style.apply(
                lambda row: ["background-color: #e8f5e9" if "鍏嶇◣" in str(row["鐘舵€?]) or "鏃犻渶缂寸撼" in str(row["鐘舵€?]) else
                             ("background-color: #fff9c4" if "寰呯即" in str(row["鐘舵€?]) else "")
                             for _ in row],
                axis=1
            ),
            use_container_width=True,
            hide_index=True,
        )

        st.caption("馃煝 缁胯壊琛?= 浜彈浼樻儬鍚庢棤闇€缂寸撼 | 馃煛 榛勮壊琛?= 闇€鎸夋湡缂寸撼")

        # ===== 3. 鍏◣涓よ垂鍑忓崐鏄庣粏 =====
        if six_two_relief > 0 or stamp_relief > 0:
            st.divider()
            st.subheader("馃巵 銆屽叚绋庝袱璐广€嶅噺鍗婂緛鏀?鈥?鏈鍑忓厤鏄庣粏")
            col_count = 3 + (1 if stamp_relief > 0 else 0)
            relief_cols = st.columns(col_count)
            relief_idx = 0
            relief_cols[relief_idx].metric("鍩庡缓绋庡噺鍏?, f"{vat5.get('鍩庡缓绋庡悕涔?, 0) - urban_tax:,.2f} 鍏?); relief_idx += 1
            relief_cols[relief_idx].metric("鏁欒偛璐归檮鍔犲噺鍏?, f"{vat5.get('鏁欒偛璐归檮鍔犲悕涔?, 0) - edu_surcharge:,.2f} 鍏?); relief_idx += 1
            relief_cols[relief_idx].metric("鍦版柟鏁欒偛闄勫姞鍑忓厤", f"{vat5.get('鍦版柟鏁欒偛闄勫姞鍚嶄箟', 0) - local_edu:,.2f} 鍏?); relief_idx += 1
            if stamp_relief > 0:
                relief_cols[relief_idx].metric("鍗拌姳绋庡噺鍏?, f"{stamp_relief:,.2f} 鍏?)
            st.success(f"馃挵 鏈銆屽叚绋庝袱璐广€嶅悎璁″噺鍏嶏細**{six_two_relief + stamp_relief:,.2f} 鍏?*")

        # ===== 4. 姹囨€绘寚鏍囧崱 =====
        st.divider()
        st.subheader("馃挵 鏈湡绋庢姹囨€伙紙浼樻儬鍚庯級")

        mc1, mc2, mc3, mc4, mc5 = st.columns(5)
        mc1.metric(
            "澧炲€肩◣鍙婇檮鍔?,
            f"{round(vat_amount + surcharge_total, 2):,.2f} 鍏?,
            delta="鍏ㄩ儴鍏嶇◣" if vat_amount == 0 else None,
        )
        mc2.metric(
            "浼佷笟鎵€寰楃◣",
            f"{corp_tax:,.2f} 鍏?,
            delta=f"鍑忓厤 {corp_relief:,.0f} 鍏? if corp_relief > 0 else ("浜忔崯鏃犻渶缂寸撼" if corp_tax == 0 else None),
        )
        mc3.metric(
            "涓汉鎵€寰楃◣",
            f"{total_personal_tax:,.2f} 鍏?,
            delta=f"{employee_count}鍚嶅憳宸? if employee_count > 0 else None,
        )
        mc4.metric(
            "馃摐 鍗拌姳绋?,
            f"{stamp_total:,.2f} 鍏?,
            delta=f"鍑忓厤 {stamp_relief:,.0f} 鍏? if stamp_relief > 0 else "鏈鏃犻渶缂寸撼",
        )
        mc5.metric(
            "馃敶 鏈湡绋庢鍚堣",
            f"{total_tax:,.2f} 鍏?,
            delta=f"鍏◣涓よ垂鍑忓厤 {six_two_relief + stamp_relief:,.0f}" if (six_two_relief + stamp_relief) > 0 else None,
        )

        # ===== 5. 绀句繚鎻愰啋 =====
        if total_company_social > 0 or total_personal_social > 0:
            st.divider()
            st.subheader("馃洝锔?绀句繚缂磋垂鎻愰啋锛堥潪绋庢锛屽崟鐙即绾筹級")
            s1, s2, s3 = st.columns(3)
            s1.metric("鍏徃鎵挎媴绀句繚", f"{total_company_social:,.2f} 鍏?)
            s2.metric("涓汉鎵挎媴绀句繚", f"{total_personal_social:,.2f} 鍏?)
            s3.metric("绀句繚鍚堣", f"{total_company_social + total_personal_social:,.2f} 鍏?)
            st.info("馃挕 绀句繚缂磋垂璇风櫥褰?*姝︽眽甯傜ぞ浼氫繚闄╃綉涓婃湇鍔″钩鍙?*鎴栭€氳繃閾惰浠ｆ墸瀹屾垚锛屾埅姝㈡棩鏈熶竴鑸负褰撴湀25鏃ャ€?)

        # ===== 6. 娈嬩繚閲?=====
        if employee_count <= 30:
            st.divider()
            st.subheader("鈾?娈嬬柧浜哄氨涓氫繚闅滈噾")
            st.success(f"鉁?鍦ㄨ亴鑱屽伐 {employee_count} 浜?鈮?30浜?鈫?**鍏嶅緛娈嬬柧浜哄氨涓氫繚闅滈噾**锛堝彂鏀逛环鏍艰銆?019銆?015鍙凤級")

        # ===== 7. 缂存鏈熼檺 =====
        st.divider()
        st.subheader("鈴?缂存鏈熼檺鎻愰啋")

        q_end_month = {1: 3, 2: 6, 3: 9, 4: 12}
        deadline_month = q_end_month.get(disp_quarter, 3) + 1
        if deadline_month > 12:
            deadline_month = 1
            deadline_year = disp_year + 1
        else:
            deadline_year = disp_year

        st.markdown(f"""
    | 绋庣 | 浼樻儬鏀跨瓥 | 瀹為檯绋庣巼 | 鐢虫姤鎴鏃ユ湡 | 鐢虫姤骞冲彴 |
    |------|---------|---------|------------|---------|
    | 澧炲€肩◣ | 灏忚妯″噺鎸?% | {vat_display_rate} | **{deadline_year}骞磠deadline_month}鏈?5鏃?* | 婀栧寳鐪佺數瀛愮◣鍔″眬 |
    | 鍩庡缓绋?| 鍏◣涓よ垂鍑忓崐 | 3.5% | 鍚屼笂 | 闅忓鍊肩◣涓€骞剁敵鎶?|
    | 鏁欒偛璐归檮鍔?鍦版柟鏁欒偛闄勫姞 | 鍏◣涓よ垂鍑忓崐 | 1.5%+1% | 鍚屼笂 | 闅忓鍊肩◣涓€骞剁敵鎶?|
    | 浼佷笟鎵€寰楃◣ | 灏忓瀷寰埄5% | {'5%' if is_small else '25%'} | **{deadline_year}骞磠deadline_month}鏈?5鏃?* | 鐢靛瓙绋庡姟灞€ 鈫?A200000 |
    | 涓汉鎵€寰楃◣ | 璧峰緛鐐?000 | 3%-45% | **娆℃湀15鏃?* | 鑷劧浜虹◣鏀剁鐞嗙郴缁?|
    | 鍗拌姳绋?| 鍏◣涓よ垂鍑忓崐 | 鍚勭◣鐩噺鍗?| 鎸夋/鎸夋湡姹囨€?| 婀栧寳鐪佺數瀛愮◣鍔″眬 |
    | 娈嬩繚閲?| 鈮?0浜哄厤寰?| 0% | 骞村害鐢虫姤 | 娈嬭仈/绋庡姟閮ㄩ棬 |
        """)

        st.markdown("""
    > **馃搶 鎿嶄綔姝ラ锛?*
    > 1. 鐧诲綍 [婀栧寳鐪佺數瀛愮◣鍔″眬](https://etax.hubei.chinatax.gov.cn/) 瀹屾垚澧炲€肩◣鍙婁紒涓氭墍寰楃◣鐢虫姤
    > 2. 鐢虫姤瀹屾垚鍚庯紝閫氳繃缃戦摱鎴栫涓夋柟鏀粯瀹屾垚绋庢鍒掔即
    > 3. 鎴浘鐣欏瓨鐢虫姤鎴愬姛椤甸潰锛屽綊鍏ョ◣鍔℃。妗?
    > 4. 涓◣閫氳繃**鑷劧浜虹◣鏀剁鐞嗙郴缁燂紙鎵ｇ即瀹㈡埛绔級**鐢虫姤骞剁即绾?
    > 5. 浼樻儬鏀跨瓥**鏃犻渶棰濆鐢宠**锛岀郴缁熻嚜鍔ㄨ瘑鍒噺鍏嶏紙婀栧寳鐪併€屽厤鐢冲嵆浜€嶏級
        """)

        # ===== 8. 涓嬭浇 =====
        st.divider()
        dl1, dl2 = st.columns(2)
        with dl1:
            csv_tax = df_tax_list.to_csv(index=False, encoding="utf-8-sig")
            st.download_button(
                label="馃摜 涓嬭浇绋庢缂寸撼娓呭崟锛圕SV锛?,
                data=csv_tax,
                file_name=f"绋庢缂寸撼娓呭崟_{disp_year}Q{disp_quarter}.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with dl2:
            summary_lines = [
                f"姝︽眽閲戣壋榫欑鎶€鏈夐檺鍏徃",
                f"{disp_year}骞寸{disp_quarter}瀛ｅ害 绋庢缂寸撼娓呭崟锛堝惈浼樻儬鏀跨瓥锛?,
                f"鐢熸垚鏃堕棿锛歿datetime.now().strftime('%Y-%m-%d %H:%M')}",
                "=" * 50,
                f"鈻?閫傜敤浼樻儬鏀跨瓥锛?,
                f"  1. 灏忚妯＄撼绋庝汉瀛ｅ害鍏嶇◣/鍑忔寜1%锛堝鈮?0涓囧厤锛岃秴鍒?%锛?,
                f"  2. 鍏◣涓よ垂鍑忓崐寰佹敹锛堝煄寤?鏁欒偛/鍦版柟鏁欒偛/鍗拌姳绋庯級",
                f"  3. 灏忓瀷寰埄浼佷笟鎵€寰楃◣5%锛堝勾鍒╂鼎鈮?00涓囷級",
                f"  4. 娈嬩繚閲戝厤寰侊紙鍛樺伐鈮?0浜猴級",
                "=" * 50,
                f"鈻?鏈湡瀹為檯搴旂即锛?,
                f"  澧炲€肩◣锛歿vat_amount:,.2f} 鍏?({vat_display_rate})",
                f"  鍩庡缓绋庯細{urban_tax:,.2f} 鍏?(3.5%)",
                f"  鏁欒偛璐归檮鍔狅細{edu_surcharge:,.2f} 鍏?(1.5%)",
                f"  鍦版柟鏁欒偛闄勫姞锛歿local_edu:,.2f} 鍏?(1%)",
                f"  浼佷笟鎵€寰楃◣锛歿corp_tax:,.2f} 鍏?({'5%' if is_small else '25%'})",
                f"  涓汉鎵€寰楃◣锛歿total_personal_tax:,.2f} 鍏?,
                f"  鍗拌姳绋庯細{stamp_total:,.2f} 鍏冿紙鍚祫閲戣处绨?璐攢鍚堝悓绛夛紝鍑忓崐鍚庯級",
                "-" * 30,
                f"  鏈湡绋庢鍚堣锛歿total_tax:,.2f} 鍏?,
                f"  鍏◣涓よ垂鍑忓厤锛堝惈鍗拌姳绋庯級锛歿six_two_relief + stamp_relief:,.2f} 鍏?,
                f"  浼佷笟鎵€寰楃◣鍑忓厤锛歿corp_relief:,.2f} 鍏?,
                "=" * 50,
                f"鐢虫姤鎴锛歿deadline_year}骞磠deadline_month}鏈?5鏃ワ紙澧炲€肩◣銆佷紒涓氭墍寰楃◣锛?,
                f"涓◣鎴锛氭鏈?5鏃?,
                f"绀句繚鎴锛氬綋鏈?5鏃?,
                f"鍗拌姳绋庯細鎸夋鎴栨寜鏈熸眹鎬荤敵鎶?,
                f"鏀跨瓥鏈夋晥鏈熻嚦锛?027骞?2鏈?1鏃?,
            ]
            summary_txt = "\n".join(summary_lines)
            st.download_button(
                label="馃摜 涓嬭浇缂寸撼閫氱煡涔︼紙TXT锛?,
                data=summary_txt,
                file_name=f"绋庢缂寸撼閫氱煡涔{disp_year}Q{disp_quarter}.txt",
                mime="text/plain",
                use_container_width=True,
            )

    # ===============================================
    #  Tab6锛氱敵鎶ユ寚鍗楋紙鍥涘悎涓€锛?
    # ===============================================

with tab8:
    st.header("馃摉 姝︽眽閲戣壋榫欑鎶€鏈夐檺鍏徃 鈥?绋庡姟鐢虫姤鎿嶄綔鎸囧崡")
    st.caption("鍩轰簬灏忚妯＄撼绋庝汉銆佸皬鍨嬪井鍒╀紒涓氬満鏅?路 婀栧寳鐪?026骞村害閫傜敤")

    guide_choice = st.radio(
        "璇烽€夋嫨瑕佹煡闃呯殑鐢虫姤璇存槑锛?,
        [
            "涓€銆佷釜浜烘墍寰楃◣鐢虫姤锛堜唬鎵ｄ唬缂达級",
            "浜屻€佷紒涓氭墍寰楃◣瀛ｆ姤 & 骞存姤",
            "涓夈€佸埄娑﹁〃缂栧埗璇存槑",
            "鍥涖€佸鍊肩◣鍙婇檮鍔犵◣绛夌◣鐩敵鎶?,
        ],
        horizontal=True,
    )

    st.divider()

    # ==========================================
    #  涓€銆佷釜浜烘墍寰楃◣鐢虫姤
    # ==========================================
    if guide_choice.startswith("涓€銆?):
        st.subheader("涓€銆佷釜浜烘墍寰楃◣鐢虫姤锛堜唬鎵ｄ唬缂达級")

        with st.expander("1.1 鐢虫姤涓讳綋涓庝箟鍔?, expanded=True):
            st.markdown("""
**鐢虫姤涓讳綋**锛氭姹夐噾鑹抽緳绉戞妧鏈夐檺鍏徃锛堜綔涓恒€屾墸缂翠箟鍔′汉銆嶏級

**娉曞緥渚濇嵁**锛氥€婁腑鍗庝汉姘戝叡鍜屽浗涓汉鎵€寰楃◣娉曘€嬬涔濇潯 鈥斺€?涓汉鎵€寰楃◣浠ユ墍寰椾汉涓虹撼绋庝汉锛屼互鏀粯鎵€寰楃殑鍗曚綅鎴栬€呬釜浜轰负鎵ｇ即涔夊姟浜恒€?

**鐢虫姤鍐呭**锛?
- 宸ヨ祫銆佽柂閲戞墍寰楋紙鍛樺伐姣忔湀宸ヨ祫锛?
- 鍔冲姟鎶ラ叕鎵€寰楋紙濡傛湁澶栭儴浜哄憳锛?
- 鍏徃闇€鍦ㄥ彂鏀惧伐璧勬椂渚濇硶浠ｆ墸浠ｇ即涓◣锛屾寜鏈堝悜绋庡姟鏈哄叧鐢虫姤

**鍏抽敭姒傚康**锛?
| 鏈 | 璇存槑 |
|------|------|
| 鎵ｇ即涔夊姟浜?| 鍏徃锛堜唬鎵ｄ唬缂达級 |
| 绾崇◣浜?| 鍛樺伐涓汉 |
| 绱棰勬墸娉?| 鎸夋湀绱璁＄畻锛屽叏骞存眹绠?|
| 涓撻」闄勫姞鎵ｉ櫎 | 瀛愬コ鏁欒偛/濠村辜鍎跨収鎶?璧″吇鑰佷汉/浣忔埧璐锋/浣忔埧绉熼噾/缁х画鏁欒偛/澶х梾鍖荤枟 |
            """)

        with st.expander("1.2 璁＄畻鍏紡锛堢疮璁￠鎵ｆ硶锛?, expanded=False):
            st.markdown("""
**鏍稿績鍏紡锛?*

鏈湡搴旈鎵ｉ缂寸◣棰?= 锛堢疮璁￠鎵ｉ缂村簲绾崇◣鎵€寰楅 脳 棰勬墸鐜?- 閫熺畻鎵ｉ櫎鏁帮級 - 绱鍑忓厤绋庨 - 绱宸查鎵ｉ缂寸◣棰?

鍏朵腑锛?
- 绱棰勬墸棰勭即搴旂撼绋庢墍寰楅 = 绱鏀跺叆 - 绱鍏嶇◣鏀跺叆 - 绱鍑忛櫎璐圭敤 - 绱涓撻」鎵ｉ櫎 - 绱涓撻」闄勫姞鎵ｉ櫎 - 绱渚濇硶纭畾鐨勫叾浠栨墸闄?

**鏈堝害璁＄畻鍏紡锛堢畝鍖栵級锛?*
```
搴旂◣鏀跺叆 = 绋庡墠宸ヨ祫 - 涓汉绀句繚 - 鍩烘湰鍑忛櫎璐圭敤(5000鍏? - 涓撻」闄勫姞鎵ｉ櫎
搴旂撼绋庨 = 搴旂◣鏀跺叆 脳 閫傜敤绋庣巼 - 閫熺畻鎵ｉ櫎鏁?
瀹炲彂宸ヨ祫 = 绋庡墠宸ヨ祫 - 涓汉绀句繚 - 搴旂撼绋庨
```

**7绾ц秴棰濈疮杩涚◣鐜囪〃锛?024-2026骞达級锛?*

| 绾ф暟 | 绱搴旂撼绋庢墍寰楅 | 绋庣巼 | 閫熺畻鎵ｉ櫎鏁?|
|------|-----------------|------|-----------|
| 1 | 鈮?3,000 | 3% | 0 |
| 2 | 3,001 ~ 12,000 | 10% | 210 |
| 3 | 12,001 ~ 25,000 | 20% | 1,410 |
| 4 | 25,001 ~ 35,000 | 25% | 2,660 |
| 5 | 35,001 ~ 55,000 | 30% | 4,410 |
| 6 | 55,001 ~ 80,000 | 35% | 7,160 |
| 7 | > 80,000 | 45% | 15,160 |

**绀轰緥璁＄畻**锛堝憳宸锛岀◣鍓?0522鍏冿紝涓汉绀句繚522鍏冿紝涓撻」闄勫姞鎵ｉ櫎5000鍏冿級锛?
```
搴旂◣鏀跺叆 = 10522 - 522 - 5000 - 5000 = 0 鍏?
搴旂撼绋庨 = 0 鍏冿紙鏃犻渶缂寸撼涓◣锛岄渶杩涜闆剁敵鎶ワ級
瀹炲彂宸ヨ祫 = 10522 - 522 = 10,000 鍏?
```
            """)

        with st.expander("1.3 涓撻」闄勫姞鎵ｉ櫎鏄庣粏", expanded=False):
            st.markdown("""
**2026骞存姹夊湴鍖洪€傜敤鏍囧噯锛?*

| 鎵ｉ櫎椤圭洰 | 姣忔湀鏍囧噯 | 璇存槑 |
|----------|---------|------|
| 瀛愬コ鏁欒偛 | 2,000鍏?浜?| 3宀佽嚦鍗氬＋鍧囧彲锛岀埗姣嶅悇鎵?0%鎴栦竴鏂瑰叏鎵?|
| 濠村辜鍎跨収鎶?| 2,000鍏?浜?| 0~3宀佸┐骞煎効 |
| 璧″吇鑰佷汉 | 鏈€楂?,000鍏?| 60宀佷互涓婄埗姣嶏紝鐙敓瀛愬コ3000鍏冿紝闈炵嫭鐢熷瓙濂冲垎鎽?|
| 浣忔埧璐锋鍒╂伅 | 1,000鍏?| 棣栧鎴胯捶娆?|
| 浣忔埧绉熼噾 | 800~1,500鍏?| 姝︽眽灞炵渷浼氬煄甯傦紝鎸?,500鍏?鏈?|
| 缁х画鏁欒偛 | 400鍏?鏈?| 瀛﹀巻缁х画鏁欒偛鏈熼棿 |
| 澶х梾鍖荤枟 | 鎹疄鎵ｉ櫎 | 骞村害姹囩畻娓呯即鏃舵墸闄?|

> **閲嶈鎻愰啋**锛氬憳宸ラ渶鍦ㄣ€屼釜浜烘墍寰楃◣APP銆嶄腑鑷濉姤涓撻」闄勫姞鎵ｉ櫎淇℃伅锛屽叕鍙哥锛堟墸缂寸锛夎嚜鍔ㄥ悓姝ャ€傚鏈夊彉鍔ㄨ鎻愰啋鍛樺伐鍙婃椂鏇存柊銆?
            """)

        with st.expander("1.4 鐢虫姤鎿嶄綔娴佺▼", expanded=False):
            st.markdown("""
**鐢虫姤骞冲彴**锛氳嚜鐒朵汉鐢靛瓙绋庡姟灞€锛堟墸缂寸锛?鈫?https://etax.chinatax.gov.cn/

**鎿嶄綔姝ラ锛堟瘡鏈堜竴娆★級锛?*

```
绗?姝ワ細浜哄憳淇℃伅閲囬泦
  鈹溾攢 鐧诲綍鎵ｇ即绔?鈫?銆屼汉鍛樹俊鎭噰闆嗐€?鈫?娣诲姞/瀵煎叆鍛樺伐淇℃伅
  鈹溾攢 蹇呭～锛氬鍚嶃€佽韩浠借瘉鍙枫€佷换鑱屽彈闆囦粠涓氱被鍨嬶紙闆囧憳锛?
  鈹斺攢 鎶ラ€?鈫?鑾峰彇鍙嶉锛堥獙璇侀€氳繃鍚庡彲鐢級

绗?姝ワ細涓撻」闄勫姞鎵ｉ櫎淇℃伅閲囬泦
  鈹溾攢 銆屼笓椤归檮鍔犳墸闄や俊鎭噰闆嗐€?鈫?銆屼笅杞芥洿鏂般€?
  鈹溾攢 绯荤粺鑷姩浠庣◣鍔″眬鍚屾鍛樺伐鍦ˋPP涓～鎶ョ殑淇℃伅
  鈹斺攢 鏍稿鏄惁鏈夋柊澧?鍙樻洿鐨勫憳宸ユ墸闄や俊鎭?

绗?姝ワ細缁煎悎鎵€寰楃敵鎶?
  鈹溾攢 銆岀患鍚堟墍寰楃敵鎶ャ€?鈫?銆屾敹鍏ュ強鍑忛櫎濉啓銆?鈫?銆屾甯稿伐璧勮柂閲戞墍寰椼€?
  鈹溾攢 瀵煎叆鎴栨墜鍔ㄥ～鍐欙細鏈湡鏀跺叆銆佷釜浜虹ぞ淇濄€佷笓椤归檮鍔犳墸闄?
  鈹溾攢 銆岀◣娆捐绠椼€?鈫?绯荤粺鑷姩璁＄畻姣忎綅鍛樺伐搴旂即涓◣
  鈹斺攢 銆岀敵鎶ヨ〃鎶ラ€併€?鈫?鏍稿鏃犺鍚庡彂閫佺敵鎶?

绗?姝ワ細绋庢缂寸撼
  鈹溾攢 銆岀◣娆剧即绾炽€?鈫?閫夋嫨涓夋柟鍗忚鎵ｆ鎴栭摱琛岀鏌ヨ缂寸◣
  鈹斺攢 鎵ｆ鎴愬姛鍚庢埅鍥剧暀瀛樺嚟璇?
```

**闆剁敵鎶ユ搷浣?*锛堝鎵€鏈夊憳宸ュ簲绋庢敹鍏ヤ负0锛夛細
- 鍚屾牱璧板畬鏁存祦绋嬶紝绋庢璁＄畻缁撴灉涓?鍏?
- 浠嶉渶鐐瑰嚮銆岀敵鎶ヨ〃鎶ラ€併€嶅畬鎴愰浂鐢虫姤
- **鍒囧嬁閬楁紡闆剁敵鎶?*锛屽惁鍒欎細浜х敓閫炬湡鏈敵鎶ヨ褰?
            """)

        with st.expander("1.5 鐢虫姤鍛ㄦ湡涓庢埅姝㈡棩鏈?, expanded=False):
            st.markdown("""
| 浜嬮」 | 鍛ㄦ湡 | 鎴鏃堕棿 | 璇存槑 |
|------|------|---------|------|
| 鏈堝害浠ｆ墸浠ｇ即 | 姣忔湀 | 娆℃湀15鏃?| 渚嬶細5鏈堝伐璧?鈫?6鏈?5鏃ュ墠鐢虫姤 |
| 骞村害姹囩畻娓呯即 | 姣忓勾 | 娆″勾3鏈?鏃6鏈?0鏃?| 鍛樺伐鑷鍦ㄤ釜绋嶢PP鎿嶄綔锛屽叕鍙告棤闇€浠ｅ姙 |
| 鎵ｇ即绔汉鍛樹俊鎭洿鏂?| 鏈夊彉鍔ㄦ椂 | 褰撴湀鐢虫姤鍓?| 鏂板叆鑱?绂昏亴鍛樺伐闇€鍙婃椂鏇存柊 |

**閫炬湡鍚庢灉**锛?
- 閫炬湡鐢虫姤锛氭寜鏃ュ姞鏀舵粸绾崇◣娆句竾鍒嗕箣浜旂殑婊炵撼閲?
- 杩炵画閫炬湡锛氬彲鑳借鍒楀叆閲嶇偣鐩戠鍚嶅崟
            """)

        with st.expander("1.6 甯歌闂", expanded=False):
            st.markdown("""
**Q1锛氬憳宸ュ伐璧勫緢浣庯紝杩橀渶瑕佺敵鎶ュ悧锛?*
A锛氶渶瑕併€傚嵆浣垮簲绋庢敹鍏ヤ负0銆佹棤闇€缂寸◣锛屼篃蹇呴』杩涜闆剁敵鎶ャ€?

**Q2锛氫笓椤归檮鍔犳墸闄よ皝濉紵**
A锛氬憳宸ュ湪涓汉鎵€寰楃◣APP涓嚜琛屽～鎶ワ紝鍏徃鍦ㄦ墸缂寸銆屼笅杞芥洿鏂般€嶅嵆鍙悓姝ャ€?

**Q3锛氱ぞ淇濆熀鏁板彉浜嗘€庝箞鍔烇紵**
A锛氭瘡骞寸ぞ淇濆熀鏁拌皟鏁村悗锛堥€氬父7鏈堬級锛屽湪鎵ｇ即绔慨鏀瑰憳宸ャ€屼釜浜虹ぞ淇濄€嶉噾棰濆嵆鍙€?

**Q4锛氬叏骞翠竴娆℃€у閲戯紙骞寸粓濂栵級鎬庝箞澶勭悊锛?*
A锛氬彲閫夋嫨鍗曠嫭璁＄◣鎴栧苟鍏ョ患鍚堟墍寰椼€傚缓璁祴绠椾袱绉嶆柟寮忓悗閫夋嫨绋庨鏇翠綆鐨勬柟妗堛€?

**Q5锛氱鑱屽憳宸ユ€庝箞澶勭悊锛?*
A锛氬湪銆屼汉鍛樹俊鎭噰闆嗐€嶄腑灏嗙姸鎬佹敼涓恒€岄潪姝ｅ父銆嶏紝濉啓绂昏亴鏃ユ湡銆?
            """)

    # ==========================================
    #  浜屻€佷紒涓氭墍寰楃◣瀛ｆ姤 & 骞存姤
    # ==========================================
    elif guide_choice.startswith("浜屻€?):
        st.subheader("浜屻€佷紒涓氭墍寰楃◣瀛ｆ姤 & 骞存姤")

        with st.expander("2.1 鐢虫姤绫诲瀷涓庡懆鏈熸€昏", expanded=True):
            st.markdown("""
**浼佷笟鎵€寰楃◣鐢虫姤鍒嗕负涓ょ被锛?*

| 鐢虫姤绫诲瀷 | 鍛ㄦ湡 | 琛ㄥ崟 | 鎴鏃堕棿 | 璇存槑 |
|----------|------|------|---------|------|
| **瀛ｅ害棰勭即** | 姣忓搴?| A200000琛?| 瀛ｅ害缁堜簡鍚?5鏃ュ唴 | 鎸夊搴﹀埄娑﹂缂?|
| **骞村害姹囩畻娓呯即** | 姣忓勾 | A绫诲勾搴︾敵鎶ヨ〃 | 娆″勾5鏈?1鏃ュ墠 | 鍏ㄥ勾姹囨€诲閫€灏戣ˉ |

**2026骞寸敵鎶ユ棩鍘嗭細**

| 鏈熼棿 | 鐢虫姤鍐呭 | 鎴鏃ユ湡 |
|------|---------|---------|
| Q1锛?-3鏈堬級 | 瀛ｅ害棰勭即 | **2026骞?鏈?5鏃?* |
| Q2锛?-6鏈堬級 | 瀛ｅ害棰勭即 | **2026骞?鏈?5鏃?* |
| Q3锛?-9鏈堬級 | 瀛ｅ害棰勭即 | **2026骞?0鏈?5鏃?* |
| Q4锛?0-12鏈堬級 | 瀛ｅ害棰勭即 | **2027骞?鏈?5鏃?* |
| 2026鍏ㄥ勾 | 骞村害姹囩畻娓呯即 | **2027骞?鏈?1鏃?* |

> 鈿狅笍 **閲嶈**锛氬搴﹂缂翠笉鑳藉皯鎶ユ垨婕忔姤锛屽惁鍒欏勾搴︽眹绠楁椂浼氳杩界即骞跺姞鏀舵粸绾抽噾銆?

**姝ゅ杩樻湁銆屽伐鍟嗗勾鎶ャ€嶏紙闈炵◣鍔★級锛?*
- 鐢虫姤骞冲彴锛氬浗瀹朵紒涓氫俊鐢ㄤ俊鎭叕绀虹郴缁?
- 鐢虫姤鏃堕棿锛氭瘡骞?鏈?鏃?~ 6鏈?0鏃?
- 鍐呭锛氫紒涓氬熀鏈俊鎭€佽偂涓滃嚭璧勩€佽祫浜х姸鍐点€佺ぞ淇濅俊鎭瓑
            """)

        with st.expander("2.2 瀛ｅ害棰勭即鐢虫姤锛圓200000琛級", expanded=False):
            st.markdown("""
**鐢虫姤骞冲彴**锛氭箹鍖楃渷鐢靛瓙绋庡姟灞€ 鈫?https://etax.hubei.chinatax.gov.cn/

**A200000琛ㄦ牳蹇冭娆★細**

```
绗?琛? 钀ヤ笟鏀跺叆          = 涓昏惀涓氬姟鏀跺叆 + 鍏朵粬涓氬姟鏀跺叆
绗?琛? 钀ヤ笟鎴愭湰          = 涓昏惀涓氬姟鎴愭湰 + 鍏朵粬涓氬姟鎴愭湰
绗?琛? 鍒╂鼎鎬婚          = 钀ヤ笟鏀跺叆 - 钀ヤ笟鎴愭湰 - 绋庨噾鍙婇檮鍔?- 绠＄悊璐圭敤 - 璐㈠姟璐圭敤 + 鎶曡祫鏀剁泭 + 钀ヤ笟澶栨敹鍏?- 钀ヤ笟澶栨敮鍑?
绗?琛? 鐗瑰畾涓氬姟璋冩暣      = 0锛堜竴鑸紒涓氭棤闇€濉啓锛?
绗?琛? 涓嶅緛绋庢敹鍏?       = 0锛堝鏈夋斂搴滆ˉ璐寸瓑闇€娉ㄦ槑锛?
绗?琛? 鍥哄畾璧勪骇鎶樻棫璋冩暣  = 0锛堥粯璁ゆ棤宸紓锛?
绗?琛? 寮ヨˉ浠ュ墠骞村害浜忔崯  = 浠ュ墠骞村害鏈讥琛ヤ簭鎹熼
绗?琛? 瀹為檯鍒╂鼎棰?       = 绗?琛?+ 绗?琛?- 绗?琛?- 绗?琛?- 绗?琛?
绗?琛? 绋庣巼              = 25%锛堝浐瀹氾級
绗?0琛?搴旂撼鎵€寰楃◣棰?     = 绗?琛?脳 25%
绗?1琛?鍑忓厤鎵€寰楃◣棰?     = 灏忓瀷寰埄浼佷笟浼樻儬鍑忓厤
绗?2琛?鏈勾绱宸查缂?   = 绱涔嬪墠瀛ｅ害宸茬即绋庨
绗?3琛?鏈湡搴旇ˉ(閫€)绋庨  = 鏈湡搴旂撼绋庨 - 绗?2琛?
```

**灏忓瀷寰埄浼佷笟浼樻儬锛圦1~Q4鑷姩閫傜敤锛夛細**
- 鏉′欢锛氬勾鍒╂鼎鈮?00涓?+ 鍛樺伐鈮?00浜?+ 璧勪骇鈮?000涓?
- 瀹為檯绋庣巼锛?*5%**锛堝噺鎸?5%璁″叆搴旂撼绋庢墍寰楅 脳 20%绋庣巼锛?
- 姝︽眽閲戣壋榫欑鎶€绗﹀悎鏉′欢 鉁?

**瀛ｅ害棰勭即涓庡埄娑﹁〃鐨勫叧绯伙細**
```
鍒╂鼎琛ㄤ腑鐨勩€屽埄娑︽€婚(绗?2琛?銆?鈮?鐢虫姤琛ㄣ€岀3琛?鍒╂鼎鎬婚銆?

宸紓璇存槑锛?
- 濡傛灉鍒╂鼎琛ㄦ寜灏忎紒涓氫細璁″噯鍒欐甯哥紪鍒讹紝鍙洿鎺ュ～鍏ョ敵鎶ヨ〃
- 濡傛湁绾崇◣璋冩暣椤癸紙濡備笉鍙墸闄よ垂鐢級锛屽湪骞村害姹囩畻鏃惰皟鏁达紝瀛ｅ害棰勭即鎸夎处闈㈠埄娑?
```
            """)

        with st.expander("2.3 骞村害姹囩畻娓呯即", expanded=False):
            st.markdown("""
**鐢虫姤骞冲彴**锛氭箹鍖楃渷鐢靛瓙绋庡姟灞€ 鈫?銆屼紒涓氭墍寰楃◣骞村害鐢虫姤銆?

**A绫诲勾搴︾敵鎶ヨ〃缁撴瀯锛?*

| 閮ㄥ垎 | 鍐呭 | 鍏抽敭琛?|
|------|------|--------|
| 涓昏〃 | A100000 骞村害绾崇◣鐢虫姤琛?| 姹囨€绘墍鏈夋暟鎹?|
| 鏀跺叆绫?| A101010 涓€鑸紒涓氭敹鍏ユ槑缁嗚〃 | 钀ヤ笟鏀跺叆/钀ヤ笟澶栨敹鍏?|
| 鎴愭湰绫?| A102010 涓€鑸紒涓氭垚鏈敮鍑烘槑缁嗚〃 | 钀ヤ笟鎴愭湰/鏈熼棿璐圭敤 |
| 鏈熼棿璐圭敤 | A104000 鏈熼棿璐圭敤鏄庣粏琛?| 绠＄悊/閿€鍞?璐㈠姟璐圭敤 |
| 绾崇◣璋冩暣 | A105000 绾崇◣璋冩暣椤圭洰鏄庣粏琛?| 璋冨/璋冨噺椤圭洰 |
| 浼樻儬绫?| A107040 鍑忓厤鎵€寰楃◣浼樻儬鏄庣粏琛?| 灏忓瀷寰埄浼佷笟绛?|
| 寮ヨˉ浜忔崯 | A106000 寮ヨˉ浜忔崯鏄庣粏琛?| 浠ュ墠骞村害浜忔崯 |
| 鑱屽伐钖叕 | A105050 鑱屽伐钖叕鏀嚭鍙婄撼绋庤皟鏁存槑缁嗚〃 | 宸ヨ祫/绂忓埄/绀句繚 |

**姹囩畻娓呯即鏍稿績閫昏緫锛?*
```
浼氳鍒╂鼎鎬婚锛堝埄娑﹁〃锛?
  + 绾崇◣璋冨椤癸紙瓒呮爣璐圭敤銆佷笉寰楁墸闄ゆ敮鍑虹瓑锛?
  - 绾崇◣璋冨噺椤癸紙鍏嶇◣鏀跺叆銆佸姞璁℃墸闄ょ瓑锛?
  - 寮ヨˉ浠ュ墠骞村害浜忔崯
  = 搴旂撼绋庢墍寰楅
  脳 閫傜敤绋庣巼锛堝皬鍨嬪井鍒?%锛?
  = 鍏ㄥ勾搴旂撼绋庨
  - 宸查缂寸◣棰濓紙Q1~Q4鍚堣锛?
  = 搴旇ˉ(閫€)绋庨
```

**甯歌绾崇◣璋冩暣椤癸細**

| 椤圭洰 | 璋冩暣鏂瑰悜 | 璇存槑 |
|------|---------|------|
| 涓氬姟鎷涘緟璐?| 璋冨 | 鎸夊彂鐢熼60%鎵ｉ櫎锛屾渶楂樹笉瓒呰繃钀ヤ笟鏀跺叆5鈥?|
| 骞垮憡瀹ｄ紶璐?| 鍙兘璋冨 | 涓嶈秴杩囪惀涓氭敹鍏?5%锛岃秴杩囬儴鍒嗗彲缁撹浆 |
| 缃氭/婊炵撼閲?| 璋冨 | 涓嶅緱绋庡墠鎵ｉ櫎 |
| 鏈彇寰楀彂绁ㄧ殑璐圭敤 | 璋冨 | 闇€鍙栧緱鍚堣鍙戠エ |
| 娈嬬柧浜哄憳宸ヨ祫 | 璋冨噺 | 鍙姞璁?00%鎵ｉ櫎 |
            """)

        with st.expander("2.4 灏忓瀷寰埄浼佷笟鍒ゅ畾涓庝紭鎯?, expanded=False):
            st.markdown("""
**灏忓瀷寰埄浼佷笟銆?35鏍囧噯銆嶏細**

| 鏉′欢 | 鏍囧噯 | 姝︽眽閲戣壋榫欑鎶€ | 鍒ゆ柇 |
|------|------|---------------|------|
| 骞村簲绾崇◣鎵€寰楅 | 鈮?300涓囧厓 | 鍙栧喅浜庡綋骞村埄娑?| 鉁?鈿狅笍 |
| 浠庝笟浜烘暟 | 鈮?300浜?| 閫氬父1-5浜?| 鉁?|
| 璧勪骇鎬婚 | 鈮?5,000涓囧厓 | 閫氬父杈冨皬 | 鉁?|

**浼樻儬鏀跨瓥锛?024-2027骞达級锛?*

```
瀹為檯绋庤礋 = 25%锛堝噺鎸夎绋庢瘮渚嬶級 脳 20%锛堜紭鎯犵◣鐜囷級 = 5%

馃敶 涓句緥锛?
  鍏ㄥ勾鍒╂鼎 100涓囧厓
  搴旂撼绋庢墍寰楅 = 100涓囷紙鍋囪鏃犺皟鏁撮」锛?
  搴旂撼绋庨锛堟爣鍑嗭級= 100涓?脳 25% = 25涓?
  搴旂撼绋庨锛堜紭鎯狅級= 100涓?脳 5% = 5涓?
  鍑忓厤绋庨 = 25涓?- 5涓?= 20涓?鉁?
```

**娉ㄦ剰浜嬮」锛?*
- 浼樻儬鏀跨瓥鏃犻渶棰濆鐢宠锛岀郴缁熻嚜鍔ㄥ垽瀹?
- 婀栧寳鐪佸凡瀹炵幇銆屽厤鐢冲嵆浜€?
- 濡傚勾搴﹀埄娑﹁秴杩?00涓囷紝褰撳勾涓嶅啀浜彈灏忓井浼樻儬锛岄渶鎸?5%缂寸撼
            """)

        with st.expander("2.5 鐢虫姤鎿嶄綔娴佺▼", expanded=False):
            st.markdown("""
**瀛ｅ害棰勭即鎿嶄綔姝ラ锛堟瘡瀛ｅ害涓€娆★級锛?*

```
绗?姝ワ細鐧诲綍 鈫?婀栧寳鐪佺數瀛愮◣鍔″眬 鈫?浼佷笟鐧诲綍锛堢◣鍙?瀵嗙爜鎴朇A璇佷功锛?
绗?姝ワ細杩涘叆 鈫?銆屾垜瑕佸姙绋庛€?鈫?銆岀◣璐圭敵鎶ュ強缂寸撼銆?
绗?姝ワ細閫夋嫨 鈫?銆屼紒涓氭墍寰楃◣鐢虫姤銆?鈫?銆屽眳姘戜紒涓氾紙鏌ヨ处寰佹敹锛夋墍寰楃◣鏈堬紙瀛ｏ級搴︾敵鎶ャ€?
绗?姝ワ細濉姤 鈫?鎸堿200000琛ㄦ牸寮忓～鍐欐敹鍏?鎴愭湰/鍒╂鼎
绗?姝ワ細绯荤粺鑷姩鍒ゅ畾灏忓瀷寰埄浼佷笟骞惰绠楀噺鍏嶇◣棰?
绗?姝ワ細棰勮 鈫?鏍稿涓庡埄娑﹁〃鏄惁涓€鑷?
绗?姝ワ細鐢虫姤 鈫?鐐瑰嚮銆岀敵鎶ャ€嶁啋 纭鎻愪氦
绗?姝ワ細缂存 鈫?濡傛湁搴旂即绋庢锛岄€氳繃涓夋柟鍗忚鎴栫綉閾剁即娆?
绗?姝ワ細鐣欏瓨 鈫?鎴浘鐢虫姤鎴愬姛椤甸潰锛屼繚瀛楶DF鍥炴墽
```

**骞村害姹囩畻娓呯即鎿嶄綔姝ラ锛堟瘡骞翠竴娆★紝5鏈?1鏃ュ墠锛夛細**

```
绗?姝ワ細鐧诲綍 鈫?婀栧寳鐪佺數瀛愮◣鍔″眬
绗?姝ワ細杩涘叆 鈫?銆屼紒涓氭墍寰楃◣骞村害鐢虫姤銆?
绗?姝ワ細鍏堝～闄勮〃锛圓101010~A107040锛夛紝鍐嶅～涓昏〃A100000
绗?姝ワ細閲嶇偣鏍稿锛氱撼绋庤皟鏁存槑缁嗚〃A105000
绗?姝ワ細绯荤粺鑷姩璁＄畻搴旇ˉ閫€绋庨
绗?姝ワ細鐢虫姤骞剁即娆撅紙濡傛湁琛ョ◣锛?
绗?姝ワ細鐣欏瓨鍏ㄥ鐢虫姤琛≒DF
```
            """)

    # ==========================================
    #  涓夈€佸埄娑﹁〃缂栧埗璇存槑
    # ==========================================
    elif guide_choice.startswith("涓夈€?):
        st.subheader("涓夈€佸埄娑﹁〃缂栧埗璇存槑锛堝皬浼佷笟浼氳鍑嗗垯锛?)

        with st.expander("3.1 浼氳鍑嗗垯涓庨€傜敤鑼冨洿", expanded=True):
            st.markdown("""
**姝︽眽閲戣壋榫欑鎶€鏈夐檺鍏徃閫傜敤锛氥€婂皬浼佷笟浼氳鍑嗗垯銆?*

| 鍑嗗垯绫诲瀷 | 閫傜敤瀵硅薄 | 姝︽眽閲戣壋榫欑鎶€ |
|----------|---------|---------------|
| 浼佷笟浼氳鍑嗗垯 | 澶т腑鍨嬩紒涓氥€佷笂甯傚叕鍙?| 鉂?|
| 灏忎紒涓氫細璁″噯鍒?| 灏忓井鍨嬩紒涓?| 鉁?|
| 涓綋宸ュ晢鎴蜂細璁″埗搴?| 涓綋鎴?| 鉂?|

**灏忎紒涓氫細璁″噯鍒欒鐐癸細**
- 绠€鍖栨牳绠楋紝鏃犻渶璁℃彁鍑忓€煎噯澶?
- 璧勪骇鎸夊巻鍙叉垚鏈閲?
- 鎵€寰楃◣璐圭敤閲囩敤搴斾粯绋庢娉曪紙涓嶇‘璁ら€掑欢鎵€寰楃◣锛?
- 鍒╂鼎琛ㄧ粨鏋勬瘮浼佷笟浼氳鍑嗗垯鏇寸畝娲?
            """)

        with st.expander("3.2 鍒╂鼎琛ㄩ」鐩€愰」瑙ｉ噴", expanded=False):
            st.markdown("""
**灏忎紒涓氫細璁″噯鍒欏埄娑﹁〃锛堟爣鍑嗘牸寮忥級锛?*

| 琛屾 | 椤圭洰 | 鍚箟 | 鏁版嵁鏉ユ簮 |
|------|------|------|---------|
| 绗?琛?| **涓€銆佽惀涓氭敹鍏?* | 涓昏惀涓氬姟鏀跺叆 + 鍏朵粬涓氬姟鏀跺叆 | 閾惰娴佹按+鍙戠エ |
| 绗?琛?| 鍑忥細钀ヤ笟鎴愭湰 | 涓昏惀涓氬姟鎴愭湰 + 鍏朵粬涓氬姟鎴愭湰 | 閲囪喘/杩涜揣/鏉愭枡鎴愭湰 |
| 绗?琛?| 绋庨噾鍙婇檮鍔?| 鍩庡缓绋?鏁欒偛璐归檮鍔?鍦版柟鏁欒偛闄勫姞+鍗拌姳绋?鎴夸骇绋庣瓑 | 绋庡姟鐢虫姤琛?|
| 绗?琛?| 閿€鍞垂鐢?| 閿€鍞浉鍏宠垂鐢紙骞垮憡銆佽繍杈撶瓑锛?| 閾惰娴佹按鍒嗙被 |
| 绗?琛?| 绠＄悊璐圭敤 | 宸ヨ祫绀句繚+鍔炲叕璐?宸梾+鎷涘緟绛?| 閾惰娴佹按+宸ヨ祫琛?|
| 绗?琛?| 璐㈠姟璐圭敤 | 鍒╂伅鏀舵敮+閾惰鎵嬬画璐?| 閾惰娴佹按 |
| 绗?琛?| 璧勪骇鍑忓€兼崯澶?| 灏忎紒涓氫細璁″噯鍒欎笅閫氬父涓? | - |
| 绗?琛?| 鍔狅細鎶曡祫鏀剁泭 | 瀵瑰鎶曡祫鍙栧緱鐨勫垎绾?鏀剁泭 | 閾惰娴佹按 |
| 绗?琛?| **浜屻€佽惀涓氬埄娑?* | = 绗?琛?- 绗?琛?- 绗?琛?- 绗?琛?- 绗?琛?- 绗?琛?- 绗?琛?+ 绗?琛?| 璁＄畻 |
| 绗?0琛?| 鍔狅細钀ヤ笟澶栨敹鍏?| 鏀垮簻琛ュ姪/缃氭鏀跺叆/鐩樼泩绛?| 閾惰娴佹按 |
| 绗?1琛?| 鍑忥細钀ヤ笟澶栨敮鍑?| 缃氭鏀嚭/鎹愯禒/璧勪骇鎹熷け | 閾惰娴佹按 |
| 绗?2琛?| **涓夈€佸埄娑︽€婚** | = 绗?琛?+ 绗?0琛?- 绗?1琛?| 璁＄畻 |
| 绗?3琛?| 鍑忥細鎵€寰楃◣璐圭敤 | 鍒╂鼎鎬婚 脳 5%锛堝皬鍨嬪井鍒╋級 | 璁＄畻 |
| 绗?4琛?| **鍥涖€佸噣鍒╂鼎** | = 绗?2琛?- 绗?3琛?| 鏈€缁堟垚鏋?|

**鈿狅笍 鍏抽敭鏍￠獙鍏紡锛?*
```
鍒╂鼎鎬婚(绗?2琛? = 浼佷笟鎵€寰楃◣棰勭即鐢虫姤琛ˋ200000 绗?琛?
钀ヤ笟鏀跺叆(绗?琛? = A200000 绗?琛?
钀ヤ笟鎴愭湰(绗?琛? = A200000 绗?琛?
```
            """)

        with st.expander("3.3 閾惰娴佹按 鈫?鍒╂鼎琛?鏄犲皠琛?, expanded=False):
            st.markdown("""
**浠庨摱琛屾祦姘村埌鍒╂鼎琛ㄧ殑鑷姩鍒嗙被閫昏緫锛?*

| 閾惰娴佹按鍏抽敭璇?| 鈫?鍒╂鼎琛ㄩ」鐩?| 鈫?鐢虫姤琛ㄨ娆?|
|---------------|-------------|------------|
| 璐ф/閿€鍞敹鍏?鏈嶅姟璐?鍜ㄨ璐?鏀舵 | **钀ヤ笟鏀跺叆** | 绗?琛?|
| 閲囪喘/杩涜揣/鏉愭枡鎴愭湰 | **钀ヤ笟鎴愭湰** | 绗?琛?|
| 绋庨噾/鍩庡缓绋?鏁欒偛璐归檮鍔?鍗拌姳绋?| **绋庨噾鍙婇檮鍔?* | 绗?琛?|
| 杩愯緭璐?骞垮憡璐?鎺ㄥ箍璐?| **閿€鍞垂鐢?* | 绗?琛?|
| 宸ヨ祫/绀句繚/鍔炲叕/宸梾/鎷涘緟/鎴跨/姘寸數 | **绠＄悊璐圭敤** | 绗?琛?|
| 鍒╂伅/閾惰鎵嬬画璐?| **璐㈠姟璐圭敤** | 绗?琛?|
| 鎶曡祫鏀剁泭/鍒嗙孩/鑲℃伅 | **鎶曡祫鏀剁泭** | 绗?琛?|
| 鏀垮簻琛ュ姪/缃氭鏀跺叆 | **钀ヤ笟澶栨敹鍏?* | 绗?0琛?|
| 缃氭鏀嚭/鎹愯禒 | **钀ヤ笟澶栨敮鍑?* | 绗?1琛?|

**娉ㄦ剰浜嬮」锛?*
- 閾惰娴佹按闇€瑕嗙洊鍏ㄩ儴鏀舵敮锛屽鏈夌幇閲戜氦鏄撻渶琛ュ厖鐧昏
- 宸ヨ祫绀句繚鏁版嵁闇€浠庡伐璧勮〃瀵煎叆锛岄摱琛屾祦姘翠粎鍙嶆槧瀹炲彂閲戦
- 绋庨噾鍙婇檮鍔犻渶浠庣◣鍔＄敵鎶ヨ〃鍙栨暟锛岄摱琛屾祦姘翠腑鐨勭◣娆炬敮鍑轰粎渚涘弬鑰?
            """)

        with st.expander("3.4 鍒╂鼎琛ㄧ紪鍒舵楠?, expanded=False):
            st.markdown("""
**鏈堝害/瀛ｅ害鍒╂鼎琛ㄧ紪鍒舵祦绋嬶細**

```
绗?姝ワ細鏀堕泦鍘熷鍑瘉
  鈹溾攢 閾惰娴佹按锛堟墍鏈夊鍏处鎴凤級
  鈹溾攢 宸ヨ祫鍙戞斁琛紙鍚ぞ淇濇槑缁嗭級
  鈹溾攢 鍙戠エ鍙拌处锛堟敹鍏?鏀嚭锛?
  鈹斺攢 绋庡姟缂寸撼璁板綍

绗?姝ワ細鍒嗙被姹囨€?
  鈹溾攢 浣跨敤鏈郴缁熺殑銆岄摱琛屾祦姘磋嚜鍔ㄥ垎绫汇€嶅姛鑳?
  鈹溾攢 鎴栨墜鍔ㄦ寜鍒╂鼎琛ㄩ」鐩綊绫?
  鈹斺攢 纭繚姣忎竴绗斾氦鏄撻兘鏈夋纭殑鍒╂鼎琛ㄥ綊灞?

绗?姝ワ細濉〃璁＄畻
  鈹溾攢 鍚勯」鐩眹鎬婚噾棰濆～鍏ュ搴旇娆?
  鈹溾攢 璁＄畻钀ヤ笟鍒╂鼎锛堢9琛岋級= 鏀跺叆 - 鎴愭湰 - 绋庤垂 - 璐圭敤 + 鎶曡祫鏀剁泭
  鈹溾攢 璁＄畻鍒╂鼎鎬婚锛堢12琛岋級= 钀ヤ笟鍒╂鼎 + 钀ヤ笟澶栨敹鍏?- 钀ヤ笟澶栨敮鍑?
  鈹斺攢 璁＄畻鎵€寰楃◣锛堢13琛岋級= 鍒╂鼎鎬婚 脳 5%

绗?姝ワ細浜ゅ弶鏍￠獙
  鈹溾攢 鍒╂鼎琛ㄧ12琛岋紙鍒╂鼎鎬婚锛?= 鐢虫姤琛ˋ200000绗?琛?
  鈹溾攢 鍒╂鼎琛ㄧ1琛岋紙钀ヤ笟鏀跺叆锛?= 鐢虫姤琛ˋ200000绗?琛?
  鈹斺攢 鍒╂鼎琛ㄧ2琛岋紙钀ヤ笟鎴愭湰锛?= 鐢虫姤琛ˋ200000绗?琛?

绗?姝ワ細褰掓。鐣欏瓨
  鈹溾攢 鍒╂鼎琛ㄥ師浠讹紙Excel/PDF锛?
  鈹溾攢 閾惰娴佹按瀹屾暣鐗?
  鈹溾攢 鍙戠エ娓呭崟
  鈹斺攢 鐢虫姤鍥炴墽
```
            """)

        with st.expander("3.5 甯歌閿欒涓庣籂姝?, expanded=False):
            st.markdown("""
| 甯歌閿欒 | 姝ｇ‘鍋氭硶 | 褰卞搷 |
|----------|---------|------|
| 鏀跺叆鎴愭湰娣锋穯锛堝噣棰濆叆璐︼級 | 鏀跺叆鎴愭湰鍒嗗紑鍒楃ず锛屼笉鍙涧宸?| 浣庝及鏀跺叆瑙勬ā |
| 绠＄悊璐圭敤鍜岄攢鍞垂鐢ㄦ贩鐢?| 鎸夎垂鐢ㄦ€ц川鍒嗙被 | 鍒╂鼎琛ㄧ粨鏋勪笉鍑嗙‘ |
| 绋庨噾鍙婇檮鍔犲繕璁板綍鍏?| 姣忓鏈牴鎹鍊肩◣璁＄畻闄勫姞绋?| 鎴愭湰灏戣銆佸埄娑﹁啫鑳€ |
| 绀句繚璐规紡璁?| 姣忔湀绀句繚缂磋垂鍚庡強鏃跺叆璐?| 璐圭敤涓嶅畬鏁?|
| 閾惰鎵嬬画璐瑰拷鐣?| 妫€鏌ユ瘡绗旈摱琛屾祦姘翠腑鐨勬墜缁垂 | 璐㈠姟璐圭敤涓嶅畬鏁?|
| 鍒╂鼎琛ㄤ笌鐢虫姤琛ㄤ笉涓€鑷?| 浜ゅ弶鏍￠獙鍚庡啀鐢虫姤 | 鐢虫姤椋庨櫓 |
| 浜忔崯鏈堜唤涓嶅仛璐?| 浜忔崯涔熻瀹屾暣璁拌处闆剁敵鎶?| 绋庡姟鍚堣椋庨櫓 |
            """)

    # ==========================================
    #  鍥涖€佸鍊肩◣鍙婇檮鍔犵◣绛夌◣鐩?
    # ==========================================
    else:
        st.subheader("鍥涖€佸鍊肩◣鍙婇檮鍔犵◣绛夌◣鐩敵鎶ヨ鏄?)

        with st.expander("4.1 绋庣鎬昏", expanded=True):
            st.markdown("""
**姝︽眽閲戣壋榫欑鎶€鏈夐檺鍏徃娑夊強鐨勭◣绉嶏紙灏忚妯＄撼绋庝汉锛夛細**

| 搴忓彿 | 绋庣 | 绋庣巼/寰佹敹鐜?| 鐢虫姤鍛ㄦ湡 | 鎴鏃ユ湡 | 浼樻儬鏀跨瓥 |
|------|------|-----------|---------|---------|---------|
| 1 | 澧炲€肩◣ | 1%锛堝噺鎸夛級 | 瀛ｅ害 | 瀛ｅ悗15鏃?| 瀛ｂ墹30涓囧厤绋?|
| 2 | 鍩庡缓绋?| 3.5%锛堝噺鍗婏級 | 瀛ｅ害 | 闅忓鍊肩◣ | 鍏◣涓よ垂鍑忓崐 |
| 3 | 鏁欒偛璐归檮鍔?| 1.5%锛堝噺鍗婏級 | 瀛ｅ害 | 闅忓鍊肩◣ | 鍏◣涓よ垂鍑忓崐 |
| 4 | 鍦版柟鏁欒偛闄勫姞 | 1%锛堝噺鍗婏級 | 瀛ｅ害 | 闅忓鍊肩◣ | 鍏◣涓よ垂鍑忓崐 |
| 5 | 浼佷笟鎵€寰楃◣ | 5%锛堝皬寰級 | 瀛ｅ害棰勭即+骞村害 | 瀛ｅ悗15鏃?娆″勾5.31 | 灏忓瀷寰埄浼樻儬 |
| 6 | 涓汉鎵€寰楃◣ | 3%-45%绱繘 | 鏈堝害 | 娆℃湀15鏃?| 璧峰緛鐐?000鍏?|
| 7 | 鍗拌姳绋?| 鍚勭◣鐩噺鍗?| 鎸夋/鎸夋湡 | 鍙戠敓鏃?| 璧勯噾璐︾翱0.0125%/璐攢0.015%绛?|
| 8 | 娈嬩繚閲?| 鍏嶅緛 | 骞村害 | 骞村害 | 鈮?0浜哄厤寰?|
            """)

        with st.expander("4.2 澧炲€肩◣鐢虫姤锛堝皬瑙勬ā绾崇◣浜猴級", expanded=False):
            st.markdown("""
**灏忚妯＄撼绋庝汉澧炲€肩◣鏀跨瓥锛?026骞绰锋箹鍖楋級锛?*

| 鍦烘櫙 | 寰佹敹鐜?| 鏀跨瓥渚濇嵁 |
|------|--------|---------|
| 瀛ｅ害涓嶅惈绋庢敹鍏?鈮?30涓囧厓 | **鍏嶅緛** | 璐㈢◣銆?023銆?9鍙?|
| 瀛ｅ害涓嶅惈绋庢敹鍏?> 30涓囧厓 | **鍑忔寜1%**锛堝師3%锛?| 璐㈢◣銆?023銆?9鍙?|

**鍏抽敭姒傚康锛?*
```
鍚◣鏀跺叆 鈫?涓嶅惈绋庢敹鍏ワ細
  涓嶅惈绋庢敹鍏?= 鍚◣鏀跺叆 梅 (1 + 寰佹敹鐜?

涓句緥锛?
  瀛ｅ害鍚◣鏀跺叆 50涓囧厓
  涓嶅惈绋庢敹鍏?= 500,000 梅 1.03 鈮?485,436.89 鍏?
  > 30涓囧厓 鈫?搴旂即澧炲€肩◣ = 485,436.89 脳 1% 鈮?4,854.37 鍏?
```

**鐢虫姤骞冲彴**锛氭箹鍖楃渷鐢靛瓙绋庡姟灞€ 鈫?銆屽鍊肩◣鍙婇檮鍔犵◣璐圭敵鎶ワ紙灏忚妯＄撼绋庝汉閫傜敤锛夈€?

**鐢虫姤姝ラ锛?*
```
绗?姝ワ細鐧诲綍 鈫?婀栧寳鐪佺數瀛愮◣鍔″眬
绗?姝ワ細閫夋嫨 鈫?銆屽鍊肩◣鍙婇檮鍔犵◣璐圭敵鎶ワ紙灏忚妯＄撼绋庝汉锛夈€?
绗?姝ワ細濉姤 鈫?鍚◣閿€鍞 鈫?绯荤粺鑷姩鎹㈢畻涓嶅惈绋庢敹鍏?
绗?姝ワ細鍏嶇◣鍒ゆ柇 鈫?鈮?0涓囪嚜鍔ㄥ厤绋庯紝>30涓囨寜1%璁＄◣
绗?姝ワ細闄勫姞绋?鈫?绯荤粺鑷姩璁＄畻鍩庡缓绋?鏁欒偛璐归檮鍔?鍦版柟鏁欒偛闄勫姞
绗?姝ワ細鏍稿 鈫?纭鏃犺鍚庣敵鎶ュ苟缂存
```

**鈿狅笍 鐗瑰埆娉ㄦ剰锛?*
- 鍗充娇浜彈鍏嶇◣浼樻儬锛屼篃蹇呴』瀹屾垚鐢虫姤锛堝～鎶ュ悗绋庨涓?锛?
- 澧炲€肩◣鍙戠エ闇€閫氳繃绋庢帶绯荤粺寮€鍏凤紙鍏ㄧ數鍙戠エ/鐢靛瓙鍙戠エ锛?
- 鍏嶇◣鏀跺叆瀵瑰簲鐨勯檮鍔犵◣鍚屾涓洪浂
            """)

        with st.expander("4.3 闄勫姞绋庤瑙ｏ紙鍏◣涓よ垂鍑忓崐锛?, expanded=False):
            st.markdown("""
**闄勫姞绋?= 浠ュ疄闄呯即绾崇殑澧炲€肩◣涓鸿绋庡熀纭€锛?*

| 闄勫姞绋庣 | 鍚嶄箟绋庣巼 | 鍑忓崐鍚庣◣鐜?| 璁＄畻鍏紡 |
|----------|---------|-----------|---------|
| 鍩庡缓绋庯紙甯傚尯锛?| 7% | **3.5%** | 澧炲€肩◣ 脳 3.5% |
| 鏁欒偛璐归檮鍔?| 3% | **1.5%** | 澧炲€肩◣ 脳 1.5% |
| 鍦版柟鏁欒偛闄勫姞 | 2% | **1%** | 澧炲€肩◣ 脳 1% |
| **闄勫姞绋庡悎璁?* | **12%** | **6%** | 澧炲€肩◣ 脳 6% |

**銆屽叚绋庝袱璐广€嶅噺鍗婃斂绛栬鏄庯細**

| 瑕佺礌 | 璇存槑 |
|------|------|
| 閫傜敤涓讳綋 | 灏忚妯＄撼绋庝汉銆佸皬鍨嬪井鍒╀紒涓氥€佷釜浣撳伐鍟嗘埛 |
| 姝︽眽閲戣壋榫欑鎶€ | 鉁?鏃㈡槸灏忚妯＄撼绋庝汉锛屽張鏄皬鍨嬪井鍒╀紒涓?|
| 鍑忓崐鑼冨洿 | 鍩庡缓绋庛€佹暀鑲茶垂闄勫姞銆佸湴鏂规暀鑲查檮鍔犮€佸嵃鑺辩◣銆佹埧浜х◣銆佸煄闀囧湡鍦颁娇鐢ㄧ◣銆佽€曞湴鍗犵敤绋庛€佽祫婧愮◣ |
| 鏀跨瓥渚濇嵁 | 璐㈢◣銆?022銆?0鍙?|
| 鏈夋晥鏈?| 鑷?027骞?2鏈?1鏃?|
| 鐢宠鏂瑰紡 | 鏃犻渶鐢宠锛屾箹鍖楃渷銆屽厤鐢冲嵆浜€?|

**涓句緥锛?*
```
澧炲€肩◣搴旂即 4,854.37 鍏?

鍑忓崐鍓嶏細                   鍑忓崐鍚庯細
  鍩庡缓绋?  4,854.37脳7% = 339.81    鍩庡缓绋?  4,854.37脳3.5% = 169.90
  鏁欒偛闄勫姞 4,854.37脳3% = 145.63    鏁欒偛闄勫姞 4,854.37脳1.5% =  72.82
  鍦版柟鏁欒偛 4,854.37脳2% =  97.09    鍦版柟鏁欒偛 4,854.37脳1%  =  48.54
  鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€    鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€
  鍚堣                582.53 鍏?    鍚堣                   291.26 鍏?

  鍑忓厤閲戦锛?82.53 - 291.26 = 291.27 鍏?鉁?
```
            """)

        with st.expander("4.4 鍏朵粬绋庣璇存槑", expanded=False):
            st.markdown("""
**馃摐 鍗拌姳绋庯紙銆婁腑鍗庝汉姘戝叡鍜屽浗鍗拌姳绋庢硶銆嬄?2022骞?鏈?鏃ユ柦琛岋級锛?*

| 绋庣洰 | 绋庣巼 | 璁＄◣鍩虹 | 浼樻儬鍚庯紙鍑忓崐锛?| 璇存槑 |
|------|------|---------|--------------|------|
| **璧勯噾璐︾翱** | 0.025%锛堜竾鍒嗕箣浜岀偣浜旓級 | 瀹炴敹璧勬湰 + 璧勬湰鍏Н | **0.0125%** | 鈿狅笍 娉ㄥ唽璧勬湰鍒颁綅/澧炶祫鏃剁即绾筹紝宸茬即閮ㄥ垎涓嶉噸澶?|
| 璐攢鍚堝悓 | 0.03%锛堜竾鍒嗕箣涓夛級 | 璐攢鍚堝悓閲戦 | **0.015%** | 鎸夊姹囨€荤敵鎶?|
| 鍊熸鍚堝悓 | 0.005%锛堝崄涓囧垎涔嬩簲锛?| 鍊熸閲戦 | **0.0025%** | 閾惰璐锋/铻嶈祫 |
| 鎶€鏈悎鍚?| 0.03%锛堜竾鍒嗕箣涓夛級 | 鎶€鏈悎鍚岄噾棰?| **0.015%** | 鎶€鏈紑鍙?杞/鍜ㄨ |
| 璐骇绉熻祦鍚堝悓 | 0.1%锛堝崈鍒嗕箣涓€锛?| 绉熻祦閲戦 | **0.05%** | 鎴垮眿/璁惧绉熻祦 |
| 鍏朵粬璐︾翱 | 鍏嶅緛 | 鈥?| 鍏嶅緛 | 璐㈢◣銆?018銆?0鍙凤紝宸插厤 |

**鈿狅笍 璧勯噾璐︾翱鍗拌姳绋庤瑙ｏ紙鏈€閲嶈锛夛細**

| 鍦烘櫙 | 璁＄◣鍩虹 | 涓句緥 |
|------|---------|------|
| 鍏徃鎴愮珛路娉ㄥ唽璧勬湰棣栨鍒颁綅 | 瀹炴敹璧勬湰 + 璧勬湰鍏Н鍏ㄩ | 娉ㄥ唽璧勬湰100涓囷紝瀹炵即鍒颁綅 鈫?100涓?脳 0.0125% = **125鍏?* |
| 鍚庣画澧炶祫 | 浠呭**鏂板瀹炴敹璧勬湰**閮ㄥ垎 | 浠?00涓囧璧勫埌200涓?鈫?100涓?脳 0.0125% = **125鍏?* |
| 娉ㄥ唽璧勬湰鏈叏棰濆埌浣?| 鎸夊疄闄呭埌浣嶉噾棰?| 娉ㄥ唽璧勬湰100涓囷紝鍙埌浣?0涓?鈫?50涓?脳 0.0125% = **62.5鍏?* |
| 宸茬即杩囧嵃鑺辩◣鐨勮祫鏈?| **涓嶉噸澶嶅緛鏀?* | 鏃犻渶鍐嶆缂寸撼 |

- 璧勯噾璐︾翱鍗拌姳绋庡湪鐢靛瓙绋庡姟灞€銆屽嵃鑺辩◣鐢虫姤銆嶄腑閫夋嫨銆岃祫閲戣处绨裤€嶇◣鐩?
- 澧炶祫鍙樻洿鍚庨渶鍦ㄥ伐鍟嗗彉鏇寸櫥璁板畬鎴愬悗鐨勭撼绋庢湡鍐呯敵鎶?
- 鐢虫姤鏂瑰紡锛氭寜娆℃垨鎸夋湡姹囨€?
- 鍏◣涓よ垂鍑忓崐鍚庯紝浠ヤ笂鍚勭◣鐩潎鎸夊崐棰濆緛鏀?
- 灏忛鍗拌姳绋庡彲閫氳繃鐢靛瓙绋庡姟灞€銆岀畝骞剁敵鎶ャ€嶅姛鑳芥眹鎬荤即绾?

**娈嬬柧浜哄氨涓氫繚闅滈噾锛堟畫淇濋噾锛夛細**

| 瑕佺礌 | 璇存槑 |
|------|------|
| 璁＄畻鍏紡 | 锛堜笂骞磋亴宸ヤ汉鏁懊?.5% - 瀹為檯娈嬬柧鑱屽伐鏁帮級脳 涓婂勾鑱屽伐骞冲潎宸ヨ祫 |
| 鍏嶅緛鏉′欢 | 鍦ㄨ亴鑱屽伐 鈮?30浜?鈫?鍏嶅緛 鉁?|
| 鏀跨瓥渚濇嵁 | 鍙戞敼浠锋牸瑙勩€?019銆?015鍙?|
| 鐢虫姤鏃堕棿 | 骞村害锛堥€氬父姣忓勾7-9鏈堬級 |

姝︽眽閲戣壋榫欑鎶€鍦ㄨ亴鑱屽伐 鈮?30浜猴紝绗﹀悎鍏嶅緛鏉′欢锛屽彧闇€杩涜銆岄浂鐢虫姤銆嶅嵆鍙€?

**鎴夸骇绋?/ 鍩庨晣鍦熷湴浣跨敤绋庯細**
- 濡傚叕鍙稿悕涓嬫棤鑷湁鎴夸骇/鍦熷湴锛屾棤闇€缂寸撼
- 濡傜璧佸姙鍏満鎵€锛岀敱鍑虹鏂圭即绾筹紝鎵跨鏂规棤闇€鐢虫姤
            """)

        with st.expander("4.5 浼樻儬鏀跨瓥鎬昏锛堟箹鍖楃渷2026锛?, expanded=False):
            st.markdown("""
**姝︽眽閲戣壋榫欑鎶€鏈夐檺鍏徃褰撳墠閫傜敤鐨勫叏閮ㄧ◣鏀朵紭鎯狅細**

| 搴忓彿 | 绋庣 | 浼樻儬鍚嶇О | 浼樻儬鍐呭 | 鏀跨瓥渚濇嵁 | 鏈夋晥鏈?|
|------|------|---------|---------|---------|--------|
| 1 | 澧炲€肩◣ | 灏忚妯＄撼绋庝汉鍏嶇◣ | 瀛ｂ墹30涓囧厤绋?| 璐㈢◣銆?023銆?9鍙?| 鑷?027.12.31 |
| 2 | 澧炲€肩◣ | 鍑忔寜1%寰佹敹 | 瀛?30涓囨寜1% | 璐㈢◣銆?023銆?9鍙?| 鑷?027.12.31 |
| 3 | 鍩庡缓绋?| 鍏◣涓よ垂鍑忓崐 | 7%鈫?.5% | 璐㈢◣銆?022銆?0鍙?| 鑷?027.12.31 |
| 4 | 鏁欒偛璐归檮鍔?| 鍏◣涓よ垂鍑忓崐 | 3%鈫?.5% | 璐㈢◣銆?022銆?0鍙?| 鑷?027.12.31 |
| 5 | 鍦版柟鏁欒偛闄勫姞 | 鍏◣涓よ垂鍑忓崐 | 2%鈫?% | 璐㈢◣銆?022銆?0鍙?| 鑷?027.12.31 |
| 6 | 浼佷笟鎵€寰楃◣ | 灏忓瀷寰埄浼樻儬 | 瀹為檯绋庤礋5% | 璐㈢◣銆?023銆?2鍙?| 鑷?027.12.31 |
| 7 | 鍗拌姳绋?| 灏忓瀷寰埄鍑忓崐 | 鍚勭◣鐩噺鍗?| 璐㈢◣銆?022銆?0鍙?| 鑷?027.12.31 |
| 8 | 娈嬩繚閲?| 灏忓井浼佷笟鍏嶅緛 | 鈮?0浜哄厤缂?| 鍙戞敼浠锋牸瑙勩€?019銆?015鍙?| 闀挎湡 |

> 馃幆 **婀栧寳鐪佺壒鑹?*锛氥€屽厤鐢冲嵆浜€嶁€斺€?浠ヤ笂鎵€鏈変紭鎯犲湪鐢虫姤鏃剁郴缁熻嚜鍔ㄥ垽瀹氬拰鍑忓厤锛屾棤闇€棰濆鎻愪氦鐢宠鎴栧妗堟潗鏂欍€?
            """)

        with st.expander("4.6 鐢虫姤鏃ュ巻涓庣即娆炬祦绋?, expanded=False):
            st.markdown("""
**鏈堝害/瀛ｅ害鐢虫姤鏃堕棿绾匡細**

```
姣忔湀 1~15鏃ワ細
  鈹斺攢 涓汉鎵€寰楃◣锛堜唬鎵ｄ唬缂达級鈥?鑷劧浜虹數瀛愮◣鍔″眬鎵ｇ即绔?

姣忓搴︾粨鏉熷悗 15鏃ュ唴锛?
  鈹溾攢 澧炲€肩◣鍙婇檮鍔犵◣ 鈥?婀栧寳鐪佺數瀛愮◣鍔″眬
  鈹溾攢 浼佷笟鎵€寰楃◣棰勭即 鈥?婀栧寳鐪佺數瀛愮◣鍔″眬
  鈹斺攢 鍗拌姳绋庯紙濡傛湁锛夆€?婀栧寳鐪佺數瀛愮◣鍔″眬

姣忓勾 1鏈?鏃6鏈?0鏃ワ細
  鈹斺攢 宸ュ晢骞存姤 鈥?鍥藉浼佷笟淇＄敤淇℃伅鍏ず绯荤粺

姣忓勾 5鏈?1鏃ュ墠锛?
  鈹斺攢 浼佷笟鎵€寰楃◣骞村害姹囩畻娓呯即 鈥?婀栧寳鐪佺數瀛愮◣鍔″眬

姣忓勾 7~9鏈堬細
  鈹斺攢 娈嬩繚閲戠敵鎶ワ紙闆剁敵鎶ワ級鈥?婀栧寳鐪佺數瀛愮◣鍔″眬
```

**缂存鏂瑰紡锛?*

| 鏂瑰紡 | 璇存槑 | 鎺ㄨ崘 |
|------|------|------|
| 涓夋柟鍗忚鎵ｆ | 浼佷笟-閾惰-绋庡姟涓夋柟绛剧害锛岃嚜鍔ㄦ墸娆?| 猸?鎺ㄨ崘 |
| 閾惰绔煡璇㈢即绋?| 鍦ㄧ數瀛愮◣鍔″眬鐢熸垚缂存涔︼紝鍒伴摱琛屾煖鍙扮即娆?| 澶囩敤 |
| 缃戦摱鐩存帴缂存 | 閮ㄥ垎閾惰鏀寔鐢靛瓙绋庡姟灞€鍐呭祵缂存 | 鍙€?|

**鐢虫姤鍚庣殑妫€鏌ユ竻鍗曪細**
- [ ] 鐢虫姤鐘舵€侊細鐢虫姤鎴愬姛
- [ ] 缂存鐘舵€侊細宸茬即娆撅紙鎴栭浂鐢虫姤鏃犻渶缂存锛?
- [ ] 鍥炴墽涓嬭浇锛氬凡鐣欏瓨PDF
- [ ] 鐢虫姤琛ㄦ墦鍗板綊妗?
- [ ] 濡傛秹鍙婇€€绋庯紝璺熻釜閫€绋庤繘搴?
            """)

    # ===== 搴曢儴閫氱敤涓嬭浇 =====
    st.divider()
    st.subheader("馃摜 涓嬭浇瀹屾暣鐢虫姤鎸囧崡")

    # 鐢熸垚瀹屾暣鎸囧崡鏂囨湰
    full_guide = """================================================================
  姝︽眽閲戣壋榫欑鎶€鏈夐檺鍏徃 路 绋庡姟鐢虫姤鎿嶄綔鎸囧崡锛?026骞村害锛?
  鍩轰簬灏忚妯＄撼绋庝汉 + 灏忓瀷寰埄浼佷笟 + 婀栧寳鐪佷紭鎯犳斂绛?
================================================================

涓€銆佷釜浜烘墍寰楃◣鐢虫姤锛堜唬鎵ｄ唬缂达級
================================================================

1. 鐢虫姤涓讳綋锛氭姹夐噾鑹抽緳绉戞妧鏈夐檺鍏徃锛堟墸缂翠箟鍔′汉锛?
2. 鐢虫姤鍛ㄦ湡锛氭湀搴?
3. 鐢虫姤骞冲彴锛氳嚜鐒朵汉鐢靛瓙绋庡姟灞€锛堟墸缂寸锛?
4. 鎴鏃ユ湡锛氭鏈?5鏃?

銆愯绠楀叕寮忥紙绱棰勬墸娉曪級銆?
  搴旂◣鏀跺叆 = 绋庡墠宸ヨ祫 - 涓汉绀句繚 - 5000锛堣捣寰佺偣锛?- 涓撻」闄勫姞鎵ｉ櫎
  搴旂撼绋庨 = 搴旂◣鏀跺叆 脳 绋庣巼 - 閫熺畻鎵ｉ櫎鏁?

銆?绾х疮杩涚◣鐜囥€?
  绾ф暟  搴旂撼绋庢墍寰楅      绋庣巼  閫熺畻鎵ｉ櫎鏁?
  1     鈮?,000            3%    0
  2     3,001~12,000      10%   210
  3     12,001~25,000     20%   1,410
  4     25,001~35,000     25%   2,660
  5     35,001~55,000     30%   4,410
  6     55,001~80,000     35%   7,160
  7     >80,000           45%   15,160

銆愪笓椤归檮鍔犳墸闄わ紙姣忔湀鏍囧噯锛夈€?
  瀛愬コ鏁欒偛锛?   2,000鍏?浜?
  濠村辜鍎跨収鎶わ細  2,000鍏?浜?
  璧″吇鑰佷汉锛?   鏈€楂?,000鍏?鏈?
  浣忔埧璐锋鍒╂伅锛?1,000鍏?鏈?
  浣忔埧绉熼噾锛?   1,500鍏?鏈堬紙姝︽眽锛?
  缁х画鏁欒偛锛?   400鍏?鏈?

銆愭搷浣滄楠ゃ€?
  1. 浜哄憳淇℃伅閲囬泦锛堟柊澧?淇敼鍛樺伐锛?
  2. 涓撻」闄勫姞鎵ｉ櫎涓嬭浇鏇存柊
  3. 缁煎悎鎵€寰楃敵鎶?鈫?姝ｅ父宸ヨ祫钖噾 鈫?绋庢璁＄畻 鈫?鐢虫姤琛ㄦ姤閫?
  4. 绋庢缂寸撼锛堥浂鐢虫姤涔熼渶瀹屾垚鎶ラ€侊級

銆愰浂鐢虫姤銆?
  鎵€鏈夊憳宸ュ簲绋庢敹鍏ヤ负0鏃朵粛闇€瀹屾垚鐢虫姤锛屼笉鍙仐婕忋€?


浜屻€佷紒涓氭墍寰楃◣瀛ｆ姤 & 骞存姤
================================================================

銆愬搴﹂缂达紙A200000琛級銆?
  鐢虫姤鍛ㄦ湡锛氭瘡瀛ｅ害涓€娆?
  鎴鏃ユ湡锛歈1鈫?/15, Q2鈫?/15, Q3鈫?0/15, Q4鈫掓骞?/15
  鐢虫姤骞冲彴锛氭箹鍖楃渷鐢靛瓙绋庡姟灞€

  A200000鍏抽敭琛屾锛?
    绗?琛?  钀ヤ笟鏀跺叆
    绗?琛?  钀ヤ笟鎴愭湰
    绗?琛?  鍒╂鼎鎬婚
    绗?琛?  瀹為檯鍒╂鼎棰?
    绗?0琛? 搴旂撼鎵€寰楃◣棰?= 绗?琛?脳 25%
    绗?1琛? 鍑忓厤鎵€寰楃◣棰濓紙灏忓井浼樻儬锛?
    绗?3琛? 鏈湡搴旇ˉ(閫€)绋庨

銆愬皬鍨嬪井鍒╀紒涓氫紭鎯犮€?
  鏉′欢锛氬勾鍒╂鼎鈮?00涓?+ 鍛樺伐鈮?00 + 璧勪骇鈮?000涓?
  瀹為檯绋庣巼锛?%锛堝噺鎸?5%脳20%锛?
  鏂瑰紡锛氱郴缁熻嚜鍔ㄥ垽瀹氾紝鏃犻渶鐢宠锛堝厤鐢冲嵆浜級

銆愬勾搴︽眹绠楁竻缂淬€?
  鎴鏃ユ湡锛氭骞?鏈?1鏃?
  鏍稿績閫昏緫锛?
    浼氳鍒╂鼎鎬婚
    + 绾崇◣璋冨椤?
    - 绾崇◣璋冨噺椤?
    - 寮ヨˉ浜忔崯
    = 搴旂撼绋庢墍寰楅 脳 5%
    - 宸查缂寸◣棰?
    = 搴旇ˉ(閫€)绋庨

銆愬伐鍟嗗勾鎶ャ€?
  骞冲彴锛氬浗瀹朵紒涓氫俊鐢ㄤ俊鎭叕绀虹郴缁?
  鏃堕棿锛氭瘡骞?鏈?鏃6鏈?0鏃?
  鍐呭锛氬熀鏈俊鎭€佽偂涓滃嚭璧勩€佽祫浜х姸鍐点€佺ぞ淇濅俊鎭瓑


涓夈€佸埄娑﹁〃缂栧埗璇存槑锛堝皬浼佷笟浼氳鍑嗗垯锛?
================================================================

銆愬埄娑﹁〃鏍囧噯鏍煎紡銆?
  绗?琛?  涓€銆佽惀涓氭敹鍏?
  绗?琛?    鍑忥細钀ヤ笟鎴愭湰
  绗?琛?    鍑忥細绋庨噾鍙婇檮鍔?
  绗?琛?    鍑忥細绠＄悊璐圭敤锛堝惈宸ヨ祫绀句繚绛夛級
  绗?琛?    鍑忥細璐㈠姟璐圭敤
  绗?琛?    鍔狅細鎶曡祫鏀剁泭
  绗?琛?  浜屻€佽惀涓氬埄娑?= 1-2-3-5-6+8
  绗?0琛?   鍔狅細钀ヤ笟澶栨敹鍏?
  绗?1琛?   鍑忥細钀ヤ笟澶栨敮鍑?
  绗?2琛? 涓夈€佸埄娑︽€婚 = 9+10-11
  绗?3琛?   鍑忥細鎵€寰楃◣璐圭敤 = 12脳5%锛堝皬鍨嬪井鍒╋級
  绗?4琛? 鍥涖€佸噣鍒╂鼎 = 12-13

銆愰摱琛屾祦姘?鈫?鍒╂鼎琛ㄦ槧灏勩€?
  璐ф/閿€鍞敹鍏?鏈嶅姟璐?    鈫?钀ヤ笟鏀跺叆锛堢1琛岋級
  閲囪喘/杩涜揣/鏉愭枡鎴愭湰       鈫?钀ヤ笟鎴愭湰锛堢2琛岋級
  绋庨噾/鍩庡缓绋?鍗拌姳绋?      鈫?绋庨噾鍙婇檮鍔狅紙绗?琛岋級
  宸ヨ祫/绀句繚/鍔炲叕/宸梾/鎷涘緟 鈫?绠＄悊璐圭敤锛堢5琛岋級
  鍒╂伅/閾惰鎵嬬画璐?         鈫?璐㈠姟璐圭敤锛堢6琛岋級
  鏀垮簻琛ュ姪/缃氭鏀跺叆        鈫?钀ヤ笟澶栨敹鍏ワ紙绗?0琛岋級
  缃氭鏀嚭/鎹愯禒            鈫?钀ヤ笟澶栨敮鍑猴紙绗?1琛岋級

銆愬叧閿牎楠屻€?
  鍒╂鼎琛ㄧ1琛岋紙钀ヤ笟鏀跺叆锛? 鐢虫姤琛ˋ200000 绗?琛?
  鍒╂鼎琛ㄧ2琛岋紙钀ヤ笟鎴愭湰锛? 鐢虫姤琛ˋ200000 绗?琛?
  鍒╂鼎琛ㄧ12琛岋紙鍒╂鼎鎬婚锛? 鐢虫姤琛ˋ200000 绗?琛?


鍥涖€佸鍊肩◣鍙婇檮鍔犵◣绛夌◣鐩敵鎶?
================================================================

銆愮◣绉嶆€昏锛堝皬瑙勬ā绾崇◣浜猴級銆?
  绋庣            绋庣巼        鍛ㄦ湡    鎴      浼樻儬
  澧炲€肩◣          1%锛堝噺鎸夛級  瀛ｅ害    瀛ｅ悗15鏃? 瀛ｂ墹30涓囧厤绋?
  鍩庡缓绋?         3.5%锛堝噺鍗婏級瀛ｅ害    闅忓鍊肩◣  鍏◣涓よ垂鍑忓崐
  鏁欒偛璐归檮鍔?     1.5%锛堝噺鍗婏級瀛ｅ害    闅忓鍊肩◣  鍏◣涓よ垂鍑忓崐
  鍦版柟鏁欒偛闄勫姞    1%锛堝噺鍗婏級  瀛ｅ害    闅忓鍊肩◣  鍏◣涓よ垂鍑忓崐
  浼佷笟鎵€寰楃◣      5%锛堝皬寰級  瀛ｅ害+骞?瀛ｅ悗/5.31  灏忓瀷寰埄浼樻儬
  涓汉鎵€寰楃◣      3%-45%      鏈堝害    娆℃湀15鏃?  璧峰緛鐐?000
  鍗拌姳绋?         0.03%鍑忓崐   鎸夋    -         灏忓井鍑忓崐
  娈嬩繚閲?         鍏嶅緛        骞村害    -         鈮?0浜哄厤寰?

銆愬鍊肩◣璁＄畻銆?
  涓嶅惈绋庢敹鍏?= 鍚◣鏀跺叆 梅 (1+寰佹敹鐜?
  瀛ｂ墹30涓?鈫?鍏嶇◣
  瀛?30涓?鈫?搴旂即 = 涓嶅惈绋庢敹鍏?脳 1%

銆愰檮鍔犵◣ = 澧炲€肩◣ 脳 浼樻儬鍚庣◣鐜囥€?
  鍩庡缓绋?     3.5%锛堝師7%鍑忓崐锛?
  鏁欒偛璐归檮鍔? 1.5%锛堝師3%鍑忓崐锛?
  鍦版柟鏁欒偛闄勫姞 1%锛堝師2%鍑忓崐锛?
  鍚堣        6%锛堝師12%鍑忓崐锛?

銆愬叚绋庝袱璐瑰噺鍗婃斂绛栥€?
  閫傜敤锛氬皬瑙勬ā绾崇◣浜?+ 灏忓瀷寰埄浼佷笟 鉁?
  鑼冨洿锛氬煄寤虹◣銆佹暀鑲茶垂闄勫姞銆佸湴鏂规暀鑲查檮鍔犮€佸嵃鑺辩◣銆佹埧浜х◣銆?
        鍩庨晣鍦熷湴浣跨敤绋庛€佽€曞湴鍗犵敤绋庛€佽祫婧愮◣
  渚濇嵁锛氳储绋庛€?022銆?0鍙?
  鏈熼檺锛氳嚦2027骞?2鏈?1鏃?

銆愮敵鎶ュ钩鍙版眹鎬汇€?
  涓汉鎵€寰楃◣锛?    鑷劧浜虹數瀛愮◣鍔″眬锛堟墸缂寸锛?
  澧炲€肩◣鍙婇檮鍔狅細   婀栧寳鐪佺數瀛愮◣鍔″眬
  浼佷笟鎵€寰楃◣锛?    婀栧寳鐪佺數瀛愮◣鍔″眬
  鍗拌姳绋庯細         婀栧寳鐪佺數瀛愮◣鍔″眬
  娈嬩繚閲戯細         婀栧寳鐪佺數瀛愮◣鍔″眬
  宸ュ晢骞存姤锛?      鍥藉浼佷笟淇＄敤淇℃伅鍏ず绯荤粺

銆愭箹鍖楃渷浼樻儬鏀跨瓥鎬昏銆?
  1. 澧炲€肩◣锛氬皬瑙勬ā鍑忔寜1%锛屽鈮?0涓囧厤绋?
  2. 鍏◣涓よ垂锛?椤瑰噺鍗婂緛鏀?
  3. 浼佷笟鎵€寰楃◣锛氬皬鍨嬪井鍒╁疄闄?%
  4. 娈嬩繚閲戯細鈮?0浜哄厤寰?
  5. 鏂瑰紡锛氬厤鐢冲嵆浜紝绯荤粺鑷姩鍑忓厤

================================================================
  鐢熸垚鏃堕棿锛?"" + datetime.now().strftime("%Y-%m-%d %H:%M") + """
  閲戣壋榫橝I绋庡姟鍔╂墜 路 浠呬緵鍙傝€冿紝浠ョ◣鍔℃満鍏虫渶鏂板叕鍛婁负鍑?
================================================================"""

    dl_g1, dl_g2 = st.columns(2)
    with dl_g1:
        st.download_button(
            label="馃摜 涓嬭浇瀹屾暣鐢虫姤鎸囧崡锛圱XT锛?,
            data=full_guide,
            file_name=f"绋庡姟鐢虫姤鎿嶄綔鎸囧崡_姝︽眽閲戣壋榫欑鎶€_{datetime.now().strftime('%Y%m%d')}.txt",
            mime="text/plain",
            use_container_width=True,
        )
    with dl_g2:
        pdf_bytes = make_pdf(
            "绋庡姟鐢虫姤鎿嶄綔鎸囧崡 - 姝︽眽閲戣壋榫欑鎶€鏈夐檺鍏徃",
            full_guide.split("\n"),
            ""
        )
        if pdf_bytes:
            st.download_button(
                label="馃摜 涓嬭浇瀹屾暣鐢虫姤鎸囧崡锛圥DF锛?,
                data=pdf_bytes,
                file_name=f"绋庡姟鐢虫姤鎿嶄綔鎸囧崡_姝︽眽閲戣壋榫欑鎶€_{datetime.now().strftime('%Y%m%d')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

# ===============================================
#  Tab7锛氭畫鐤句汉灏变笟淇濋殰閲戠敵鎶?
# ===============================================

with tab6:
    st.header("鈾?娈嬬柧浜哄氨涓氫繚闅滈噾锛堟畫淇濋噾锛夌敵鎶?)
    st.caption("姝︽眽閲戣壋榫欑鎶€鏈夐檺鍏徃 路 灏忚妯＄撼绋庝汉 路 灏忓瀷寰埄浼佷笟")

    # 鈹€鈹€ 璇诲彇渚ц竟鏍忓弬鏁?鈹€鈹€
    def_prev_employees = st.session_state.get("def_prev_employees", 5)
    def_prev_disabled = st.session_state.get("def_prev_disabled", 0)
    def_prev_avg_salary = st.session_state.get("def_prev_avg_salary", 60000.0)
    def_local_avg_salary = st.session_state.get("def_local_avg_salary", 90000.0)
    def_year = st.session_state.get("def_year", 2026)

    # ===== 璁＄畻娈嬩繚閲?=====
    fund_data = calc_disabled_employment_fund(
        prev_year_employees=def_prev_employees,
        prev_year_disabled_employees=def_prev_disabled,
        prev_year_avg_salary=def_prev_avg_salary,
        local_avg_salary=def_local_avg_salary,
        year=def_year,
    )

    # ===== 鏀跨瓥閫熻鍖?=====
    col_a, col_b = st.columns(2)
    with col_a:
        st.metric("馃懃 涓婂勾鍦ㄨ亴鑱屽伐", f"{fund_data['涓婂勾鑱屽伐浜烘暟']} 浜?)
        st.metric("馃鈥嶐煢?瀹夋帓娈嬬柧浜?, f"{fund_data['涓婂勾娈嬬柧鑱屽伐浜烘暟']} 浜?)
        st.metric("馃搻 娉曞畾搴斿畨鎺掓瘮渚?, "1.5%")
    with col_b:
        st.metric("馃挵 鑱屽伐骞村潎宸ヨ祫", f"{fund_data['涓婂勾鑱屽伐骞村潎宸ヨ祫']:,.0f} 鍏?)
        st.metric("馃懁 搴斿畨鎺掍汉鏁?, f"{fund_data['搴斿畨鎺掍汉鏁?]:.2f} 浜?)
        if fund_data["鏄惁灏忓井浼佷笟鍏嶅緛"] == "鏄?鉁?:
            st.metric("馃帀 搴旂即娈嬩繚閲?, "0 鍏冿紙鍏嶅緛锛?, delta="鍏ㄩ鍑忓厤")
        else:
            st.metric("馃搵 搴旂即娈嬩繚閲?, f"{fund_data['搴旂即娈嬩繚閲?]:,.2f} 鍏?)

    st.divider()

    # ===== Part 1锛氬皬寰紒涓氬厤寰佸垽鏂?=====
    st.subheader("1锔忊儯 灏忓井浼佷笟鍏嶅緛鍒ゅ畾")

    if fund_data["鏄惁灏忓井浼佷笟鍏嶅緛"] == "鏄?鉁?:
        st.success(
            f"鉁?**鍏嶅緛娈嬬柧浜哄氨涓氫繚闅滈噾**\n\n"
            f"| 鏉′欢 | 鏁版嵁 | 鍒ゅ畾 |\n"
            f"|------|------|------|\n"
            f"| 鍦ㄨ亴鑱屽伐浜烘暟 | **{fund_data['涓婂勾鑱屽伐浜烘暟']} 浜?* | 鈮?30浜?鉁?|\n"
            f"| 鏀跨瓥渚濇嵁 | 鍙戞敼浠锋牸瑙勩€?019銆?015鍙?| 鈥?|\n\n"
            f"**缁撹锛氭偍鏃犻渶缂寸撼娈嬩繚閲戙€?* 浣嗛渶瑕佸湪瑙勫畾鏃堕棿鍐呯櫥褰曠數瀛愮◣鍔″眬瀹屾垚銆岄浂鐢虫姤銆嶃€?
        )
    else:
        st.warning(
            f"鈿狅笍 鍦ㄨ亴鑱屽伐 {fund_data['涓婂勾鑱屽伐浜烘暟']} 浜?> 30浜猴紝涓嶇鍚堝皬寰紒涓氬厤寰佹潯浠讹紝闇€鎸夎瀹氳绠楃即绾炽€?
        )

    # ===== Part 2锛氳绠楁槑缁?=====
    st.subheader("2锔忊儯 娈嬩繚閲戣绠楁槑缁?)

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("""
**鍩虹鍏紡锛?*
```
淇濋殰閲?= (涓婂勾鑱屽伐浜烘暟 脳 1.5% - 涓婂勾娈嬬柧鑱屽伐浜烘暟)
       脳 涓婂勾鑱屽伐骞村钩鍧囧伐璧?
       脳 鍒嗘。寰佹敹姣斾緥
```

**鍙傛暟浠ｅ叆锛?*
| 鍙傛暟 | 鏁板€?|
|------|------|
| (A) 涓婂勾鍦ㄨ亴鑱屽伐浜烘暟 | {} 浜?|
| (B) 娉曞畾姣斾緥 | 1.5% |
| (C) 搴斿畨鎺掍汉鏁?= A脳B | {:.2f} 浜?|
| (D) 瀹為檯瀹夋帓娈嬬柧浜?| {} 浜?|
| (E) 宸浜烘暟 = C-D | {:.2f} 浜?|
| (F) 鑱屽伐骞村潎宸ヨ祫 | {:,.0f} 鍏?|
| (G) 宸ヨ祫鍩烘暟锛堝皝椤跺悗锛?| {:,.0f} 鍏?|
| (H) 鍒嗘。寰佹敹姣斾緥 | {} |
""".format(
    fund_data["涓婂勾鑱屽伐浜烘暟"],
    fund_data["搴斿畨鎺掍汉鏁?],
    fund_data["涓婂勾娈嬬柧鑱屽伐浜烘暟"],
    fund_data["宸浜烘暟"],
    fund_data["涓婂勾鑱屽伐骞村潎宸ヨ祫"],
    fund_data["宸ヨ祫璁＄畻鍩烘暟"],
    fund_data["鍒嗘。寰佹敹姣斾緥"],
))

    with col2:
        if fund_data["搴旂即娈嬩繚閲?] > 0:
            st.markdown(f"""
**閫愭璁＄畻锛?*

1. 搴斿畨鎺掍汉鏁扮己鍙ｏ細
   `{fund_data['涓婂勾鑱屽伐浜烘暟']} 脳 1.5% - {fund_data['涓婂勾娈嬬柧鑱屽伐浜烘暟']} = {fund_data['宸浜烘暟']:.2f} 浜篳

2. 宸ヨ祫鍩烘暟锛堢ぞ骞?鍊嶅皝椤讹級锛?
   `min({fund_data['涓婂勾鑱屽伐骞村潎宸ヨ祫']:,.0f}, 绀惧钩宸ヨ祫脳2) = {fund_data['宸ヨ祫璁＄畻鍩烘暟']:,.0f} 鍏僠
   {fund_data['宸ヨ祫灏侀《璇存槑']}

3. 鍏ㄩ搴旂即锛?
   `{fund_data['宸浜烘暟']:.2f} 脳 {fund_data['宸ヨ祫璁＄畻鍩烘暟']:,.0f} = {fund_data['搴旂即娈嬩繚閲戯紙鍏ㄩ锛?]:,.2f} 鍏僠

4. {fund_data['鍒嗘。璇存槑']}

5. 瀹為檯搴旂即锛?
   `{fund_data['搴旂即娈嬩繚閲戯紙鍏ㄩ锛?]:,.2f} 脳 {fund_data['鍒嗘。寰佹敹姣斾緥']} = **{fund_data['搴旂即娈嬩繚閲?]:,.2f} 鍏?*`

6. 鍑忓厤閲戦锛?
   `{fund_data['搴旂即娈嬩繚閲戯紙鍏ㄩ锛?]:,.2f} - {fund_data['搴旂即娈嬩繚閲?]:,.2f} = {fund_data['鍑忓厤閲戦']:,.2f} 鍏僠 鉁?
""")
        else:
            st.info(
                f"**{fund_data['璁＄畻璇存槑']}**\n\n"
                f"灏忓井浼佷笟鍏嶅緛锛屽簲缂撮噾棰濅负 0 鍏冦€?
            )

    st.divider()

    # ===== Part 3锛氫紭鎯犳斂绛栧叏瑙?=====
    st.subheader("3锔忊儯 婀栧寳鐪?姝︽眽甯傛畫淇濋噾浼樻儬鏀跨瓥锛?026锛?)

    policy_cols = st.columns(3)
    with policy_cols[0]:
        st.markdown("""
**馃幆 鍏嶅緛鏉′欢锛堝叏棰濓級**

| 鏉′欢 | 閫傜敤 |
|------|------|
| 鍦ㄨ亴鑱屽伐 鈮?30浜?| 鉁?|
| 瀹夋帓娈嬬柧浜烘瘮渚?鈮?1.5% | 鈥?|

""")
        if def_prev_employees <= 30:
            st.success("鉁?婊¤冻鍏嶅緛鏉′欢")
        else:
            st.info("涓嶆弧瓒冲厤寰佹潯浠?)

    with policy_cols[1]:
        st.markdown("""
**馃搲 鍒嗘。鍑忓緛锛?30浜猴級**

| 瀹夋帓姣斾緥 | 寰佹敹姣斾緥 |
|----------|---------|
| 鈮?1.5% | 鍏嶅緛 |
| 1%~1.5% | 鍑忔寜 50% |
| < 1% | 鎸?90% |

""")

    with policy_cols[2]:
        st.markdown("""
**馃挵 宸ヨ祫鍩烘暟灏侀《**

| 椤圭洰 | 璇存槑 |
|------|------|
| 灏侀《绾?| 褰撳湴绀惧钩 脳 2 |
| 瓒呭皝椤?| 鎸夊皝椤惰绠?|
| 宸ヨ祫鍙ｅ緞 | 搴斿彂宸ヨ祫 |

""")

    st.markdown("""
| 搴忓彿 | 浼樻儬鏀跨瓥 | 鍐呭 | 鏀跨瓥渚濇嵁 | 鏈夋晥鏈?|
|------|---------|------|---------|--------|
| 1 | 灏忓井浼佷笟鍏嶅緛 | 鍦ㄨ亴鑱屽伐 鈮?30浜猴紝鏆傚厤寰佹敹 | 鍙戞敼浠锋牸瑙勩€?019銆?015鍙?| 闀挎湡 |
| 2 | 鍒嗘。鍑忓緛 | 杈炬爣1%鍑忓崐銆佷笉瓒?%鎸?0% | 璐㈢◣銆?019銆?8鍙?| 鑷?027.12.31 |
| 3 | 宸ヨ祫鍩烘暟灏侀《 | 瓒呯ぞ骞?鍊嶉儴鍒嗕笉璁″叆 | 璐㈢◣銆?018銆?9鍙?| 闀挎湡 |
| 4 | 婀栧寳鐪佸欢缁?| 涓婅堪浼樻儬寤剁画鎵ц | 婀栧寳鐪?025鍔╀紒鎺柦 | 鑷?027.12.31 |
""")

    # ===== Part 4锛氱敵鎶ユ祦绋?=====
    st.subheader("4锔忊儯 鐢虫姤鎿嶄綔娴佺▼")

    st.markdown("""
**鐢虫姤骞冲彴**锛氭箹鍖楃渷鐢靛瓙绋庡姟灞€ 鈫?銆岄潪绋庢敹鍏ラ€氱敤鐢虫姤銆嶆垨銆屾畫鐤句汉灏变笟淇濋殰閲戠敵鎶ャ€?

**鐢虫姤鏃堕棿**锛氭瘡骞?**7鏈?鏃?~ 9鏈?0鏃?*锛堜互褰撳湴绋庡姟鏈哄叧鍏憡涓哄噯锛?

**鎿嶄綔姝ラ锛堥浂鐢虫姤鍦烘櫙锛夛細**

```
绗?姝ワ細鐧诲綍 鈫?婀栧寳鐪佺數瀛愮◣鍔″眬 鈫?浼佷笟鐧诲綍
绗?姝ワ細杩涘叆 鈫?銆屾垜瑕佸姙绋庛€?鈫?銆岀◣璐圭敵鎶ュ強缂寸撼銆?
绗?姝ワ細鏌ユ壘 鈫?銆岄潪绋庢敹鍏ラ€氱敤鐢虫姤銆嶆垨鎼滅储銆屾畫淇濋噾銆?
绗?姝ワ細濉啓 鈫?涓婂勾鍦ㄨ亴鑱屽伐浜烘暟銆佸畨鎺掓畫鐤句汉鏁般€佸勾鍧囧伐璧?
绗?姝ワ細绯荤粺鑷姩鍒ゆ柇 鈫?鈮?0浜鸿嚜鍔ㄥ厤寰侊紙搴旂即=0锛?
绗?姝ワ細鐢虫姤 鈫?纭鏃犺鍚庣偣鍑汇€岀敵鎶ャ€?
绗?姝ワ細鎴 鈫?鍗充娇涓洪浂涔熼渶瑕佸畬鎴愭姤閫?
绗?姝ワ細鐣欏瓨 鈫?涓嬭浇鐢虫姤鍥炴墽PDF褰掓。
```

**鎿嶄綔姝ラ锛堥渶缂存鍦烘櫙锛?30浜猴級锛?*

```
绗?-4姝ワ細鍚屼笂
绗?姝ワ細绯荤粺鑷姩璁＄畻搴旂即閲戦
绗?姝ワ細鏍稿璁＄畻缁撴灉鏄惁姝ｇ‘
绗?姝ワ細鐢虫姤 鈫?纭鏃犺鍚庣偣鍑汇€岀敵鎶ャ€?
绗?姝ワ細缂存 鈫?閫氳繃涓夋柟鍗忚鎵ｆ鎴栭摱琛岀缂存
绗?姝ワ細鐣欏瓨 鈫?涓嬭浇瀹岀◣鍑瘉PDF 鈫?璐㈠姟璁拌处
```

**鈿狅笍 娉ㄦ剰浜嬮」锛?*
- **涓婂勾鏁版嵁** = 鐢虫姤骞村害鐨勪笂涓€骞达紙濡?026骞寸敵鎶ワ紝濉啓2025骞存暟鎹級
- **鍦ㄨ亴鑱屽伐浜烘暟** = 鍏ㄥ勾鍚勬湀骞冲潎浜烘暟锛堝惈瀛ｈ妭鎬х敤宸ラ渶鎶樼畻锛?
- **娈嬬柧鑱屽伐浜烘暟** = 椤荤粡娈嬭仈瀹℃牳纭鐨勬畫鐤句汉灏变笟浜烘暟
- 濡傚叕鍙告棤娈嬬柧鑱屽伐锛屽～鍐?0 浜?
- 閫炬湡涓嶇敵鎶ュ皢琚瀹氫负鏈畨缃畫鐤句汉锛屾寜鍏ㄩ璁＄畻骞跺彲鑳藉姞鏀舵粸绾抽噾
""")

    # ===== Part 5锛氬父瑙侀棶棰?=====
    st.subheader("5锔忊儯 甯歌闂")

    with st.expander("Q1锛氭姹夐噾鑹抽緳绉戞妧闇€瑕佺即绾虫畫淇濋噾鍚楋紵"):
        st.markdown(f"""
**绛旓細涓嶉渶瑕侊紙鍏嶅緛锛夈€?*

姝︽眽閲戣壋榫欑鎶€鏈夐檺鍏徃鍦ㄨ亴鑱屽伐 **{def_prev_employees} 浜?鈮?30浜?*锛岀鍚堝皬寰紒涓氬厤寰佹潯浠讹紙鍙戞敼浠锋牸瑙勩€?019銆?015鍙凤級銆?

浣嗗繀椤绘瘡骞村湪瑙勫畾鏃堕棿鍐咃紙7~9鏈堬級鐧诲綍鐢靛瓙绋庡姟灞€瀹屾垚闆剁敵鎶ユ墜缁紝涓嶅彲閬楁紡銆?
        """)

    with st.expander("Q2锛氶浂鐢虫姤涔熷繀椤诲仛鍚楋紵涓嶅仛浼氭€庢牱锛?):
        st.markdown("""
**绛旓細蹇呴』鍋氥€?*

鍗充娇搴旂即閲戦涓?锛屼篃蹇呴』鎸夋椂瀹屾垚鐢虫姤銆傞€炬湡鏈敵鎶ュ皢琚涓猴細
- 鏈寜瑙勫畾瀹夋帓娈嬬柧浜哄氨涓?
- 鍙兘琚寜鍏ㄩ璁＄畻骞惰拷缂存畫淇濋噾
- 绾冲叆淇＄敤璁板綍
- 褰卞搷浼佷笟绾崇◣淇＄敤绛夌骇
        """)

    with st.expander("Q3锛氬憳宸ヤ汉鏁版€庝箞绠楋紵鍖呭惈鑰佹澘鍚楋紵"):
        st.markdown("""
**绛旓細鍖呭惈鎵€鏈夊湪鑱岃亴宸ャ€?*

- 銆屽湪鑱岃亴宸ャ€? 涓庡崟浣嶇璁㈠姵鍔ㄥ悎鍚屻€佺敱鍗曚綅鏀粯宸ヨ祫鐨勬墍鏈変汉鍛?
- 鍖呮嫭锛氭寮忓憳宸ャ€佸悎鍚屽伐銆佸鑺傛€х敤宸ャ€佷复鏃跺伐锛堜笉鍚姵鍔℃淳閬ｏ級
- **鑰佹澘锛堟硶瀹氫唬琛ㄤ汉锛夊浠庡叕鍙搁鍙栧伐璧勶紝涔熻鍏ュ湪鍐?*
- 璁＄畻鏂瑰紡锛氬叏骞村悇鏈堝钩鍧?= 鍚勬湀鏈堟湯浜烘暟涔嬪拰 梅 12
- 瀛ｈ妭鎬х敤宸ラ渶鎶樼畻涓哄勾骞冲潎鐢ㄥ伐浜烘暟
        """)

    with st.expander("Q4锛氬彲浠ヨ仒鐢ㄦ畫鐤句汉鏉ユ姷鎵ｆ畫淇濋噾鍚楋紵"):
        st.markdown("""
**绛旓細鍙互銆?*

瀹夋帓娈嬬柧浜哄氨涓氭槸鍑忓厤娈嬩繚閲戠殑鏈€鏈夋晥鏂瑰紡锛?

| 瀹夋帓浜烘暟 | 鏁堟灉 |
|----------|------|
| 瀹夋帓1鍚嶉噸搴︽畫鐤句汉 | 鎸?鍚嶈绠?|
| 杈惧埌1.5%姣斾緥 | 鍏ㄩ鍏嶅緛 |

浣嗛渶娉ㄦ剰锛?
- 娈嬬柧鑱屽伐椤绘寔鏈夈€婃畫鐤句汉璇併€?
- 闇€绛捐1骞翠互涓婂姵鍔ㄥ悎鍚?
- 闇€瓒抽缂寸撼绀句繚
- 闇€鍒板綋鍦版畫鑱斿姙鐞嗐€屾寜姣斾緥灏变笟瀹℃牳銆嶇‘璁?
        """)

    with st.expander("Q5锛氬幓骞村垰鎴愮珛鐨勫叕鍙歌涓嶈鐢虫姤锛?):
        st.markdown("""
**绛旓細瑕併€?*

- 鎴愮珛涓嶈冻1骞寸殑锛屾寜瀹為檯鏈堜唤璁＄畻骞冲潎浜烘暟
- 浠嶉€傜敤鈮?0浜哄厤寰佹斂绛?
- 绗竴娆＄敵鎶ユ椂闇€鍏堝湪鐢靛瓙绋庡姟灞€鍋氱◣璐圭璁ゅ畾锛堝鏈嚜鍔ㄥ甫鍑猴級

**璁＄畻鍏紡**锛堟垚绔嬩笉瓒?骞达級锛?
```
鏈堝钩鍧囦汉鏁?= 鍚勬湀浜烘暟涔嬪拰 梅 瀹為檯缁忚惀鏈堟暟
```
        """)

    with st.expander("Q6锛氭畫淇濋噾鍜岀ぞ淇濋噷鐨勫伐浼や繚闄╂槸涓€鍥炰簨鍚楋紵"):
        st.markdown("""
**绛旓細涓嶆槸銆?*

| 椤圭洰 | 鎬ц川 | 寰佹敹閮ㄩ棬 | 鐢ㄩ€?|
|------|------|---------|------|
| 宸ヤ激淇濋櫓 | 绀句繚浜旈櫓涔嬩竴 | 绀句繚灞€ | 鍥犲伐鍙椾激璧斿伩 |
| 娈嬩繚閲?| 鏀垮簻鎬у熀閲?| 绋庡姟浠ｅ緛鈫掓畫鑱斾娇鐢?| 娈嬬柧浜哄氨涓氬煿璁€佽ˉ璐?|

涓よ€呮槸瀹屽叏涓嶅悓鐨勬椤癸紝涓嶅彲娣锋穯銆傚嵆浣跨即绾虫畫淇濋噾锛屼篃涓嶈兘鏇夸唬宸ヤ激淇濋櫓銆?
        """)

    # ===== 搴曢儴涓嬭浇 =====
    st.divider()
    st.subheader("馃摜 涓嬭浇娈嬩繚閲戞祴绠楀簳绋?)

    # 鐢熸垚閫氱煡鏂囨湰
    if fund_data["搴旂即娈嬩繚閲?] == 0:
        status_text = "闆剁敵鎶ワ紙鍏嶅緛锛?
        amount_text = "0 鍏?
        notice_text = "璐靛叕鍙哥鍚堝皬寰紒涓氬厤寰佹潯浠讹紙鍦ㄨ亴鑱屽伐鈮?0浜猴級锛屽簲缂撮噾棰濅负0鍏冦€傝鍦ㄨ瀹氭湡闄愬唴瀹屾垚闆剁敵鎶ュ嵆鍙€?
    else:
        status_text = f"闇€缂寸撼 {fund_data['搴旂即娈嬩繚閲?]:,.2f} 鍏?
        amount_text = f"{fund_data['搴旂即娈嬩繚閲?]:,.2f} 鍏?
        notice_text = f"璇蜂簬 {fund_data['鐢虫姤骞村害']} 骞寸敵鎶ユ湡鍐呭畬鎴愮敵鎶ュ強缂存銆?

    fund_report = f"""================================================================
       娈嬬柧浜哄氨涓氫繚闅滈噾娴嬬畻鎶ュ憡
  濉姤鍗曚綅锛氭姹夐噾鑹抽緳绉戞妧鏈夐檺鍏徃
  鐢虫姤骞村害锛歿fund_data['鐢虫姤骞村害']}骞?
================================================================

涓€銆佸熀鏈俊鎭?
  涓婂勾锛坽fund_data['鐢虫姤骞村害']-1}骞达級鍦ㄨ亴鑱屽伐浜烘暟锛歿fund_data['涓婂勾鑱屽伐浜烘暟']} 浜?
  涓婂勾瀹夋帓娈嬬柧浜哄氨涓氫汉鏁帮細        {fund_data['涓婂勾娈嬬柧鑱屽伐浜烘暟']} 浜?
  涓婂勾鑱屽伐骞村钩鍧囧伐璧勶細            {fund_data['涓婂勾鑱屽伐骞村潎宸ヨ祫']:,.2f} 鍏?
  娉曞畾瀹夋帓姣斾緥锛?                 {fund_data['娉曞畾瀹夋帓姣斾緥']}
  搴斿畨鎺掓畫鐤句汉鏁帮細                {fund_data['搴斿畨鎺掍汉鏁?]:.4f} 浜?

浜屻€佷紭鎯犳斂绛栭€傜敤
  {fund_data['璁＄畻璇存槑']}
  鏀跨瓥渚濇嵁锛歿fund_data['鏀跨瓥渚濇嵁']}

涓夈€佽绠楃粨鏋?
  搴旂即娈嬩繚閲戯紙鍏ㄩ锛夛細{fund_data['搴旂即娈嬩繚閲戯紙鍏ㄩ锛?]:,.2f} 鍏?
  鍒嗘。寰佹敹姣斾緥锛?     {fund_data['鍒嗘。寰佹敹姣斾緥']}
  鍑忓厤閲戦锛?         {fund_data['鍑忓厤閲戦']:,.2f} 鍏?
  鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€
  瀹為檯搴旂即娈嬩繚閲戯細    {fund_data['搴旂即娈嬩繚閲?]:,.2f} 鍏?

鍥涖€佺敵鎶ョ粨璁?
  鐘舵€侊細{status_text}
  {notice_text}
  鐢虫姤鎴锛歿fund_data['鐢虫姤鎴']}

浜斻€侀噸瑕佹彁绀?
  1. 鍗充娇鍏嶅緛涔熷繀椤绘寜鏃跺湪鐢靛瓙绋庡姟灞€瀹屾垚闆剁敵鎶ワ紱
  2. 涓婂勾鏁版嵁闇€涓庝笂涓€骞村害涓◣鐢虫姤浜烘暟涓€鑷达紝绋庡姟绯荤粺浼氳嚜鍔ㄦ瘮瀵癸紱
  3. 濡傚畨鎺掓湁娈嬬柧浜猴紝闇€鍦ㄧ敵鎶ュ墠瀹屾垚娈嬭仈瀹℃牳璁ゅ畾锛?
  4. 璇蜂繚瀛樻湰娴嬬畻鎶ュ憡鍜岀敵鎶ュ洖鎵ц嚦灏?骞淬€?

================================================================
  鐢熸垚鏃堕棿锛歿datetime.now().strftime('%Y-%m-%d %H:%M')}
  閲戣壋榫橝I绋庡姟鍔╂墜 路 浠呬緵鍙傝€冿紝浠ョ◣鍔℃満鍏虫渶鏂板叕鍛婁负鍑?
================================================================"""

    dl_f1, dl_f2 = st.columns(2)
    with dl_f1:
        st.download_button(
            label=f"馃摜 涓嬭浇娈嬩繚閲戞祴绠楁姤鍛婏紙TXT锛?,
            data=fund_report,
            file_name=f"娈嬩繚閲戞祴绠楁姤鍛奯{def_year}骞確姝︽眽閲戣壋榫欑鎶€.txt",
            mime="text/plain",
            use_container_width=True,
        )
    with dl_f2:
        pdf_bytes = make_pdf(
            f"娈嬬柧浜哄氨涓氫繚闅滈噾娴嬬畻鎶ュ憡 - {def_year}骞?,
            fund_report.split("\n"),
            ""
        )
        if pdf_bytes:
            st.download_button(
                label=f"馃摜 涓嬭浇娈嬩繚閲戞祴绠楁姤鍛婏紙PDF锛?,
                data=pdf_bytes,
                file_name=f"娈嬩繚閲戞祴绠楁姤鍛奯{def_year}骞確姝︽眽閲戣壋榫欑鎶€.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

# ===============================================
#  Tab8锛氬勾鎶ユ暟鎹鍏?
# ===============================================

with tab1:
    st.header("馃梻锔?骞存姤鏁版嵁瀵煎叆")
    st.caption("鏀寔 Excel / PDF 涓ょ鏍煎紡銆傚鍏ュ悗骞存姤鏁版嵁鑷姩鎷嗗垎涓?4 涓搴︾敵鎶ュ簳绋裤€傜◣鍔″勾鎶ヤ笌鍐呴儴搴曠涓嶄竴鑷存椂锛岄噸鏂板鍏ュ嵆鍙籂鍋忋€?)
    st.success("鉁?v1.6.7 鈥?骞存姤瀵煎叆妯″潡宸插氨缁紙2026-06-03 build锛?)

    # 鈹€鈹€ 妫€鏌ユ槸鍚︽湁鍘嗗彶瀵煎叆 鈹€鈹€
    snapshot_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "骞存姤瀵煎叆蹇収.json")
    prev_snapshot = None
    if os.path.exists(snapshot_file):
        try:
            with open(snapshot_file, "r", encoding="utf-8") as f:
                prev_snapshot = json.load(f)
        except Exception:
            prev_snapshot = None

    if prev_snapshot:
        prev_summary = prev_snapshot.get("summary", {})
        st.info(f"""
        馃搶 **宸叉湁瀵煎叆璁板綍**锛坽prev_snapshot.get('imported_at', '鏈煡鏃堕棿')}锛?
        鈥?钀ヤ笟鏀跺叆 {prev_summary.get('annual_revenue', 0):,.0f} 鍏?|
        鍒╂鼎 {prev_summary.get('annual_profit', 0):,.0f} 鍏?|
        {prev_snapshot.get('employee_count', 0)} 鍚嶅憳宸?
        鈥?涓嬫柟涓婁紶鏂版暟鎹皢杩涘叆銆岀籂鍋忔ā寮忋€嶏紝鍙姣斿樊寮傚悗瑕嗙洊鏇存柊銆?
        """)

    # 鈹€鈹€ Step 1锛氫笅杞芥ā鏉?鈹€鈹€
    st.subheader("馃摜 绗竴姝ワ細涓嬭浇瀵煎叆妯℃澘")
    st.markdown("""
    妯℃澘鍖呭惈 3 涓?Sheet锛?
    - **骞存姤姹囨€?* 鈥?鍏ㄥ勾鏀跺叆/鎴愭湰/鍒╂鼎/娉ㄥ唽璧勬湰锛堥儴鍒嗗疄缂达級/娈嬩繚閲戝弬鏁?
    - **鍛樺伐淇℃伅** 鈥?姣忎綅鍛樺伐鐨勬湀鍧囧伐璧勫拰涓撻」闄勫姞鎵ｉ櫎
    - **瀛ｅ害鍒嗘憡鏄庣粏** 鈥?鍙€夛紝濡傞渶鎸夊搴︿笉鍧囧垎鍒欏～鍐?
    """)

    template_bytes = gen_annual_report_template_bytes()
    pdf_template_bytes = gen_annual_report_template_pdf_bytes()

    dl_t1, dl_t2 = st.columns(2)
    with dl_t1:
        st.download_button(
            label="馃摜 涓嬭浇妯℃澘锛圗xcel 路 鍙～鍐欎笂浼狅級",
            data=template_bytes,
            file_name="骞存姤鏁版嵁瀵煎叆妯℃澘.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with dl_t2:
        if pdf_template_bytes:
            st.download_button(
                label="馃摜 涓嬭浇妯℃澘锛圥DF 路 鎵撳嵃瀛樻。锛?,
                data=pdf_template_bytes,
                file_name="骞存姤鏁版嵁瀵煎叆妯℃澘.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        else:
            st.download_button(
                label="馃摜 PDF 涓嶅彲鐢紙闇€ fpdf2锛?,
                data=b"",
                file_name="",
                mime="text/plain",
                use_container_width=True,
                disabled=True,
            )

    st.divider()

    # 鈹€鈹€ Step 2锛氫笂浼?Excel 鎴?PDF 鈹€鈹€
    st.subheader("馃摛 绗簩姝ワ細涓婁紶骞存姤鏂囦欢锛圗xcel 鎴?PDF锛?)

    up_col1, up_col2 = st.columns(2)
    with up_col1:
        uploaded_excel = st.file_uploader(
            "涓婁紶 Excel 骞存姤",
            type=["xlsx", "xls"],
            key="annual_report_excel_uploader",
        )
    with up_col2:
        uploaded_pdf = st.file_uploader(
            "涓婁紶 PDF 骞存姤锛堢◣鍔¤繑鍥炵増锛?,
            type=["pdf"],
            key="annual_report_pdf_uploader",
        )

    uploaded = uploaded_excel or uploaded_pdf
    is_pdf = uploaded_pdf is not None

    if uploaded is not None:
        fname = uploaded.name
        fsize = uploaded.size
        cache_key = f"parsed_annual_{fname}_{fsize}"
        if cache_key not in st.session_state or st.button("馃攧 閲嶆柊瑙ｆ瀽", key="reparse_annual"):
            with st.spinner(f"姝ｅ湪{'AI ' if is_pdf else ''}瑙ｆ瀽骞存姤鏁版嵁锛坽fname}锛?.."):
                file_bytes = uploaded.getvalue()
                if is_pdf:
                    st.session_state[cache_key] = parse_annual_report_pdf(file_bytes)
                else:
                    st.session_state[cache_key] = parse_annual_report_excel(file_bytes)
            st.rerun()

        parsed = st.session_state[cache_key]
        summary = parsed["summary"]
        employees = parsed["employees"]
        quarterly = parsed["quarterly"]
        warnings = parsed["warnings"]

        if is_pdf:
            st.caption("馃 宸查€氳繃 AI 浠?PDF 涓彁鍙栨暟鎹紝璇蜂粩缁嗘牳瀵瑰悇瀛楁鏄惁姝ｇ‘")

        # 鏄剧ず璀﹀憡
        if warnings:
            with st.expander(f"馃攳 鏁版嵁鏍￠獙锛坽len(warnings)} 鏉℃彁绀猴級", expanded=True):
                for w in warnings:
                    st.warning(w)

        # 鈹€鈹€ 绾犲亸瀵规瘮锛堟湁鍘嗗彶鏁版嵁鏃讹級 鈹€鈹€
        if prev_snapshot:
            st.subheader("馃攳 绾犲亸瀵规瘮锛氭柊鏁版嵁 vs 鍘嗗彶瀵煎叆")
            prev_s = prev_snapshot.get("summary", {})
            diff_fields = [
                ("annual_revenue", "鍏ㄥ勾钀ヤ笟鏀跺叆锛堝惈绋庯級", "鍏?),
                ("annual_cost", "鍏ㄥ勾钀ヤ笟鎴愭湰", "鍏?),
                ("annual_profit", "鍏ㄥ勾鍒╂鼎鎬婚", "鍏?),
                ("annual_vat_revenue", "澧炲€肩◣璁＄◣鏀跺叆锛堜笉鍚◣锛?, "鍏?),
                ("avg_employees", "骞冲潎浠庝笟浜烘暟", "浜?),
                ("avg_assets", "骞冲潎璧勪骇鎬婚", "涓囧厓"),
                ("reg_capital", "娉ㄥ唽璧勬湰瀹炵即棰?, "鍏?),
                ("capital_increase", "鏈勾澧炶祫棰?, "鍏?),
                ("total_salary", "鍏ㄥ勾宸ヨ祫鎬婚", "鍏?),
                ("prev_employees", "涓婂勾鑱屽伐浜烘暟", "浜?),
            ]
            diff_rows = []
            has_diff = False
            for key, label, unit in diff_fields:
                old_val = prev_s.get(key, 0) or 0
                new_val = summary.get(key, 0) or 0
                try:
                    old_val = float(old_val)
                    new_val = float(new_val)
                except (ValueError, TypeError):
                    old_val = 0
                    new_val = 0
                delta = new_val - old_val
                if abs(delta) > 0.01:
                    has_diff = True
                    diff_rows.append({
                        "椤圭洰": label,
                        "鍘嗗彶鍊?: f"{old_val:,.2f} {unit}",
                        "鏂板€?: f"{new_val:,.2f} {unit}",
                        "宸紓": f"{delta:+,.2f} {unit}",
                    })
            if has_diff:
                st.dataframe(pd.DataFrame(diff_rows), use_container_width=True, hide_index=True)
                st.warning("鈿狅笍 绋庡姟骞存姤涓庡唴閮ㄥ簳绋垮瓨鍦ㄥ樊寮傦紝纭鍚庡皢鐢ㄦ柊鏁版嵁瑕嗙洊鏃ф暟鎹紙鍚搴︾敵鎶ユ暟鎹級")
            else:
                st.success("鉁?鏂版暟鎹笌鍘嗗彶鏁版嵁涓€鑷达紝鏃犻渶绾犲亸")

        # 鈹€鈹€ 瑙ｆ瀽缁撴灉棰勮 鈹€鈹€
        st.subheader("馃搵 绗笁姝ワ細纭瑙ｆ瀽缁撴灉")

        prev_col1, prev_col2, prev_col3 = st.columns(3)

        with prev_col1:
            st.markdown("**馃搳 缁忚惀鏁版嵁**")
            st.metric("鍏ㄥ勾钀ヤ笟鏀跺叆", f"{summary.get('annual_revenue', 0):,.0f} 鍏?)
            st.metric("鍏ㄥ勾钀ヤ笟鎴愭湰", f"{summary.get('annual_cost', 0):,.0f} 鍏?)
            st.metric("鍏ㄥ勾鍒╂鼎鎬婚", f"{summary.get('annual_profit', 0):,.0f} 鍏?)
            st.metric("澧炲€肩◣璁＄◣鏀跺叆", f"{summary.get('annual_vat_revenue', 0):,.0f} 鍏?)

        with prev_col2:
            st.markdown("**馃彚 浼佷笟淇℃伅**")
            st.metric("骞冲潎浠庝笟浜烘暟", f"{int(summary.get('avg_employees', 0))} 浜?)
            st.metric("骞冲潎璧勪骇鎬婚", f"{summary.get('avg_assets', 0):,.1f} 涓囧厓")
            st.metric("娉ㄥ唽璧勬湰瀹炵即", f"{summary.get('reg_capital', 0):,.0f} 鍏?)
            st.metric("鏈勾澧炶祫棰?, f"{summary.get('capital_increase', 0):,.0f} 鍏?)
            q_split = summary.get("split_method", "骞冲潎")
            st.caption(f"瀛ｅ害鍒嗘憡鏂瑰紡锛?*{q_split}**")

        with prev_col3:
            st.markdown("**馃懃 钖叕 & 娈嬩繚閲?*")
            st.metric("鍏ㄥ勾宸ヨ祫鎬婚", f"{summary.get('total_salary', 0):,.0f} 鍏?)
            st.metric("绀句繚鍏徃鎵挎媴", f"{summary.get('total_si_company', 0):,.0f} 鍏?)
            st.metric("涓婂勾鑱屽伐浜烘暟", f"{int(summary.get('prev_employees', 0))} 浜?)
            st.metric("涓婂勾瀹夋帓娈嬬柧浜?, f"{int(summary.get('prev_disabled', 0))} 浜?)

        # 鍛樺伐棰勮
        if employees:
            st.markdown(f"**馃懃 鍛樺伐淇℃伅锛坽len(employees)} 浜猴級**")
            emp_df = pd.DataFrame(employees)
            emp_df_display = emp_df.rename(columns={
                "name": "濮撳悕", "gross_salary": "绋庡墠鏈堝伐璧?,
                "si_base": "绀句繚鍩烘暟", "si_personal_actual": "涓汉绀句繚",
                "special_deductions": "涓撻」鎵ｉ櫎", "child_education": "瀛愬コ鏁欒偛",
                "infant_care": "濠村辜鍎跨収鎶?, "elderly_care": "璧″吇鑰佷汉",
            })
            st.dataframe(emp_df_display, use_container_width=True, hide_index=True)

        # 瀛ｅ害鏄庣粏棰勮
        if quarterly:
            st.markdown("**馃搮 瀛ｅ害鍒嗘憡鏄庣粏**")
            q_rows = []
            q_total_rev = q_total_cost = q_total_profit = q_total_vat = 0
            for q in ["Q1", "Q2", "Q3", "Q4"]:
                qd = quarterly.get(q, {})
                r = qd.get("revenue", 0)
                c = qd.get("cost", 0)
                p = qd.get("period_profit", 0)
                v = qd.get("vat_revenue", 0)
                q_total_rev += r
                q_total_cost += c
                q_total_profit += p
                q_total_vat += v
                q_rows.append({
                    "瀛ｅ害": q, "钀ヤ笟鏀跺叆": r, "钀ヤ笟鎴愭湰": c, "鍒╂鼎": p, "澧炲€肩◣鏀跺叆": v,
                })
            # 鍚堣琛?
            q_rows.append({
                "瀛ｅ害": "鍚堣", "钀ヤ笟鏀跺叆": q_total_rev, "钀ヤ笟鎴愭湰": q_total_cost,
                "鍒╂鼎": q_total_profit, "澧炲€肩◣鏀跺叆": q_total_vat,
            })
            st.dataframe(pd.DataFrame(q_rows), use_container_width=True, hide_index=True)

            # 鏍￠獙锛氬洓瀛ｅ害鍚堣 = 骞存姤鎬绘暟
            annual_rev = summary.get("annual_revenue", 0)
            annual_cost = summary.get("annual_cost", 0)
            annual_profit = summary.get("annual_profit", 0)
            annual_vat = summary.get("annual_vat_revenue", 0)

            rev_ok = abs(q_total_rev - annual_rev) < 1
            cost_ok = abs(q_total_cost - annual_cost) < 1
            profit_ok = abs(q_total_profit - annual_profit) < 1
            vat_ok = abs(q_total_vat - annual_vat) < 1

            if not (rev_ok and cost_ok and profit_ok and vat_ok):
                issues = []
                if not rev_ok:
                    issues.append(f"钀ヤ笟鏀跺叆锛氬洓瀛ｅ害鍚堣 {q_total_rev:,.0f} 鈮?骞存姤 {annual_rev:,.0f}")
                if not cost_ok:
                    issues.append(f"钀ヤ笟鎴愭湰锛氬洓瀛ｅ害鍚堣 {q_total_cost:,.0f} 鈮?骞存姤 {annual_cost:,.0f}")
                if not profit_ok:
                    issues.append(f"鍒╂鼎鎬婚锛氬洓瀛ｅ害鍚堣 {q_total_profit:,.0f} 鈮?骞存姤 {annual_profit:,.0f}")
                if not vat_ok:
                    issues.append(f"澧炲€肩◣鏀跺叆锛氬洓瀛ｅ害鍚堣 {q_total_vat:,.0f} 鈮?骞存姤 {annual_vat:,.0f}")
                st.warning("鈿狅笍 骞存姤姹囨€?鈮?鍥涘搴﹀悎璁★紝璇锋牳瀵癸細\n" + "\n".join(f"  路 {i}" for i in issues))
            else:
                st.success("鉁?骞存姤姹囨€?= 鍥涘搴﹀悎璁★紝鏁版嵁涓€鑷?)
        else:
            # 鏄剧ず鑷姩鍧囧垎棰勮
            rev = summary.get("annual_revenue", 0)
            cost = summary.get("annual_cost", 0)
            profit = summary.get("annual_profit", 0)
            vat_rev = summary.get("annual_vat_revenue", 0)
            st.caption("馃挕 鏈～鍐欏搴︽槑缁嗭紝灏嗘寜 4 瀛ｅ害骞冲潎鍒嗘憡锛?)
            q_avg_rows = []
            for q in ["Q1", "Q2", "Q3", "Q4"]:
                q_avg_rows.append({
                    "瀛ｅ害": q, "钀ヤ笟鏀跺叆": round(rev / 4, 2),
                    "钀ヤ笟鎴愭湰": round(cost / 4, 2), "鍒╂鼎": round(profit / 4, 2),
                    "澧炲€肩◣鏀跺叆": round(vat_rev / 4, 2),
                })
            q_avg_rows.append({
                "瀛ｅ害": "鍚堣", "钀ヤ笟鏀跺叆": rev, "钀ヤ笟鎴愭湰": cost,
                "鍒╂鼎": profit, "澧炲€肩◣鏀跺叆": vat_rev,
            })
            st.dataframe(pd.DataFrame(q_avg_rows), use_container_width=True, hide_index=True)

        # 鈹€鈹€ 宸ヨ祫鏁版嵁鏍￠獙锛堜笁閲嶏級 鈹€鈹€
        if employees:
            with st.expander("馃攷 宸ヨ祫鏁版嵁鏍￠獙锛堥摱琛屾祦姘?vs 涓◣鐢虫姤 vs 骞存姤锛?, expanded=False):
                st.caption("涓婁紶閾惰娴佹按鍜?鎴栦釜绋庣敵鎶ヨ褰曪紝涓庡鍏ョ殑鍛樺伐宸ヨ祫浜ゅ弶姣斿")

                vac1, vac2 = st.columns(2)
                with vac1:
                    val_bank_file = st.file_uploader(
                        "涓婁紶閾惰娴佹按锛堟牎楠屽伐璧勬敮鍑猴級",
                        type=["csv", "xlsx", "xls"],
                        key="annual_val_bank",
                    )
                with vac2:
                    val_tax_file = st.file_uploader(
                        "涓婁紶涓◣鐢虫姤璁板綍锛堟牎楠岀疮璁℃敹鍏ワ級",
                        type=["csv", "xlsx", "xls"],
                        key="annual_val_tax",
                    )

                val_annual_salary = st.number_input(
                    "骞存姤銆屽叏骞村伐璧勬€婚銆嶏紙鍏冿紝閫夊～锛?,
                    min_value=0.0, value=float(summary.get("total_salary", 0) or 0.0), step=1000.0,
                    key="annual_val_salary_input",
                )

                if st.button("馃攳 寮€濮嬫牎楠?, key="run_annual_salary_val", use_container_width=True):
                    bank_df = None
                    tax_df = None

                    if val_bank_file:
                        try:
                            if val_bank_file.name.endswith(".csv"):
                                try:
                                    bank_df = pd.read_csv(val_bank_file, encoding="utf-8-sig")
                                except Exception:
                                    val_bank_file.seek(0)
                                    bank_df = pd.read_csv(val_bank_file, encoding="gbk")
                            else:
                                bank_df = pd.read_excel(val_bank_file)
                        except Exception as e:
                            st.error(f"閾惰娴佹按璇诲彇澶辫触锛歿e}")

                    if val_tax_file:
                        try:
                            if val_tax_file.name.endswith(".csv"):
                                try:
                                    tax_df = pd.read_csv(val_tax_file, encoding="utf-8-sig")
                                except Exception:
                                    val_tax_file.seek(0)
                                    tax_df = pd.read_csv(val_tax_file, encoding="gbk")
                            else:
                                tax_df = pd.read_excel(val_tax_file)
                        except Exception as e:
                            st.error(f"涓◣鐢虫姤璁板綍璇诲彇澶辫触锛歿e}")

                    with st.spinner("姝ｅ湪鏍￠獙..."):
                        val_result = validate_salary_data(
                            employees=employees,
                            bank_df=bank_df,
                            tax_filing_df=tax_df,
                            annual_total_salary=val_annual_salary if val_annual_salary > 0 else 0.0,
                        )

                    st.divider()

                    # 鏍￠獙1锛氶摱琛屾祦姘?
                    if bank_df is not None:
                        st.markdown("**馃彟 鏍￠獙涓€锛氶摱琛屾祦姘?vs 绯荤粺宸ヨ祫**")
                        bm = val_result.get("bank_match")
                        if isinstance(bm, dict) and bm:
                            c1, c2, c3 = st.columns(3)
                            c1.metric("閾惰娴佹按宸ヨ祫鏀嚭", f"{bm['bank_salary_total']:,.0f} 鍏?)
                            c2.metric("绯荤粺骞村伐璧勫悎璁?, f"{bm['sys_annual_total']:,.0f} 鍏?)
                            c3.metric("宸紓", f"{bm['diff']:+,.0f} 鍏?, delta=f"{bm['diff_pct']:+.1f}%")
                            if bm["match"]:
                                st.success("鉁?閾惰娴佹按涓庣郴缁熷伐璧勪竴鑷?)
                            else:
                                st.error(f"鈿狅笍 宸紓杈冨ぇ锛佸缓璁牳鏌ワ紙璇嗗埆鍒?{bm['txn_count']} 鏉″伐璧勭被浜ゆ槗锛?)
                        else:
                            st.warning("鏈湪閾惰娴佹按涓瘑鍒埌宸ヨ祫绫绘敮鍑猴紙鎽樿闇€鍚€屽伐璧勩€嶃€屽閲戙€嶃€岀哗鏁堛€嶇瓑鍏抽敭璇嶏級")
                        st.divider()

                    # 鏍￠獙2锛氫釜绋庣敵鎶?
                    if tax_df is not None:
                        st.markdown("**馃搵 鏍￠獙浜岋細涓◣鐢虫姤璁板綍 vs 绯荤粺宸ヨ祫**")
                        tm = val_result.get("tax_match", [])
                        if tm:
                            tm_rows = []
                            for r in tm:
                                tm_rows.append({
                                    "濮撳悕": r["name"],
                                    "涓◣鐢虫姤绱鏀跺叆": f"{r['tax_filing_income']:,.0f}",
                                    "绯荤粺骞村伐璧?: f"{r['sys_annual']:,.0f}",
                                    "宸紓": f"{r['diff']:+,.0f}",
                                    "鐘舵€?: "鉁? if r["match"] else "鈿狅笍",
                                })
                            st.dataframe(pd.DataFrame(tm_rows), use_container_width=True, hide_index=True)
                        else:
                            st.warning("涓◣鐢虫姤璁板綍涓湭鎵惧埌涓庣郴缁熷憳宸ュ尮閰嶇殑濮撳悕")
                        st.divider()

                    # 鏍￠獙3锛氬勾鎶ュ伐璧勬€婚
                    if val_annual_salary > 0:
                        st.markdown("**馃搳 鏍￠獙涓夛細骞存姤宸ヨ祫鎬婚 vs 绯荤粺骞村伐璧勫悎璁?*")
                        am = val_result.get("annual_match", {})
                        if am:
                            c1, c2, c3 = st.columns(3)
                            c1.metric("骞存姤宸ヨ祫鎬婚", f"{am['annual_total_salary']:,.0f} 鍏?)
                            c2.metric("绯荤粺骞村伐璧勫悎璁?, f"{am['sys_annual_total']:,.0f} 鍏?)
                            c3.metric("宸紓", f"{am['diff']:+,.0f} 鍏?, delta=f"{am['diff_pct']:+.1f}%")
                            if am["match"]:
                                st.success("鉁?骞存姤宸ヨ祫鎬婚涓庣郴缁熷伐璧勪竴鑷?)
                            else:
                                st.error("鈿狅笍 骞存姤宸ヨ祫鎬婚涓庣郴缁熷勾宸ヨ祫鍚堣宸紓杈冨ぇ锛?)
                        st.divider()

                    # 姹囨€?
                    warnings = val_result.get("warnings", [])
                    if warnings:
                        with st.expander(f"馃搵 鏍￠獙璇存槑锛坽len(warnings)} 鏉★級", expanded=True):
                            for w in warnings:
                                if "鉁? in w:
                                    st.success(w)
                                elif "鈿狅笍" in w or "宸? in w:
                                    st.warning(w)
                                else:
                                    st.info(w)

        st.divider()

        # 鈹€鈹€ Step 4锛氱‘璁ゅ鍏?鈹€鈹€
        if prev_snapshot:
            st.subheader("馃殌 绗洓姝ワ細纭瀵煎叆锛堢籂鍋忔ā寮忥級")
            st.markdown("""
            鐐瑰嚮涓嬫柟鎸夐挳灏嗙敤鏂版暟鎹?*瑕嗙洊**浠ヤ笅鍐呭锛?
            - 鉁?鍛樺伐淇℃伅 鈫?鏇存柊銆屽伐璧勮绠椼€嶈崏绋?
            - 鉁?4 涓搴︽暟鎹?鈫?**瑕嗙洊**銆屽搴︾敵鎶ャ€嶅瓨妗?
            - 鉁?鍗拌姳绋庡弬鏁?鈫?鏇存柊娉ㄥ唽璧勬湰/澧炶祫
            - 鉁?娈嬩繚閲戝弬鏁?鈫?鏇存柊涓婂勾鑱屽伐浜烘暟/宸ヨ祫绛?
            """)
            import_label = "鉁?纭瀵煎叆锛堣鐩栨棫鏁版嵁锛?
        else:
            st.subheader("馃殌 绗洓姝ワ細纭瀵煎叆")
            st.markdown("""
            鐐瑰嚮涓嬫柟鎸夐挳灏嗕竴娆℃€у畬鎴愪互涓嬫搷浣滐細
            - 鉁?灏嗗憳宸ヤ俊鎭繚瀛樹负銆屽伐璧勮绠椼€嶈崏绋?
            - 鉁?灏?4 涓搴︽暟鎹啓鍏ャ€屽搴︾敵鎶ャ€嶅瓨妗?
            - 鉁?璁剧疆渚ц竟鏍忕殑鍗拌姳绋庯紙娉ㄥ唽璧勬湰/澧炶祫锛夊拰娈嬩繚閲戝弬鏁?
            """)
            import_label = "鉁?纭瀵煎叆 2025 骞村勾鎶ユ暟鎹?

        import_warning = st.checkbox(
            "鈿狅笍 鎴戠‘璁や互涓婃暟鎹纭紝瀵煎叆灏嗚鐩栫幇鏈夎崏绋垮拰瀛ｅ害鐢虫姤鏁版嵁",
            key="confirm_import",
        )

        if st.button(import_label, type="primary", use_container_width=True,
                     disabled=not import_warning):
            import_count = 0

            # 1. 淇濆瓨鍛樺伐鑽夌
            if employees:
                EMP_DATA_FILE_ANNUAL = os.path.join(os.path.dirname(os.path.abspath(__file__)), "鍛樺伐鏁版嵁_鑽夌.json")
                with open(EMP_DATA_FILE_ANNUAL, "w", encoding="utf-8") as f:
                    json.dump(employees, f, ensure_ascii=False, indent=2)
                st.session_state["employees_saved"] = employees
                import_count += 1
                st.success(f"鉁?宸插鍏?{len(employees)} 鍚嶅憳宸ユ暟鎹?)

            # 2. 鍐欏叆瀛ｅ害鐢虫姤鏁版嵁锛?025骞达級
            rev = summary.get("annual_revenue", 0)
            cost = summary.get("annual_cost", 0)
            profit = summary.get("annual_profit", 0)
            vat_rev = summary.get("annual_vat_revenue", 0)
            avg_emp = int(summary.get("avg_employees", 0))
            avg_assets = summary.get("avg_assets", 0.0)

            if quarterly:
                q_data_map = quarterly
            else:
                q_data_map = {}
                for q in ["Q1", "Q2", "Q3", "Q4"]:
                    q_data_map[q] = {
                        "revenue": rev / 4,
                        "cost": cost / 4,
                        "period_profit": profit / 4,
                        "vat_revenue": vat_rev / 4,
                        "avg_employees": avg_emp,
                        "avg_assets": avg_assets,
                    }

            quarter_imported = 0
            for q_name, q_data in q_data_map.items():
                q_num = int(q_name.replace("Q", ""))
                save_quarter_data(2025, q_num, {
                    "revenue": q_data.get("revenue", 0),
                    "cost": q_data.get("cost", 0),
                    "period_profit": q_data.get("period_profit", 0),
                    "vat_revenue": q_data.get("vat_revenue", 0),
                    "avg_employees": q_data.get("avg_employees", avg_emp),
                    "avg_assets": q_data.get("avg_assets", avg_assets),
                })
                quarter_imported += 1
            import_count += 1
            st.success(f"鉁?宸插啓鍏?2025 骞?{quarter_imported} 涓搴︾敵鎶ユ暟鎹?)

            # 3. 璁剧疆渚ц竟鏍忓弬鏁?
            st.session_state["stamp_reg_capital"] = float(summary.get("reg_capital", 0))
            st.session_state["stamp_capital_increase"] = float(summary.get("capital_increase", 0))
            st.session_state["def_prev_employees"] = int(summary.get("prev_employees", 0))
            st.session_state["def_prev_disabled"] = int(summary.get("prev_disabled", 0))
            st.session_state["def_prev_avg_salary"] = float(summary.get("prev_avg_salary", 60000))
            st.session_state["def_local_avg_salary"] = float(summary.get("local_avg_salary", 90000))

            # 4. 淇濆瓨骞存姤蹇収
            snapshot = {
                "imported_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "year": 2025,
                "source": "PDF" if is_pdf else "Excel",
                "summary": {k: (float(v) if isinstance(v, (int, float)) else v) for k, v in summary.items()},
                "employee_count": len(employees),
                "quarter_count": quarter_imported,
            }
            with open(snapshot_file, "w", encoding="utf-8") as f:
                json.dump(snapshot, f, ensure_ascii=False, indent=2)

            if prev_snapshot:
                prev_s = prev_snapshot.get("summary", {})
                old_rev = prev_s.get("annual_revenue", 0) or 0
                new_rev = summary.get("annual_revenue", 0) or 0
                delta_rev = new_rev - old_rev
                st.balloons()
                st.success(f"""
                馃帀 **骞存姤绾犲亸瀹屾垚锛?*

                | 椤圭洰 | 鏃у€?| 鏂板€?| 鍙樺寲 |
                |------|------|------|------|
                | 钀ヤ笟鏀跺叆 | {old_rev:,.0f} 鍏?| {new_rev:,.0f} 鍏?| {delta_rev:+,.0f} 鍏?|
                | 鍛樺伐鏁版嵁 | {prev_snapshot.get('employee_count', 0)} 浜?| {len(employees)} 浜?| 鈥?|
                | 瀛ｅ害鐢虫姤 | 宸茶鐩?| 2025 Q1-Q4 | 鈥?|
                | 鍗拌姳绋?娈嬩繚閲?| 宸叉洿鏂?| 鈥?| 鈥?|

                馃憠 鍒囨崲鍒般€岎煋?瀛ｅ害鐢虫姤銆嶉€夋嫨 2026 骞村嵆鍙紑濮嬫湰骞寸敵鎶ャ€?
                """)
            else:
                st.balloons()
                st.success(f"""
                馃帀 **2025 骞村勾鎶ユ暟鎹鍏ュ畬鎴愶紒**

                | 椤圭洰 | 鐘舵€?|
                |------|------|
                | 鍛樺伐鏁版嵁 | {len(employees)} 浜哄凡淇濆瓨 |
                | 瀛ｅ害鐢虫姤 | 2025 骞?Q1-Q4 宸插啓鍏?|
                | 鍗拌姳绋庡弬鏁?| 娉ㄥ唽璧勬湰 {summary.get('reg_capital', 0):,.0f} 鍏?|
                | 娈嬩繚閲戝弬鏁?| 涓婂勾 {int(summary.get('prev_employees', 0))} 浜?|

                馃憠 鐜板湪鍒囨崲鍒般€岎煋?瀛ｅ害鐢虫姤銆嶅苟閫夋嫨 2026 骞村嵆鍙紑濮嬫湰骞寸敵鎶ャ€?
                """)

# ===============================================
#  鍏ㄥ眬椤佃剼
# ===============================================
st.divider()
footer_col1, footer_col2, footer_col3 = st.columns([1, 2, 1])
with footer_col2:
    st.caption(
        f"漏 {datetime.now().year} 閲戣壋榫橝I绋庡姟鍔╂墜 v1.6 路 "
        "姝︽眽閲戣壋榫欑鎶€鏈夐檺鍏徃 路 "
        "鍩轰簬灏忚妯＄撼绋庝汉+灏忓瀷寰埄浼佷笟鍦烘櫙 路 "
        "浠呬緵鍙傝€冿紝浠ョ◣鍔℃満鍏虫渶鏂板叕鍛婁负鍑?
    )
